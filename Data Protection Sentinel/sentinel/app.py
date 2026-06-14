"""
Data Protection Sentinel — Flask application.
Unified platform: RoPA + DPIA + Breaches + DSR + Vendors +
Privacy Notices + Consent + Controllers + Audit Log + AI + Users.
"""
import os
import json
from datetime import datetime
from functools import wraps
from flask import (Flask, render_template, request, jsonify,
                   send_file, abort, session, redirect, url_for)
from dotenv import load_dotenv

load_dotenv()

import database as db
import ai_service as ai
from docx_export import generate_docx
try:
    from xlsx_export import generate_ropa_xlsx
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "sentinel-dev-secret-change-me")

with app.app_context():
    db.init_db()
    db.ensure_default_admin()


# ── Auth helpers ──────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        if session.get("user_role") not in ("admin",):
            return jsonify({"error": "Forbidden — admin only"}), 403
        return f(*args, **kwargs)
    return decorated


def current_user():
    return {
        "id": session.get("user_id"),
        "username": session.get("username"),
        "full_name": session.get("full_name"),
        "role": session.get("user_role"),
        "avatar_initials": session.get("avatar_initials"),
        "email": session.get("user_email"),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# AUTH ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/login", methods=["GET"])
def login_page():
    if "user_id" in session:
        return redirect(url_for("index"))
    return render_template("login.html")


@app.route("/api/auth/login", methods=["POST"])
def api_login():
    body = request.get_json(force=True, silent=True) or {}
    username = body.get("username", "").strip()
    password = body.get("password", "")
    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400
    user = db.authenticate_user(username, password)
    if not user:
        return jsonify({"error": "Invalid credentials"}), 401
    session.permanent = True
    session["user_id"]         = user["id"]
    session["username"]        = user["username"]
    session["full_name"]       = user["full_name"]
    session["user_role"]       = user["role"]
    session["avatar_initials"] = user["avatar_initials"]
    session["user_email"]      = user["email"]
    return jsonify({
        "ok": True,
        "user": {
            "id": user["id"], "username": user["username"],
            "full_name": user["full_name"], "role": user["role"],
            "avatar_initials": user["avatar_initials"],
            "email": user["email"],
        }
    })


@app.route("/api/auth/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/auth/me", methods=["GET"])
@login_required
def api_me():
    return jsonify(current_user())


@app.route("/api/auth/change-password", methods=["POST"])
@login_required
def api_change_password():
    body = request.get_json(force=True, silent=True) or {}
    old_pw = body.get("old_password", "")
    new_pw = body.get("new_password", "")
    if not new_pw or len(new_pw) < 8:
        return jsonify({"error": "New password must be at least 8 characters"}), 400
    user = db.authenticate_user(session["username"], old_pw)
    if not user:
        return jsonify({"error": "Current password is incorrect"}), 401
    db.update_user(session["user_id"], {"password": new_pw})
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# USER MANAGEMENT (admin only)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/users", methods=["GET"])
@login_required
def api_user_list():
    users = db.list_users()
    # Don't send password hashes
    for u in users:
        u.pop("password_hash", None)
    return jsonify(users)


@app.route("/api/users", methods=["POST"])
@admin_required
def api_user_create():
    data = request.get_json(force=True, silent=True) or {}
    if not data.get("username") or not data.get("email"):
        return jsonify({"error": "username and email required"}), 400
    try:
        new_id = db.create_user(data)
        user = db.get_user(new_id)
        user.pop("password_hash", None)
        return jsonify(user), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/users/<int:user_id>", methods=["PUT"])
@login_required
def api_user_update(user_id):
    # Users can update their own profile; only admins can change others
    if session.get("user_role") != "admin" and session.get("user_id") != user_id:
        return jsonify({"error": "Forbidden"}), 403
    data = request.get_json(force=True, silent=True) or {}
    # Non-admins can't change roles
    if session.get("user_role") != "admin":
        data.pop("role", None)
        data.pop("is_active", None)
    db.update_user(user_id, data)
    user = db.get_user(user_id)
    user.pop("password_hash", None)
    return jsonify(user)


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@admin_required
def api_user_delete(user_id):
    if user_id == session.get("user_id"):
        return jsonify({"error": "Cannot delete your own account"}), 400
    db.delete_user(user_id)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN SPA
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/")
@login_required
def index():
    return render_template("index.html",
                           regulations=ai.REGULATIONS,
                           activity_types=ai.ACTIVITY_TYPES,
                           data_categories=ai.DATA_CATEGORIES,
                           special_categories=ai.SPECIAL_CATEGORIES,
                           current_user=current_user())


# ═══════════════════════════════════════════════════════════════════════════════
# STATS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/stats")
@login_required
def api_stats():
    return jsonify(db.get_stats())


# ═══════════════════════════════════════════════════════════════════════════════
# ROPA
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/ropa", methods=["GET"])
def api_ropa_list():
    entries = db.list_ropa(
        search=request.args.get("q"),
        regulation=request.args.get("regulation"),
        status=request.args.get("status"),
        risk=request.args.get("risk"),
    )
    return jsonify(entries)


@app.route("/api/ropa", methods=["POST"])
def api_ropa_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_ropa(data)
    entry = db.get_ropa(new_id)
    return jsonify(entry), 201


@app.route("/api/ropa/<int:ropa_id>", methods=["GET"])
def api_ropa_get(ropa_id):
    entry = db.get_ropa(ropa_id)
    if not entry:
        abort(404)
    return jsonify(entry)


@app.route("/api/ropa/<int:ropa_id>", methods=["PUT"])
def api_ropa_update(ropa_id):
    data = request.get_json(force=True, silent=True) or {}
    db.update_ropa(ropa_id, data)
    return jsonify(db.get_ropa(ropa_id))


@app.route("/api/ropa/<int:ropa_id>", methods=["DELETE"])
def api_ropa_delete(ropa_id):
    db.delete_ropa(ropa_id)
    return jsonify({"ok": True})


@app.route("/api/ropa/export/xlsx")
@login_required
def api_ropa_export_xlsx():
    if not XLSX_AVAILABLE:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500
    entries = db.list_ropa(
        search=request.args.get("q"),
        regulation=request.args.get("regulation"),
        status=request.args.get("status"),
    )
    settings = db.get_all_settings()
    buf = generate_ropa_xlsx(entries, settings)
    org = settings.get("org_name", "Organisation").replace(" ", "_")
    filename = f"RoPA_{org}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# DPIA
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/dpias", methods=["GET"])
def api_dpia_list():
    dpias = db.list_dpias(
        search=request.args.get("q"),
        regulation=request.args.get("regulation"),
        status=request.args.get("status"),
    )
    return jsonify(dpias)


@app.route("/api/dpias", methods=["POST"])
def api_dpia_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_dpia(data)
    return jsonify(db.get_dpia(new_id)), 201


@app.route("/api/dpias/<int:dpia_id>", methods=["GET"])
def api_dpia_get(dpia_id):
    dpia = db.get_dpia(dpia_id)
    if not dpia:
        abort(404)
    return jsonify(dpia)


@app.route("/api/dpias/<int:dpia_id>", methods=["PUT"])
def api_dpia_update(dpia_id):
    data = request.get_json(force=True, silent=True) or {}
    db.update_dpia(dpia_id, data)
    return jsonify(db.get_dpia(dpia_id))


@app.route("/api/dpias/<int:dpia_id>", methods=["DELETE"])
def api_dpia_delete(dpia_id):
    db.delete_dpia(dpia_id)
    return jsonify({"ok": True})


@app.route("/api/dpias/<int:dpia_id>/download")
def api_dpia_download(dpia_id):
    dpia = db.get_dpia(dpia_id)
    if not dpia:
        abort(404)
    buf = generate_docx(dpia)
    filename = f"DPIA_{dpia['ref_number']}.docx"
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# BREACHES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/breaches", methods=["GET"])
def api_breach_list():
    return jsonify(db.list_breaches(
        search=request.args.get("q"),
        status=request.args.get("status"),
        severity=request.args.get("severity"),
    ))


@app.route("/api/breaches", methods=["POST"])
def api_breach_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_breach(data)
    return jsonify(db.get_breach(new_id)), 201


@app.route("/api/breaches/<int:breach_id>", methods=["GET"])
def api_breach_get(breach_id):
    b = db.get_breach(breach_id)
    if not b:
        abort(404)
    return jsonify(b)


@app.route("/api/breaches/<int:breach_id>", methods=["PUT"])
def api_breach_update(breach_id):
    data = request.get_json(force=True, silent=True) or {}
    db.update_breach(breach_id, data)
    return jsonify(db.get_breach(breach_id))


@app.route("/api/breaches/<int:breach_id>", methods=["DELETE"])
def api_breach_delete(breach_id):
    db.delete_breach(breach_id)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# DSR
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/dsrs", methods=["GET"])
def api_dsr_list():
    return jsonify(db.list_dsrs(
        search=request.args.get("q"),
        status=request.args.get("status"),
        request_type=request.args.get("type"),
    ))


@app.route("/api/dsrs", methods=["POST"])
def api_dsr_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_dsr(data)
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM dsr_requests WHERE id=?", (new_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/dsrs/<int:dsr_id>", methods=["PUT"])
def api_dsr_update(dsr_id):
    data = request.get_json(force=True, silent=True) or {}
    db.update_dsr(dsr_id, data)
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM dsr_requests WHERE id=?", (dsr_id,)).fetchone()
    return jsonify(dict(row))


@app.route("/api/dsrs/<int:dsr_id>", methods=["DELETE"])
def api_dsr_delete(dsr_id):
    db.delete_dsr(dsr_id)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# VENDORS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/vendors", methods=["GET"])
def api_vendor_list():
    return jsonify(db.list_vendors(
        search=request.args.get("q"),
        risk=request.args.get("risk"),
        dpa_status=request.args.get("dpa_status"),
    ))


@app.route("/api/vendors", methods=["POST"])
def api_vendor_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_vendor(data)
    return jsonify(db.get_vendor(new_id)), 201


@app.route("/api/vendors/<int:vendor_id>", methods=["GET"])
def api_vendor_get(vendor_id):
    v = db.get_vendor(vendor_id)
    if not v:
        abort(404)
    return jsonify(v)


@app.route("/api/vendors/<int:vendor_id>", methods=["PUT"])
def api_vendor_update(vendor_id):
    data = request.get_json(force=True, silent=True) or {}
    db.update_vendor(vendor_id, data)
    return jsonify(db.get_vendor(vendor_id))


@app.route("/api/vendors/<int:vendor_id>", methods=["DELETE"])
def api_vendor_delete(vendor_id):
    db.delete_vendor(vendor_id)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# PRIVACY NOTICES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/notices", methods=["GET"])
def api_notice_list():
    return jsonify(db.list_notices())


@app.route("/api/notices", methods=["POST"])
def api_notice_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_notice(data)
    return jsonify(db.get_notice(new_id)), 201


@app.route("/api/notices/<int:notice_id>", methods=["GET"])
def api_notice_get(notice_id):
    n = db.get_notice(notice_id)
    if not n:
        abort(404)
    return jsonify(n)


@app.route("/api/notices/<int:notice_id>", methods=["PUT"])
def api_notice_update(notice_id):
    data = request.get_json(force=True, silent=True) or {}
    db.update_notice(notice_id, data)
    return jsonify(db.get_notice(notice_id))


@app.route("/api/notices/<int:notice_id>", methods=["DELETE"])
def api_notice_delete(notice_id):
    db.delete_notice(notice_id)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# CONSENT RECORDS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/consent", methods=["GET"])
def api_consent_list():
    return jsonify(db.list_consent(
        search=request.args.get("q"),
        status=request.args.get("status"),
    ))


@app.route("/api/consent", methods=["POST"])
def api_consent_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_consent(data)
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM consent_records WHERE id=?", (new_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/consent/<int:consent_id>", methods=["GET"])
@login_required
def api_consent_get(consent_id):
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM consent_records WHERE id=?", (consent_id,)).fetchone()
    if not row:
        abort(404)
    return jsonify(dict(row))


@app.route("/api/consent/<int:consent_id>", methods=["PUT"])
def api_consent_update(consent_id):
    data = request.get_json(force=True, silent=True) or {}
    db.update_consent(consent_id, data)
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM consent_records WHERE id=?", (consent_id,)).fetchone()
    return jsonify(dict(row))


@app.route("/api/consent/<int:consent_id>", methods=["DELETE"])
def api_consent_delete(consent_id):
    db.delete_consent(consent_id)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# CONTROLLERS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/controllers", methods=["GET"])
@login_required
def api_controller_list():
    return jsonify(db.list_controllers())


@app.route("/api/controllers", methods=["POST"])
@login_required
def api_controller_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_controller(data)
    return jsonify(db.get_controller(new_id)), 201


@app.route("/api/controllers/<int:ctrl_id>", methods=["GET"])
@login_required
def api_controller_get(ctrl_id):
    c = db.get_controller(ctrl_id)
    if not c:
        abort(404)
    return jsonify(c)


@app.route("/api/controllers/<int:ctrl_id>", methods=["PUT"])
@login_required
def api_controller_update(ctrl_id):
    data = request.get_json(force=True, silent=True) or {}
    db.update_controller(ctrl_id, data)
    return jsonify(db.get_controller(ctrl_id))


@app.route("/api/controllers/<int:ctrl_id>", methods=["DELETE"])
@login_required
def api_controller_delete(ctrl_id):
    db.delete_controller(ctrl_id)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# TRANSFERS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/transfers", methods=["GET"])
def api_transfer_list():
    return jsonify(db.list_transfers())


@app.route("/api/transfers", methods=["POST"])
def api_transfer_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_transfer(data)
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM transfers WHERE id=?", (new_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/transfers/<int:tid>", methods=["GET"])
@login_required
def api_transfer_get(tid):
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM transfers WHERE id=?", (tid,)).fetchone()
    if not row:
        abort(404)
    return jsonify(dict(row))


@app.route("/api/transfers/<int:tid>", methods=["PUT"])
def api_transfer_update(tid):
    data = request.get_json(force=True, silent=True) or {}
    db.update_transfer(tid, data)
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM transfers WHERE id=?", (tid,)).fetchone()
    return jsonify(dict(row))


@app.route("/api/transfers/<int:tid>", methods=["DELETE"])
def api_transfer_delete(tid):
    db.delete_transfer(tid)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# RETENTION
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/retention", methods=["GET"])
def api_retention_list():
    return jsonify(db.list_retention())


@app.route("/api/retention", methods=["POST"])
def api_retention_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_retention(data)
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM retention_schedules WHERE id=?", (new_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/retention/<int:rid>", methods=["GET"])
@login_required
def api_retention_get(rid):
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM retention_schedules WHERE id=?", (rid,)).fetchone()
    if not row:
        abort(404)
    return jsonify(dict(row))


@app.route("/api/retention/<int:rid>", methods=["PUT"])
def api_retention_update(rid):
    data = request.get_json(force=True, silent=True) or {}
    db.update_retention(rid, data)
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM retention_schedules WHERE id=?", (rid,)).fetchone()
    return jsonify(dict(row))


@app.route("/api/retention/<int:rid>", methods=["DELETE"])
def api_retention_delete(rid):
    db.delete_retention(rid)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# SECURITY MEASURES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/security", methods=["GET"])
def api_security_list():
    return jsonify(db.list_security())


@app.route("/api/security", methods=["POST"])
def api_security_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_security(data)
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM security_measures WHERE id=?", (new_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/security/<int:sid>", methods=["GET"])
@login_required
def api_security_get(sid):
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM security_measures WHERE id=?", (sid,)).fetchone()
    if not row:
        abort(404)
    return jsonify(dict(row))


@app.route("/api/security/<int:sid>", methods=["PUT"])
def api_security_update(sid):
    data = request.get_json(force=True, silent=True) or {}
    db.update_security(sid, data)
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM security_measures WHERE id=?", (sid,)).fetchone()
    return jsonify(dict(row))


@app.route("/api/security/<int:sid>", methods=["DELETE"])
def api_security_delete(sid):
    db.delete_security(sid)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# AUDIT LOG
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/audit", methods=["GET"])
def api_audit_list():
    limit = int(request.args.get("limit", 200))
    return jsonify(db.list_audit(limit=limit))


# ═══════════════════════════════════════════════════════════════════════════════
# SETTINGS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/settings", methods=["GET"])
def api_settings_get():
    return jsonify(db.get_all_settings())


@app.route("/api/settings", methods=["PUT"])
def api_settings_update():
    data = request.get_json(force=True, silent=True) or {}
    for k, v in data.items():
        db.set_setting(k, v)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# POLICIES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/policies", methods=["GET"])
@login_required
def api_policies_list():
    return jsonify(db.list_policies(
        search=request.args.get("q"),
        status=request.args.get("status"),
        policy_type=request.args.get("type"),
    ))

@app.route("/api/policies", methods=["POST"])
@login_required
def api_policies_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_policy(data)
    return jsonify(db.get_policy(new_id)), 201

@app.route("/api/policies/<int:policy_id>", methods=["GET"])
@login_required
def api_policies_get(policy_id):
    p = db.get_policy(policy_id)
    if not p:
        abort(404)
    return jsonify(p)

@app.route("/api/policies/<int:policy_id>", methods=["PUT"])
@login_required
def api_policies_update(policy_id):
    data = request.get_json(force=True, silent=True) or {}
    db.update_policy(policy_id, data)
    return jsonify(db.get_policy(policy_id))

@app.route("/api/policies/<int:policy_id>", methods=["DELETE"])
@login_required
def api_policies_delete(policy_id):
    db.delete_policy(policy_id)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# TRAINING
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/training", methods=["GET"])
@login_required
def api_training_list():
    return jsonify(db.list_training(
        search=request.args.get("q"),
        department=request.args.get("department"),
    ))

@app.route("/api/training", methods=["POST"])
@login_required
def api_training_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_training(data)
    return jsonify(db.get_training(new_id)), 201

@app.route("/api/training/<int:training_id>", methods=["GET"])
@login_required
def api_training_get(training_id):
    t = db.get_training(training_id)
    if not t:
        abort(404)
    return jsonify(t)

@app.route("/api/training/<int:training_id>", methods=["PUT"])
@login_required
def api_training_update(training_id):
    data = request.get_json(force=True, silent=True) or {}
    db.update_training(training_id, data)
    return jsonify(db.get_training(training_id))

@app.route("/api/training/<int:training_id>", methods=["DELETE"])
@login_required
def api_training_delete(training_id):
    db.delete_training(training_id)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# TRAINING EXCEL UPLOAD & ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/training/analyse-excel", methods=["POST"])
@login_required
def api_training_analyse_excel():
    """
    Accept a multipart Excel file upload.
    Returns attendance analysis: who attended each session, who missed,
    overall attendance ratio per person and per session.
    Also bulk-imports records into training_records.
    """
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files['file']
    if not f.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        return jsonify({"error": "Please upload an Excel file (.xlsx / .xls)"}), 400

    import io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return jsonify({"error": f"Could not read Excel file: {e}"}), 400

    ws = wb.active

    # --- Auto-detect header row (first non-empty row) ---
    headers = []
    header_row_idx = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if any(c for c in row):
            headers = [str(c).strip() if c is not None else '' for c in row]
            header_row_idx = row_idx
            break

    if not headers:
        return jsonify({"error": "Excel file appears to be empty"}), 400

    # --- Normalize headers to known field names ---
    def norm(s):
        return s.lower().replace(' ', '').replace('_', '').replace('-', '').replace('/', '')

    header_map = {}
    known = {
        'staffname': 'staff_name',
        'name': 'staff_name',
        'employee': 'staff_name',
        'employeename': 'staff_name',
        'staffemail': 'staff_email',
        'email': 'staff_email',
        'emailaddress': 'staff_email',
        'department': 'department',
        'dept': 'department',
        'division': 'department',
        'training': 'training_name',
        'trainingname': 'training_name',
        'course': 'training_name',
        'coursename': 'training_name',
        'module': 'training_name',
        'trainingtype': 'training_type',
        'type': 'training_type',
        'completiondate': 'completion_date',
        'dateofcompletion': 'completion_date',
        'completeddate': 'completion_date',
        'date': 'completion_date',
        'expirydate': 'expiry_date',
        'expiresdate': 'expiry_date',
        'expiry': 'expiry_date',
        'result': 'passed',
        'passed': 'passed',
        'status': 'passed',
        'score': 'score',
        'mark': 'score',
        'grade': 'score',
        'certificate': 'certificate_no',
        'certificateno': 'certificate_no',
        'certnumber': 'certificate_no',
        'trainer': 'trainer',
        'provider': 'trainer',
        'trainingprovider': 'trainer',
        'regulation': 'regulation',
    }
    for i, h in enumerate(headers):
        mapped = known.get(norm(h))
        if mapped:
            header_map[i] = mapped

    # --- Parse rows ---
    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(c for c in row):
            continue
        rec = {}
        for col_i, val in enumerate(row):
            field = header_map.get(col_i)
            if field and val is not None:
                rec[field] = str(val).strip() if not isinstance(val, (int, float)) else val
        if rec.get('staff_name') or rec.get('training_name'):
            # Normalise 'passed' field
            p = str(rec.get('passed', '')).lower()
            rec['passed'] = 1 if p in ('yes', 'passed', 'true', '1', 'complete', 'completed') else 0
            records.append(rec)

    if not records:
        return jsonify({"error": "No data rows found. Check that your headers match: Staff Name, Training Name, Completion Date, Passed, Department, etc."}), 400

    # --- Compute attendance analytics ---
    from collections import defaultdict

    # Group by training_name (session)
    sessions = defaultdict(list)
    all_staff = set()
    for r in records:
        sn = r.get('training_name', 'Unknown Training')
        staff = r.get('staff_name', 'Unknown')
        sessions[sn].append(r)
        all_staff.add(staff)

    total_staff = len(all_staff)

    session_summary = []
    for session_name, rows in sessions.items():
        attendees = [r.get('staff_name') for r in rows if r.get('passed') or r.get('completion_date')]
        passed_count = sum(1 for r in rows if r.get('passed'))
        attendance_pct = round(len(attendees) / total_staff * 100) if total_staff else 0
        pass_rate = round(passed_count / len(rows) * 100) if rows else 0
        absent = sorted(all_staff - set(attendees))
        session_summary.append({
            "session": session_name,
            "total_registered": len(rows),
            "attended": len(attendees),
            "passed": passed_count,
            "absent_count": len(absent),
            "attendance_pct": attendance_pct,
            "pass_rate": pass_rate,
            "absent_staff": absent[:20],  # cap at 20 for response size
        })

    # Per-staff summary
    staff_sessions = defaultdict(lambda: {"attended": 0, "passed": 0, "sessions": []})
    for r in records:
        s = r.get('staff_name', 'Unknown')
        staff_sessions[s]["attended"] += 1 if (r.get('passed') or r.get('completion_date')) else 0
        staff_sessions[s]["passed"] += 1 if r.get('passed') else 0
        staff_sessions[s]["sessions"].append(r.get('training_name', ''))

    total_sessions = len(sessions)
    staff_summary = []
    for name, data in sorted(staff_sessions.items()):
        att = data['attended']
        overall_pct = round(att / total_sessions * 100) if total_sessions else 0
        missed = [s for s in sessions if s not in data['sessions']]
        staff_summary.append({
            "name": name,
            "department": records[next((i for i, r in enumerate(records) if r.get('staff_name') == name), 0)].get('department', ''),
            "attended": att,
            "passed": data['passed'],
            "total_sessions": total_sessions,
            "attendance_pct": overall_pct,
            "missed_sessions": missed,
        })

    # Overall stats
    total_possible = len(records)
    total_attended = sum(1 for r in records if r.get('passed') or r.get('completion_date'))
    total_passed = sum(1 for r in records if r.get('passed'))
    overall_attendance = round(total_attended / total_possible * 100) if total_possible else 0
    overall_pass = round(total_passed / total_possible * 100) if total_possible else 0

    # Bulk import into DB
    imported = 0
    skipped = 0
    for r in records:
        try:
            db.create_training(r)
            imported += 1
        except Exception:
            skipped += 1

    return jsonify({
        "ok": True,
        "summary": {
            "total_staff": total_staff,
            "total_sessions": total_sessions,
            "total_records": len(records),
            "overall_attendance_pct": overall_attendance,
            "overall_pass_rate": overall_pass,
            "imported": imported,
            "skipped_duplicates": skipped,
        },
        "sessions": sorted(session_summary, key=lambda x: x['attendance_pct']),
        "staff": sorted(staff_summary, key=lambda x: x['attendance_pct']),
        "headers_detected": [h for h in headers if h],
        "mapped_columns": {headers[i]: v for i, v in header_map.items()},
    })


# ═══════════════════════════════════════════════════════════════════════════════
# DATA FLOWS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/dataflows", methods=["GET"])
@login_required
def api_dataflows_list():
    return jsonify(db.list_dataflows())

@app.route("/api/dataflows", methods=["POST"])
@login_required
def api_dataflows_create():
    data = request.get_json(force=True, silent=True) or {}
    new_id = db.create_dataflow(data)
    return jsonify({"id": new_id}), 201

@app.route("/api/dataflows/<int:flow_id>", methods=["PUT"])
@login_required
def api_dataflows_update(flow_id):
    data = request.get_json(force=True, silent=True) or {}
    db.update_dataflow(flow_id, data)
    return jsonify({"ok": True})

@app.route("/api/dataflows/<int:flow_id>", methods=["DELETE"])
@login_required
def api_dataflows_delete(flow_id):
    db.delete_dataflow(flow_id)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# CALENDAR & COMPLIANCE SCORE
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/calendar")
@login_required
def api_calendar():
    days = int(request.args.get("days", 90))
    return jsonify(db.get_calendar_events(days))


@app.route("/api/compliance-score")
@login_required
def api_compliance_score():
    return jsonify(db.get_compliance_score())


# ═══════════════════════════════════════════════════════════════════════════════
# ROPA → DPIA SPAWN
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/ropa/<int:ropa_id>/spawn-dpia", methods=["POST"])
@login_required
def api_spawn_dpia_from_ropa(ropa_id):
    ropa = db.get_ropa(ropa_id)
    if not ropa:
        abort(404)
    # Pre-populate DPIA from RoPA data
    dpia_data = {
        "title": f"DPIA — {ropa.get('processing_name', '')}",
        "status": "draft",
        "regulation": ropa.get("regulation", "GDPR"),
        "activity_desc": ropa.get("purpose", ""),
        "purpose": ropa.get("purpose", ""),
        "legal_basis": ropa.get("legal_basis", ""),
        "data_categories": ropa.get("data_categories", ""),
        "data_subjects": ropa.get("data_subjects", ""),
        "subject_count": ropa.get("subject_count", ""),
        "retention": ropa.get("retention_period", ""),
        "systems": ropa.get("systems", ""),
        "processors": ropa.get("processors", ""),
        "intl_transfer": ropa.get("intl_transfers", ""),
        "transfer_dest": ropa.get("transfer_dest", ""),
        "controller_name": ropa.get("controller_name", ""),
        "dpo_name": ropa.get("dpo_name", ""),
        "dpo_email": ropa.get("dpo_email", ""),
    }
    new_id = db.create_dpia(dpia_data)
    # Link the DPIA back to the RoPA
    db.update_ropa(ropa_id, {"dpia_id": new_id})
    dpia = db.get_dpia(new_id)
    return jsonify({"ok": True, "dpia_id": new_id, "dpia": dpia}), 201


# ═══════════════════════════════════════════════════════════════════════════════
# AUDIT EVIDENCE PACK
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/audit-export")
@login_required
def api_audit_export():
    """Export a ZIP containing: RoPA xlsx, DPIA summaries, breach log, DSR log, vendor list, audit log."""
    import io, zipfile, json as _json
    settings = db.get_all_settings()
    org = settings.get("org_name", "Organisation").replace(" ", "_")
    date_str = datetime.utcnow().strftime("%Y%m%d")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # RoPA Excel
        try:
            from xlsx_export import generate_ropa_xlsx
            ropa_entries = db.list_ropa()
            ropa_buf = generate_ropa_xlsx(ropa_entries, settings)
            zf.writestr(f"RoPA_{org}_{date_str}.xlsx", ropa_buf.read())
        except Exception as e:
            zf.writestr("RoPA_error.txt", str(e))

        # DPIA JSON summary
        dpias = db.list_dpias()
        zf.writestr(f"DPIA_Summary_{date_str}.json", _json.dumps(dpias, indent=2, default=str))

        # Breach log
        breaches = db.list_breaches()
        zf.writestr(f"Breach_Log_{date_str}.json", _json.dumps(breaches, indent=2, default=str))

        # DSR log
        dsrs = db.list_dsrs()
        zf.writestr(f"DSR_Log_{date_str}.json", _json.dumps(dsrs, indent=2, default=str))

        # Vendor register
        vendors = db.list_vendors()
        zf.writestr(f"Vendor_Register_{date_str}.json", _json.dumps(vendors, indent=2, default=str))

        # Audit trail
        audit = db.list_audit(limit=10000)
        zf.writestr(f"Audit_Trail_{date_str}.json", _json.dumps(audit, indent=2, default=str))

        # Policies
        policies = db.list_policies()
        zf.writestr(f"Policy_Register_{date_str}.json", _json.dumps(policies, indent=2, default=str))

        # Compliance score
        score = db.get_compliance_score()
        zf.writestr(f"Compliance_Score_{date_str}.json", _json.dumps(score, indent=2, default=str))

        # README
        readme = f"""DATA PROTECTION SENTINEL — AUDIT EVIDENCE PACK
Generated: {datetime.utcnow().strftime('%d %B %Y %H:%M UTC')}
Organisation: {settings.get('org_name', '—')}
DPO: {settings.get('dpo_name', '—')} ({settings.get('dpo_email', '—')})
Primary Regulation: {settings.get('primary_regulation', '—')}
Overall Compliance Score: {score.get('overall', '—')}% (Grade {score.get('grade', '—')})

CONTENTS:
- RoPA_{org}_{date_str}.xlsx  — Record of Processing Activities (Excel)
- DPIA_Summary_{date_str}.json — DPIA Register
- Breach_Log_{date_str}.json — Data Breach Register
- DSR_Log_{date_str}.json — Data Subject Request Log
- Vendor_Register_{date_str}.json — Processor/Vendor Register
- Audit_Trail_{date_str}.json — System Audit Log
- Policy_Register_{date_str}.json — Policy & Document Register
- Compliance_Score_{date_str}.json — Compliance Scorecard

Generated by Data Protection Sentinel · by Ali Moyo
"""
        zf.writestr("README.txt", readme)

    zip_buf.seek(0)
    return send_file(
        zip_buf, as_attachment=True,
        download_name=f"AuditPack_{org}_{date_str}.zip",
        mimetype="application/zip"
    )


# ── Legal bases by regulation ─────────────────────────────────────────────────
@app.route("/api/legal-bases/<regulation>")
def api_legal_bases(regulation):
    bases = ai.LEGAL_BASES.get(regulation, ai.LEGAL_BASES.get("GDPR", []))
    return jsonify(bases)


# ═══════════════════════════════════════════════════════════════════════════════
# AI ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/ai/research", methods=["POST"])
def api_ai_research():
    body = request.get_json(force=True, silent=True) or {}
    activity   = body.get("activity_type", "")
    regulation = body.get("regulation", "GDPR")
    context    = body.get("context", "")
    if not activity:
        return jsonify({"error": "activity_type is required"}), 400
    text, err = ai.ai_research(activity, regulation, context)
    if err:
        return jsonify({"error": err}), 500
    return jsonify({"research": text})


@app.route("/api/ai/generate/<int:dpia_id>", methods=["POST"])
def api_ai_generate(dpia_id):
    dpia = db.get_dpia(dpia_id)
    if not dpia:
        return jsonify({"error": "DPIA not found"}), 404
    text, err = ai.ai_generate_full_dpia(dpia)
    if err:
        return jsonify({"error": err}), 500
    db.update_dpia(dpia_id, {"ai_full_dpia": text})
    return jsonify({"content": text})


@app.route("/api/ai/risks", methods=["POST"])
def api_ai_risks():
    body = request.get_json(force=True, silent=True) or {}
    activity   = body.get("activity_type", "")
    regulation = body.get("regulation", "GDPR")
    categories = body.get("data_categories", [])
    if not activity:
        return jsonify({"error": "activity_type is required"}), 400
    risks, err = ai.ai_suggest_risks(activity, regulation, categories)
    if err:
        return jsonify({"error": err}), 500
    return jsonify({"risks": risks})


@app.route("/api/ai/score-ropa/<int:ropa_id>", methods=["POST"])
def api_ai_score_ropa(ropa_id):
    entry = db.get_ropa(ropa_id)
    if not entry:
        return jsonify({"error": "RoPA entry not found"}), 404
    result, err = ai.ai_score_ropa(entry)
    if err:
        return jsonify({"error": err}), 500
    # Auto-update the entry with the AI score
    update_data = {"risk_score": result.get("risk_score", "medium"), "ai_risk_notes": result.get("rationale", "")}
    if result.get("dpia_required"):
        update_data["dpia_required"] = 1
    db.update_ropa(ropa_id, update_data)
    return jsonify(result)


@app.route("/api/ai/breach-impact/<int:breach_id>", methods=["POST"])
def api_ai_breach_impact(breach_id):
    breach = db.get_breach(breach_id)
    if not breach:
        return jsonify({"error": "Breach not found"}), 404
    text, err = ai.ai_assess_breach(breach)
    if err:
        return jsonify({"error": err}), 500
    db.update_breach(breach_id, {"ai_assessment": text})
    return jsonify({"assessment": text})


@app.route("/api/ai/dsr-draft/<int:dsr_id>", methods=["POST"])
def api_ai_dsr_draft(dsr_id):
    with db.get_db() as conn:
        row = conn.execute("SELECT * FROM dsr_requests WHERE id=?", (dsr_id,)).fetchone()
    if not row:
        return jsonify({"error": "DSR not found"}), 404
    dsr = dict(row)
    text, err = ai.ai_draft_dsr_response(dsr)
    if err:
        return jsonify({"error": err}), 500
    db.update_dsr(dsr_id, {"ai_draft": text})
    return jsonify({"draft": text})


@app.route("/api/ai/privacy-notice", methods=["POST"])
def api_ai_privacy_notice():
    body = request.get_json(force=True, silent=True) or {}
    # Optionally load from org settings
    settings = db.get_all_settings()
    body.setdefault("org_name", settings.get("org_name", ""))
    body.setdefault("dpo_name", settings.get("dpo_name", ""))
    body.setdefault("dpo_email", settings.get("dpo_email", ""))
    text, err = ai.ai_generate_privacy_notice(body)
    if err:
        return jsonify({"error": err}), 500
    return jsonify({"notice": text})


@app.route("/api/ai/vendor-check/<int:vendor_id>", methods=["POST"])
def api_ai_vendor_check(vendor_id):
    vendor = db.get_vendor(vendor_id)
    if not vendor:
        return jsonify({"error": "Vendor not found"}), 404
    text, err = ai.ai_vendor_check(vendor)
    if err:
        return jsonify({"error": err}), 500
    db.update_vendor(vendor_id, {"ai_assessment": text})
    return jsonify({"assessment": text})


@app.route("/api/ai/chat", methods=["POST"])
def api_ai_chat():
    body = request.get_json(force=True, silent=True) or {}
    message    = body.get("message", "")
    regulation = body.get("regulation")
    history    = body.get("history", [])
    if not message:
        return jsonify({"error": "message is required"}), 400
    text, err = ai.ai_chat(message, regulation=regulation, history=history)
    if err:
        return jsonify({"error": err}), 500
    return jsonify({"response": text})


@app.route("/api/ai/gap-analysis", methods=["POST"])
def api_ai_gap_analysis():
    body = request.get_json(force=True, silent=True) or {}
    reg_from   = body.get("regulation_from", "GDPR")
    reg_to     = body.get("regulation_to", "South Africa POPIA")
    activities = body.get("activities", "")
    text, err = ai.ai_gap_analysis(reg_from, reg_to, activities)
    if err:
        return jsonify({"error": err}), 500
    return jsonify({"analysis": text})


# ═══════════════════════════════════════════════════════════════════════════════
# RUN
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    debug = os.getenv("FLASK_DEBUG", "false").lower() == "true"
    port  = int(os.getenv("PORT", 5000))
    print(f"""
╔══════════════════════════════════════════════════════╗
║      DATA PROTECTION SENTINEL — by Ali Moyo          ║
║      http://localhost:{port}                           ║
╚══════════════════════════════════════════════════════╝
    """)
    app.run(debug=debug, host="0.0.0.0", port=port)
