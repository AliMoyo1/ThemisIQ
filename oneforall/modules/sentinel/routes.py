"""
Sentinel module — Data Protection routes.

FastAPI router ported from the original Flask-based Data Protection Sentinel.
Sentinel is an SPA: one HTML template serves the UI, everything else is JSON API.
All routes are prefixed with /sentinel and guarded by capability-based RBAC.
"""
import io
import json
import os
import zipfile
from datetime import datetime
from core.timeutils import utcnow, to_dt

from fastapi import APIRouter, Request, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from config import settings
from core.middleware import require_module, require_capability
from core.shell_context import shell_ctx
from core.events import (
    emit, SENTINEL_BREACH_CONFIRMED, SENTINEL_BREACH_RESOLVED,
    SENTINEL_DPIA_COMPLETED, SENTINEL_DSR_OVERDUE,
)

from modules.sentinel import data_service as ds

router = APIRouter(prefix="/sentinel", tags=["sentinel"])

# Templates instantiated once at module load — not per-request
_tpl_dir = os.path.join(os.path.dirname(__file__), "templates")
_templates = Jinja2Templates(directory=[_tpl_dir, "templates"])


# ── Helper ───────────────────────────────────────────────────────────────────

async def _json_body(request: Request) -> dict:
    try:
        return await request.json()
    except Exception:
        return {}


# ═════════════════════════════════════════════════════════════════════════════
# SPA INDEX
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/", response_class=HTMLResponse)
@require_module("sentinel")
async def sentinel_index(request: Request):
    """Serve the Sentinel SPA shell."""
    templates = _templates  # use module-level instance

    # Import reference data for the SPA
    try:
        from modules.sentinel.ai_service import (
            REGULATIONS, ACTIVITY_TYPES, DATA_CATEGORIES, SPECIAL_CATEGORIES
        )
    except ImportError:
        REGULATIONS = []
        ACTIVITY_TYPES = []
        DATA_CATEGORIES = []
        SPECIAL_CATEGORIES = []

    user = request.state.user
    # Build current_user dict matching the SPA's expected format
    current_user = {
        "id": user["id"],
        "username": user["username"],
        "full_name": user.get("full_name", ""),
        "email": user.get("email", ""),
        "role": "admin" if any(
            r in user.get("roles", [])
            for r in ("super_admin", "org_admin")
        ) else "user",
        "avatar_initials": user.get("avatar_initials", "??"),
    }
    return templates.TemplateResponse(request, "index.html", {
        "user": user,
        "current_user": current_user,
        "module": "sentinel",
        "regulations": REGULATIONS,
        "activity_types": ACTIVITY_TYPES,
        "data_categories": DATA_CATEGORIES,
        "special_categories": SPECIAL_CATEGORIES,
        **shell_ctx(request, active_module="sentinel", active_section="dashboard"),
    })


# SPA catch-all — lets browser refresh on /sentinel/ropa etc. work
_SPA_PAGES = {
    "ropa", "dpia", "breaches", "dsr", "vendors", "consent",
    "reports", "chat", "settings",
    # Previously API-only, now have UI (Item 4)
    "transfers", "retention", "security", "policies", "notices",
    "dataflows", "controllers", "training",
    # New features
    "lia",           # SENT-14: Legitimate Interest Assessments
    "jurisdictions", # Multi-jurisdiction manager
}

@router.get("/{page}", response_class=HTMLResponse)
@require_module("sentinel")
async def sentinel_spa_page(request: Request, page: str):
    if page.startswith("api") or page not in _SPA_PAGES:
        raise HTTPException(404)
    return await sentinel_index(request)


# ═════════════════════════════════════════════════════════════════════════════
# CROSS-MODULE: Active breach status (used by BCM, GRID, ORM dashboards)
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/active-breach")
@require_module("sentinel")
async def api_active_breach(request: Request):
    """
    Lightweight endpoint — returns {active: bool, id: int|null, severity: str|null}
    Used by BCM, GRID, and ORM module dashboards to check for active data breaches
    without coupling directly to sentinel data_service.
    """
    from database import get_db as _get_db
    db = _get_db()
    try:
        row = db.execute(
            "SELECT id, title, severity FROM sentinel_breaches "
            "WHERE status NOT IN ('closed','resolved') "
            "ORDER BY CASE severity WHEN 'critical' THEN 0 WHEN 'high' THEN 1 ELSE 2 END "
            "LIMIT 1"
        ).fetchone()
        if row:
            return JSONResponse({
                "active": True,
                "id": row["id"],
                "title": row["title"],
                "severity": row["severity"],
            })
        return JSONResponse({"active": False, "id": None, "title": None, "severity": None})
    except Exception:
        return JSONResponse({"active": False, "id": None, "title": None, "severity": None})
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# STATS
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/stats")
@require_capability("module.sentinel.access")
async def api_stats(request: Request):
    return JSONResponse(ds.get_stats())


# ═════════════════════════════════════════════════════════════════════════════
# ROPA
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/ropa")
@require_capability("module.sentinel.access")
async def api_ropa_list(request: Request):
    entries = ds.list_ropa(
        search=request.query_params.get("q"),
        regulation=request.query_params.get("regulation"),
        status=request.query_params.get("status"),
        risk=request.query_params.get("risk"),
    )
    return JSONResponse(entries)


@router.post("/api/ropa", status_code=201)
@require_capability("sentinel.ropa.manage")
async def api_ropa_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_ropa(data)
    entry = ds.get_ropa(new_id)
    return JSONResponse(entry, status_code=201)


@router.get("/api/ropa/{ropa_id}")
@require_capability("module.sentinel.access")
async def api_ropa_get(request: Request, ropa_id: int):
    entry = ds.get_ropa(ropa_id)
    if not entry:
        raise HTTPException(404, "RoPA entry not found")
    return JSONResponse(entry)


@router.put("/api/ropa/{ropa_id}")
@require_capability("sentinel.ropa.manage")
async def api_ropa_update(request: Request, ropa_id: int):
    data = await _json_body(request)
    ds.update_ropa(ropa_id, data)
    return JSONResponse(ds.get_ropa(ropa_id))


@router.delete("/api/ropa/{ropa_id}")
@require_capability("sentinel.ropa.manage")
async def api_ropa_delete(request: Request, ropa_id: int):
    ds.delete_ropa(ropa_id)
    return JSONResponse({"ok": True})


@router.post("/api/ropa/{ropa_id}/spawn-dpia", status_code=201)
@require_capability("sentinel.dpia.manage")
async def api_spawn_dpia_from_ropa(request: Request, ropa_id: int):
    ropa = ds.get_ropa(ropa_id)
    if not ropa:
        raise HTTPException(404, "RoPA entry not found")
    def _str(val):
        """Ensure lists/dicts become JSON strings for text columns."""
        if isinstance(val, (list, dict)):
            return json.dumps(val)
        return val or ""

    dpia_data = {
        "title": f"DPIA — {ropa.get('processing_name', '')}",
        "status": "draft",
        "regulation": ropa.get("regulation", "GDPR"),
        "activity_desc": _str(ropa.get("purpose", "")),
        "purpose": _str(ropa.get("purpose", "")),
        "legal_basis": _str(ropa.get("legal_basis", "")),
        "data_categories": _str(ropa.get("data_categories", "")),
        "data_subjects": _str(ropa.get("data_subjects", "")),
        "subject_count": _str(ropa.get("subject_count", "")),
        "retention": _str(ropa.get("retention_period", "")),
        "systems": _str(ropa.get("systems", "")),
        "processors": _str(ropa.get("processors", "")),
        "intl_transfer": _str(ropa.get("intl_transfers", "")),
        "transfer_dest": _str(ropa.get("transfer_dest", "")),
        "controller_name": _str(ropa.get("controller_name", "")),
        "dpo_name": _str(ropa.get("dpo_name", "")),
        "dpo_email": _str(ropa.get("dpo_email", "")),
    }
    new_id = ds.create_dpia(dpia_data)
    ds.update_ropa(ropa_id, {"dpia_id": new_id})
    dpia = ds.get_dpia(new_id)
    return JSONResponse({"ok": True, "dpia_id": new_id, "dpia": dpia}, status_code=201)


# ═════════════════════════════════════════════════════════════════════════════
# DPIA
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/dpias")
@require_capability("sentinel.dpia.manage")
async def api_dpia_list(request: Request):
    dpias = ds.list_dpias(
        search=request.query_params.get("q"),
        regulation=request.query_params.get("regulation"),
        status=request.query_params.get("status"),
    )
    return JSONResponse(dpias)


@router.post("/api/dpias", status_code=201)
@require_capability("sentinel.dpia.manage")
async def api_dpia_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_dpia(data)
    return JSONResponse(ds.get_dpia(new_id), status_code=201)


@router.get("/api/dpias/{dpia_id}")
@require_capability("sentinel.dpia.manage")
async def api_dpia_get(request: Request, dpia_id: int):
    dpia = ds.get_dpia(dpia_id)
    if not dpia:
        raise HTTPException(404, "DPIA not found")
    return JSONResponse(dpia)


@router.put("/api/dpias/{dpia_id}")
@require_capability("sentinel.dpia.manage")
async def api_dpia_update(request: Request, dpia_id: int):
    data = await _json_body(request)
    ds.update_dpia(dpia_id, data)

    # Emit DPIA completed event when status changes to completed/approved
    new_status = (data.get("status") or "").lower()
    if new_status in ("completed", "approved"):
        dpia = ds.get_dpia(dpia_id)
        emit(
            SENTINEL_DPIA_COMPLETED,
            source_module="sentinel",
            entity_type="dpia",
            entity_id=dpia_id,
            payload={
                "title": dpia.get("title", "") if dpia else "",
                "risk_level": dpia.get("risk_level", "") if dpia else "",
                "recommendations": dpia.get("recommendations", "") if dpia else "",
                "findings": dpia.get("findings", "") if dpia else "",
            },
            user_id=request.state.user["id"],
        )
    return JSONResponse(ds.get_dpia(dpia_id))


@router.delete("/api/dpias/{dpia_id}")
@require_capability("sentinel.dpia.manage")
async def api_dpia_delete(request: Request, dpia_id: int):
    ds.delete_dpia(dpia_id)
    return JSONResponse({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# BREACHES
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/breaches")
@require_capability("sentinel.breach.manage")
async def api_breach_list(request: Request):
    return JSONResponse(ds.list_breaches(
        search=request.query_params.get("q"),
        status=request.query_params.get("status"),
        severity=request.query_params.get("severity"),
    ))


@router.post("/api/breaches", status_code=201)
@require_capability("sentinel.breach.manage")
async def api_breach_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_breach(data)

    # Emit breach confirmed event
    emit(
        SENTINEL_BREACH_CONFIRMED,
        source_module="sentinel",
        entity_type="breach",
        entity_id=new_id,
        payload={
            "title": data.get("title", ""),
            "severity": data.get("severity", ""),
            "category": data.get("category", ""),
            "affected_records": data.get("affected_records", 0),
            "description": data.get("description", ""),
            "regulation": data.get("regulation") or settings.DEFAULT_REGULATION,
            "active_jurisdictions": [j["jurisdiction_key"] for j in ds.get_active_jurisdictions()],
        },
        user_id=request.state.user["id"],
    )
    return JSONResponse(ds.get_breach(new_id), status_code=201)


@router.get("/api/breaches/{breach_id}")
@require_capability("sentinel.breach.manage")
async def api_breach_get(request: Request, breach_id: int):
    b = ds.get_breach(breach_id)
    if not b:
        raise HTTPException(404, "Breach not found")
    return JSONResponse(b)


@router.put("/api/breaches/{breach_id}")
@require_capability("sentinel.breach.manage")
async def api_breach_update(request: Request, breach_id: int):
    data = await _json_body(request)
    ds.update_breach(breach_id, data)

    # Emit events based on status transition
    new_status = (data.get("status") or "").lower()
    breach = ds.get_breach(breach_id)
    if new_status == "confirmed":
        emit(
            SENTINEL_BREACH_CONFIRMED,
            source_module="sentinel",
            entity_type="breach",
            entity_id=breach_id,
            payload={
                "title": breach.get("title", "") if breach else "",
                "severity": breach.get("severity", "") if breach else "",
                "affected_records": breach.get("affected_records", 0) if breach else 0,
                "regulation": (breach.get("regulation") if breach else None) or settings.DEFAULT_REGULATION,
                "active_jurisdictions": [j["jurisdiction_key"] for j in ds.get_active_jurisdictions()],
            },
            user_id=request.state.user["id"],
        )
    elif new_status in ("closed", "resolved", "contained"):
        emit(
            SENTINEL_BREACH_RESOLVED,
            source_module="sentinel",
            entity_type="breach",
            entity_id=breach_id,
            payload={
                "title": breach.get("title", "") if breach else "",
                "severity": breach.get("severity", "") if breach else "",
                "resolution_notes": data.get("resolution_notes", ""),
                "closed_status": new_status,
            },
            user_id=request.state.user["id"],
        )
    return JSONResponse(breach)


@router.delete("/api/breaches/{breach_id}")
@require_capability("sentinel.breach.manage")
async def api_breach_delete(request: Request, breach_id: int):
    ds.delete_breach(breach_id)
    return JSONResponse({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# DSR (Data Subject Requests)
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/dsrs")
@require_capability("sentinel.dsr.manage")
async def api_dsr_list(request: Request):
    return JSONResponse(ds.list_dsrs(
        search=request.query_params.get("q"),
        status=request.query_params.get("status"),
        request_type=request.query_params.get("type"),
    ))


@router.post("/api/dsrs", status_code=201)
@require_capability("sentinel.dsr.manage")
async def api_dsr_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_dsr(data)
    return JSONResponse(ds.get_dsr(new_id), status_code=201)


@router.get("/api/dsrs/{dsr_id}")
@require_capability("sentinel.dsr.manage")
async def api_dsr_get(request: Request, dsr_id: int):
    d = ds.get_dsr(dsr_id)
    if not d:
        raise HTTPException(404, "DSR not found")
    return JSONResponse(d)


@router.put("/api/dsrs/{dsr_id}")
@require_capability("sentinel.dsr.manage")
async def api_dsr_update(request: Request, dsr_id: int):
    data = await _json_body(request)
    ds.update_dsr(dsr_id, data)

    # Check if DSR is now overdue and emit event
    dsr = ds.get_dsr(dsr_id)
    if dsr:
        deadline = dsr.get("deadline_date") or ""
        status = (dsr.get("status") or "").lower()
        if deadline and status not in ("completed", "closed"):
            try:
                due = to_dt(deadline[:10])
                if due < utcnow():
                    emit(
                        SENTINEL_DSR_OVERDUE,
                        source_module="sentinel",
                        entity_type="dsr",
                        entity_id=dsr_id,
                        payload={
                            "subject_name": dsr.get("subject_name", ""),
                            "request_type": dsr.get("request_type", ""),
                            "deadline": deadline,
                            "days_overdue": (utcnow() - due).days,
                        },
                        user_id=request.state.user["id"],
                    )
            except (ValueError, TypeError):
                pass
    return JSONResponse(ds.get_dsr(dsr_id))


@router.delete("/api/dsrs/{dsr_id}")
@require_capability("sentinel.dsr.manage")
async def api_dsr_delete(request: Request, dsr_id: int):
    ds.delete_dsr(dsr_id)
    return JSONResponse({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# VENDORS
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/vendors")
@require_capability("sentinel.vendor.manage")
async def api_vendor_list(request: Request):
    return JSONResponse(ds.list_vendors(
        search=request.query_params.get("q"),
        risk=request.query_params.get("risk"),
        dpa_status=request.query_params.get("dpa_status"),
    ))


@router.post("/api/vendors", status_code=201)
@require_capability("sentinel.vendor.manage")
async def api_vendor_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_vendor(data)
    vendor = ds.get_vendor(new_id)
    emit("vendor.created", source_module="sentinel", entity_type="vendor", entity_id=new_id,
         payload={"name": vendor.get("name", ""), "canonical_id": vendor.get("canonical_id"),
                  "source_module": "sentinel"},
         user_id=request.state.user["id"])
    return JSONResponse(vendor, status_code=201)


@router.get("/api/vendors/{vendor_id}")
@require_capability("sentinel.vendor.manage")
async def api_vendor_get(request: Request, vendor_id: int):
    v = ds.get_vendor(vendor_id)
    if not v:
        raise HTTPException(404, "Vendor not found")
    return JSONResponse(v)


@router.put("/api/vendors/{vendor_id}")
@require_capability("sentinel.vendor.manage")
async def api_vendor_update(request: Request, vendor_id: int):
    data = await _json_body(request)
    ds.update_vendor(vendor_id, data)
    return JSONResponse(ds.get_vendor(vendor_id))


@router.delete("/api/vendors/{vendor_id}")
@require_capability("sentinel.vendor.manage")
async def api_vendor_delete(request: Request, vendor_id: int):
    ds.delete_vendor(vendor_id)
    return JSONResponse({"ok": True})


@router.get("/api/vendors/{vendor_id}/cross-module")
@require_capability("sentinel.vendor.manage")
async def api_vendor_cross_module(request: Request, vendor_id: int):
    v = ds.get_vendor(vendor_id)
    if not v or not v.get("canonical_id"):
        return JSONResponse({"modules": {}, "flags": [], "canonical_id": None})
    from core.vendor_link import get_cross_module_profile
    from database import get_db
    db = get_db()
    try:
        return JSONResponse(get_cross_module_profile(db, v["canonical_id"]))
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# PRIVACY NOTICES
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/notices")
@require_capability("sentinel.privacy_notice.manage")
async def api_notice_list(request: Request):
    return JSONResponse(ds.list_notices())


@router.post("/api/notices", status_code=201)
@require_capability("sentinel.privacy_notice.manage")
async def api_notice_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_notice(data)
    return JSONResponse(ds.get_notice(new_id), status_code=201)


@router.get("/api/notices/{notice_id}")
@require_capability("sentinel.privacy_notice.manage")
async def api_notice_get(request: Request, notice_id: int):
    n = ds.get_notice(notice_id)
    if not n:
        raise HTTPException(404, "Notice not found")
    return JSONResponse(n)


@router.put("/api/notices/{notice_id}")
@require_capability("sentinel.privacy_notice.manage")
async def api_notice_update(request: Request, notice_id: int):
    data = await _json_body(request)
    ds.update_notice(notice_id, data)
    return JSONResponse(ds.get_notice(notice_id))


@router.delete("/api/notices/{notice_id}")
@require_capability("sentinel.privacy_notice.manage")
async def api_notice_delete(request: Request, notice_id: int):
    ds.delete_notice(notice_id)
    return JSONResponse({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# CONSENT RECORDS
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/consent")
@require_capability("sentinel.consent.manage")
async def api_consent_list(request: Request):
    return JSONResponse(ds.list_consent(
        search=request.query_params.get("q"),
        status=request.query_params.get("status"),
    ))


@router.post("/api/consent", status_code=201)
@require_capability("sentinel.consent.manage")
async def api_consent_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_consent(data)
    return JSONResponse(ds.get_consent(new_id), status_code=201)


@router.get("/api/consent/{consent_id}")
@require_capability("sentinel.consent.manage")
async def api_consent_get(request: Request, consent_id: int):
    c = ds.get_consent(consent_id)
    if not c:
        raise HTTPException(404, "Consent record not found")
    return JSONResponse(c)


@router.put("/api/consent/{consent_id}")
@require_capability("sentinel.consent.manage")
async def api_consent_update(request: Request, consent_id: int):
    data = await _json_body(request)
    ds.update_consent(consent_id, data)
    return JSONResponse(ds.get_consent(consent_id))


@router.delete("/api/consent/{consent_id}")
@require_capability("sentinel.consent.manage")
async def api_consent_delete(request: Request, consent_id: int):
    ds.delete_consent(consent_id)
    return JSONResponse({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# CONTROLLERS
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/controllers")
@require_capability("sentinel.controller.manage")
async def api_controller_list(request: Request):
    return JSONResponse(ds.list_controllers())


@router.post("/api/controllers", status_code=201)
@require_capability("sentinel.controller.manage")
async def api_controller_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_controller(data)
    return JSONResponse(ds.get_controller(new_id), status_code=201)


@router.get("/api/controllers/{ctrl_id}")
@require_capability("sentinel.controller.manage")
async def api_controller_get(request: Request, ctrl_id: int):
    c = ds.get_controller(ctrl_id)
    if not c:
        raise HTTPException(404, "Controller not found")
    return JSONResponse(c)


@router.put("/api/controllers/{ctrl_id}")
@require_capability("sentinel.controller.manage")
async def api_controller_update(request: Request, ctrl_id: int):
    data = await _json_body(request)
    ds.update_controller(ctrl_id, data)
    return JSONResponse(ds.get_controller(ctrl_id))


@router.delete("/api/controllers/{ctrl_id}")
@require_capability("sentinel.controller.manage")
async def api_controller_delete(request: Request, ctrl_id: int):
    ds.delete_controller(ctrl_id)
    return JSONResponse({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# TRANSFERS
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/transfers")
@require_capability("sentinel.transfer.manage")
async def api_transfer_list(request: Request):
    return JSONResponse(ds.list_transfers())


@router.post("/api/transfers", status_code=201)
@require_capability("sentinel.transfer.manage")
async def api_transfer_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_transfer(data)
    return JSONResponse(ds.get_transfer(new_id), status_code=201)


@router.get("/api/transfers/{tid}")
@require_capability("sentinel.transfer.manage")
async def api_transfer_get(request: Request, tid: int):
    t = ds.get_transfer(tid)
    if not t:
        raise HTTPException(404, "Transfer not found")
    return JSONResponse(t)


@router.put("/api/transfers/{tid}")
@require_capability("sentinel.transfer.manage")
async def api_transfer_update(request: Request, tid: int):
    data = await _json_body(request)
    ds.update_transfer(tid, data)
    return JSONResponse(ds.get_transfer(tid))


@router.delete("/api/transfers/{tid}")
@require_capability("sentinel.transfer.manage")
async def api_transfer_delete(request: Request, tid: int):
    ds.delete_transfer(tid)
    return JSONResponse({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# RETENTION SCHEDULES
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/retention")
@require_capability("sentinel.retention.manage")
async def api_retention_list(request: Request):
    return JSONResponse(ds.list_retention())


@router.post("/api/retention", status_code=201)
@require_capability("sentinel.retention.manage")
async def api_retention_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_retention(data)
    return JSONResponse(ds.get_retention(new_id), status_code=201)


@router.get("/api/retention/{rid}")
@require_capability("sentinel.retention.manage")
async def api_retention_get(request: Request, rid: int):
    r = ds.get_retention(rid)
    if not r:
        raise HTTPException(404, "Retention schedule not found")
    return JSONResponse(r)


@router.put("/api/retention/{rid}")
@require_capability("sentinel.retention.manage")
async def api_retention_update(request: Request, rid: int):
    data = await _json_body(request)
    ds.update_retention(rid, data)
    return JSONResponse(ds.get_retention(rid))


@router.delete("/api/retention/{rid}")
@require_capability("sentinel.retention.manage")
async def api_retention_delete(request: Request, rid: int):
    ds.delete_retention(rid)
    return JSONResponse({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# SECURITY MEASURES
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/security")
@require_capability("module.sentinel.access")
async def api_security_list(request: Request):
    return JSONResponse(ds.list_security())


@router.post("/api/security", status_code=201)
@require_capability("sentinel.ropa.manage")
async def api_security_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_security(data)
    return JSONResponse(ds.get_security(new_id), status_code=201)


@router.get("/api/security/{sid}")
@require_capability("module.sentinel.access")
async def api_security_get(request: Request, sid: int):
    s = ds.get_security(sid)
    if not s:
        raise HTTPException(404, "Security measure not found")
    return JSONResponse(s)


@router.put("/api/security/{sid}")
@require_capability("sentinel.ropa.manage")
async def api_security_update(request: Request, sid: int):
    data = await _json_body(request)
    ds.update_security(sid, data)
    return JSONResponse(ds.get_security(sid))


@router.delete("/api/security/{sid}")
@require_capability("sentinel.ropa.manage")
async def api_security_delete(request: Request, sid: int):
    ds.delete_security(sid)
    return JSONResponse({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# POLICIES
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/policies")
@require_capability("module.sentinel.access")
async def api_policies_list(request: Request):
    return JSONResponse(ds.list_policies(
        search=request.query_params.get("q"),
        status=request.query_params.get("status"),
        policy_type=request.query_params.get("type"),
    ))


@router.post("/api/policies", status_code=201)
@require_capability("sentinel.ropa.manage")
async def api_policies_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_policy(data)
    return JSONResponse(ds.get_policy(new_id), status_code=201)


@router.get("/api/policies/{policy_id}")
@require_capability("module.sentinel.access")
async def api_policies_get(request: Request, policy_id: int):
    p = ds.get_policy(policy_id)
    if not p:
        raise HTTPException(404, "Policy not found")
    return JSONResponse(p)


@router.put("/api/policies/{policy_id}")
@require_capability("sentinel.ropa.manage")
async def api_policies_update(request: Request, policy_id: int):
    data = await _json_body(request)
    ds.update_policy(policy_id, data)
    return JSONResponse(ds.get_policy(policy_id))


@router.delete("/api/policies/{policy_id}")
@require_capability("sentinel.ropa.manage")
async def api_policies_delete(request: Request, policy_id: int):
    ds.delete_policy(policy_id)
    return JSONResponse({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# TRAINING
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/training")
@require_capability("module.sentinel.access")
async def api_training_list(request: Request):
    return JSONResponse(ds.list_training(
        search=request.query_params.get("q"),
        department=request.query_params.get("department"),
    ))


@router.post("/api/training", status_code=201)
@require_capability("sentinel.ropa.manage")
async def api_training_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_training(data)
    return JSONResponse(ds.get_training(new_id), status_code=201)


@router.get("/api/training/{training_id}")
@require_capability("module.sentinel.access")
async def api_training_get(request: Request, training_id: int):
    t = ds.get_training(training_id)
    if not t:
        raise HTTPException(404, "Training record not found")
    return JSONResponse(t)


@router.put("/api/training/{training_id}")
@require_capability("sentinel.ropa.manage")
async def api_training_update(request: Request, training_id: int):
    data = await _json_body(request)
    ds.update_training(training_id, data)
    return JSONResponse(ds.get_training(training_id))


@router.delete("/api/training/{training_id}")
@require_capability("sentinel.ropa.manage")
async def api_training_delete(request: Request, training_id: int):
    ds.delete_training(training_id)
    return JSONResponse({"ok": True})


@router.post("/api/training/analyse-excel")
@require_capability("sentinel.ropa.manage")
async def api_training_analyse_excel(request: Request, file: UploadFile = File(...)):
    """Accept Excel upload, parse training attendance, bulk-import, return analytics."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        raise HTTPException(400, "Please upload an Excel file (.xlsx / .xls)")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Could not read Excel file: {e}")

    ws = wb.active

    # Auto-detect header row
    headers = []
    header_row_idx = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if any(c for c in row):
            headers = [str(c).strip() if c is not None else '' for c in row]
            header_row_idx = row_idx
            break

    if not headers:
        raise HTTPException(400, "Excel file appears to be empty")

    # Normalize headers
    def norm(s):
        return s.lower().replace(' ', '').replace('_', '').replace('-', '').replace('/', '')

    known = {
        'staffname': 'staff_name', 'name': 'staff_name', 'employee': 'staff_name',
        'employeename': 'staff_name', 'staffemail': 'staff_email', 'email': 'staff_email',
        'emailaddress': 'staff_email', 'department': 'department', 'dept': 'department',
        'division': 'department', 'training': 'training_name', 'trainingname': 'training_name',
        'course': 'training_name', 'coursename': 'training_name', 'module': 'training_name',
        'trainingtype': 'training_type', 'type': 'training_type',
        'completiondate': 'completion_date', 'dateofcompletion': 'completion_date',
        'completeddate': 'completion_date', 'date': 'completion_date',
        'expirydate': 'expiry_date', 'expiresdate': 'expiry_date', 'expiry': 'expiry_date',
        'result': 'passed', 'passed': 'passed', 'status': 'passed',
        'score': 'score', 'mark': 'score', 'grade': 'score',
        'certificate': 'certificate_no', 'certificateno': 'certificate_no',
        'certnumber': 'certificate_no', 'trainer': 'trainer', 'provider': 'trainer',
        'trainingprovider': 'trainer', 'regulation': 'regulation',
    }
    header_map = {}
    for i, h in enumerate(headers):
        mapped = known.get(norm(h))
        if mapped:
            header_map[i] = mapped

    # Parse rows
    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(c for c in row):
            continue
        rec = {}
        for col_i, val in enumerate(row):
            field = header_map.get(col_i)
            if field and val is not None:
                rec[field] = str(val).strip() if not isinstance(val, (int, float)) else val
        if rec.get('staff_name') or rec.get('training_name'):
            p = str(rec.get('passed', '')).lower()
            rec['passed'] = 1 if p in ('yes', 'passed', 'true', '1', 'complete', 'completed') else 0
            records.append(rec)

    if not records:
        raise HTTPException(400, "No data rows found")

    # Analytics
    from collections import defaultdict
    sessions = defaultdict(list)
    all_staff = set()
    for r in records:
        sn = r.get('training_name', 'Unknown Training')
        staff = r.get('staff_name', 'Unknown')
        sessions[sn].append(r)
        all_staff.add(staff)

    total_staff = len(all_staff)
    session_summary = []
    for session_name, rows in sessions.items():
        attendees = [r.get('staff_name') for r in rows if r.get('passed') or r.get('completion_date')]
        passed_count = sum(1 for r in rows if r.get('passed'))
        attendance_pct = round(len(attendees) / total_staff * 100) if total_staff else 0
        pass_rate = round(passed_count / len(rows) * 100) if rows else 0
        absent = sorted(all_staff - set(attendees))
        session_summary.append({
            "session": session_name, "total_registered": len(rows),
            "attended": len(attendees), "passed": passed_count,
            "absent_count": len(absent), "attendance_pct": attendance_pct,
            "pass_rate": pass_rate, "absent_staff": absent[:20],
        })

    staff_sessions = defaultdict(lambda: {"attended": 0, "passed": 0, "sessions": []})
    for r in records:
        s = r.get('staff_name', 'Unknown')
        staff_sessions[s]["attended"] += 1 if (r.get('passed') or r.get('completion_date')) else 0
        staff_sessions[s]["passed"] += 1 if r.get('passed') else 0
        staff_sessions[s]["sessions"].append(r.get('training_name', ''))

    total_sessions = len(sessions)
    staff_summary = []
    for name, data in sorted(staff_sessions.items()):
        att = data['attended']
        overall_pct = round(att / total_sessions * 100) if total_sessions else 0
        missed = [s for s in sessions if s not in data['sessions']]
        staff_summary.append({
            "name": name, "attended": att, "passed": data['passed'],
            "total_sessions": total_sessions, "attendance_pct": overall_pct,
            "missed_sessions": missed,
        })

    total_possible = len(records)
    total_attended = sum(1 for r in records if r.get('passed') or r.get('completion_date'))
    total_passed = sum(1 for r in records if r.get('passed'))
    overall_attendance = round(total_attended / total_possible * 100) if total_possible else 0
    overall_pass = round(total_passed / total_possible * 100) if total_possible else 0

    # Bulk import
    imported = skipped = 0
    for r in records:
        try:
            ds.create_training(r)
            imported += 1
        except Exception:
            skipped += 1

    return JSONResponse({
        "ok": True,
        "summary": {
            "total_staff": total_staff, "total_sessions": total_sessions,
            "total_records": len(records), "overall_attendance_pct": overall_attendance,
            "overall_pass_rate": overall_pass, "imported": imported,
            "skipped_duplicates": skipped,
        },
        "sessions": sorted(session_summary, key=lambda x: x['attendance_pct']),
        "staff": sorted(staff_summary, key=lambda x: x['attendance_pct']),
        "headers_detected": [h for h in headers if h],
        "mapped_columns": {headers[i]: v for i, v in header_map.items()},
    })


# ═════════════════════════════════════════════════════════════════════════════
# DATA FLOWS
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/dataflows")
@require_capability("module.sentinel.access")
async def api_dataflows_list(request: Request):
    return JSONResponse(ds.list_dataflows())


@router.get("/api/dataflows/{flow_id}")
@require_capability("module.sentinel.access")
async def api_dataflows_detail(request: Request, flow_id: int):
    row = ds.get_dataflow(flow_id)
    if not row:
        raise HTTPException(404, "Data flow not found")
    return JSONResponse(row)


@router.post("/api/dataflows", status_code=201)
@require_capability("sentinel.ropa.manage")
async def api_dataflows_create(request: Request):
    data = await _json_body(request)
    new_id = ds.create_dataflow(data)
    return JSONResponse({"id": new_id}, status_code=201)


@router.put("/api/dataflows/{flow_id}")
@require_capability("sentinel.ropa.manage")
async def api_dataflows_update(request: Request, flow_id: int):
    data = await _json_body(request)
    ds.update_dataflow(flow_id, data)
    return JSONResponse({"ok": True})


@router.delete("/api/dataflows/{flow_id}")
@require_capability("sentinel.ropa.manage")
async def api_dataflows_delete(request: Request, flow_id: int):
    ds.delete_dataflow(flow_id)
    return JSONResponse({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# SETTINGS
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/settings")
@require_capability("module.sentinel.access")
async def api_settings_get(request: Request):
    return JSONResponse(ds.get_all_settings())


@router.put("/api/settings")
@require_capability("sentinel.ropa.manage")
async def api_settings_update(request: Request):
    data = await _json_body(request)
    for k, v in data.items():
        ds.set_setting(k, v)
    return JSONResponse({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# AUDIT LOG
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/audit")
@require_capability("module.sentinel.access")
async def api_audit_list(request: Request):
    limit = int(request.query_params.get("limit", "200"))
    return JSONResponse(ds.list_audit(limit=limit))


# ═════════════════════════════════════════════════════════════════════════════
# CALENDAR & COMPLIANCE SCORE
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/compliance-score")
@require_capability("module.sentinel.access")
async def api_compliance_score(request: Request):
    return JSONResponse(ds.get_compliance_score())


# ═════════════════════════════════════════════════════════════════════════════
# LEGAL BASES (per regulation)
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/legal-bases/{regulation}")
@require_capability("module.sentinel.access")
async def api_legal_bases(request: Request, regulation: str):
    try:
        from modules.sentinel.ai_service import LEGAL_BASES
        bases = LEGAL_BASES.get(regulation, LEGAL_BASES.get("GDPR", []))
    except ImportError:
        bases = []
    # Return both formats: list for backward compat, dict for new jurisdiction-aware UI
    return JSONResponse({"regulation": regulation, "bases": bases})


# ═════════════════════════════════════════════════════════════════════════════
# AUDIT EVIDENCE PACK (ZIP export)
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/audit-export")
@require_capability("module.sentinel.access")
async def api_audit_export(request: Request):
    """Export ZIP: RoPA, DPIA summaries, breach log, DSR log, vendor list, audit trail."""
    settings = ds.get_all_settings() or {}
    org = settings.get("org_name", "Organisation").replace(" ", "_")
    date_str = utcnow().strftime("%Y%m%d")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # DPIA JSON summary
        dpias = ds.list_dpias()
        zf.writestr(f"DPIA_Summary_{date_str}.json", json.dumps(dpias, indent=2, default=str))

        # Breach log
        breaches = ds.list_breaches()
        zf.writestr(f"Breach_Log_{date_str}.json", json.dumps(breaches, indent=2, default=str))

        # DSR log
        dsrs = ds.list_dsrs()
        zf.writestr(f"DSR_Log_{date_str}.json", json.dumps(dsrs, indent=2, default=str))

        # Vendor register
        vendors = ds.list_vendors()
        zf.writestr(f"Vendor_Register_{date_str}.json", json.dumps(vendors, indent=2, default=str))

        # Audit trail
        audit = ds.list_audit(limit=10000)
        zf.writestr(f"Audit_Trail_{date_str}.json", json.dumps(audit, indent=2, default=str))

        # Policies
        policies = ds.list_policies()
        zf.writestr(f"Policy_Register_{date_str}.json", json.dumps(policies, indent=2, default=str))

        # Compliance score
        score = ds.get_compliance_score() or {"overall": 0, "grade": "N/A", "breakdown": {}}
        zf.writestr(f"Compliance_Score_{date_str}.json", json.dumps(score, indent=2, default=str))

        # README
        readme = f"""DATA PROTECTION SENTINEL — AUDIT EVIDENCE PACK
Generated: {utcnow().strftime('%d %B %Y %H:%M UTC')}
Organisation: {settings.get('org_name', '—')}
DPO: {settings.get('dpo_name', '—')} ({settings.get('dpo_email', '—')})
Primary Regulation: {settings.get('primary_regulation', '—')}
Overall Compliance Score: {score.get('overall', '—')}% (Grade {score.get('grade', '—')})

Generated by One For All — Data Protection Sentinel Module
"""
        zf.writestr("README.txt", readme)

    zip_buf.seek(0)
    filename = f"AuditPack_{org}_{date_str}.zip"
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═════════════════════════════════════════════════════════════════════════════
# AI ENDPOINTS
# ═════════════════════════════════════════════════════════════════════════════

# ═════════════════════════════════════════════════════════════════════════════
# LEGITIMATE INTEREST ASSESSMENTS (SENT-14)
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/lia")
@require_capability("module.sentinel.access")
async def api_lia_list(request: Request):
    return JSONResponse(ds.list_lia(
        ropa_id=int(request.query_params["ropa_id"]) if request.query_params.get("ropa_id") else None,
        result=request.query_params.get("result"),
    ))


@router.post("/api/lia", status_code=201)
@require_capability("sentinel.ropa.manage")
async def api_lia_create(request: Request):
    data = await _json_body(request)
    data["created_by"] = request.state.user["id"]
    new_id = ds.create_lia(data)
    return JSONResponse(ds.get_lia(new_id), status_code=201)


@router.get("/api/lia/{lia_id}")
@require_capability("module.sentinel.access")
async def api_lia_get(request: Request, lia_id: int):
    r = ds.get_lia(lia_id)
    if not r:
        raise HTTPException(404, "LIA not found")
    return JSONResponse(r)


@router.put("/api/lia/{lia_id}")
@require_capability("sentinel.ropa.manage")
async def api_lia_update(request: Request, lia_id: int):
    data = await _json_body(request)
    ds.update_lia(lia_id, data)
    return JSONResponse(ds.get_lia(lia_id))


@router.delete("/api/lia/{lia_id}")
@require_capability("sentinel.ropa.manage")
async def api_lia_delete(request: Request, lia_id: int):
    ds.delete_lia(lia_id)
    return JSONResponse({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# Jurisdiction Management
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/api/jurisdictions/rules")
@require_module("sentinel")
async def api_jurisdiction_rules(request: Request):
    """All known jurisdictions from the registry, grouped by region."""
    from modules.sentinel.jurisdictions import list_jurisdictions, jurisdictions_by_region
    return JSONResponse({
        "items": list_jurisdictions(),
        "by_region": jurisdictions_by_region(),
    })


@router.get("/api/jurisdictions/active")
@require_module("sentinel")
async def api_active_jurisdictions(request: Request):
    """Org's currently active jurisdiction configs."""
    return JSONResponse({"items": ds.get_active_jurisdictions()})


@router.post("/api/jurisdictions/active")
@require_capability("module.sentinel.access")
async def api_set_active_jurisdictions(request: Request):
    """Bulk update active jurisdictions. Body: {keys:[...], primary:'GDPR'}"""
    body = await _json_body(request)
    keys = body.get("keys", [])
    primary = body.get("primary")

    from modules.sentinel.jurisdictions import JURISDICTION_RULES
    valid_keys = [k for k in keys if k in JURISDICTION_RULES]

    existing = {j["jurisdiction_key"] for j in ds.get_all_jurisdiction_configs()}
    for key in existing:
        if key not in valid_keys:
            ds.deactivate_jurisdiction(key)
    for key in valid_keys:
        ds.activate_jurisdiction(key, is_primary=(key == primary))

    return JSONResponse({"ok": True, "active": len(valid_keys)})


@router.get("/api/jurisdictions/{key}")
@require_module("sentinel")
async def api_jurisdiction_detail(request: Request, key: str):
    """Single jurisdiction: registry rules + org config + record counts."""
    from modules.sentinel.jurisdictions import get_jurisdiction
    rules = get_jurisdiction(key)
    if not rules:
        raise HTTPException(status_code=404, detail="Unknown jurisdiction")
    configs = ds.get_all_jurisdiction_configs()
    config = next((c for c in configs if c.get("jurisdiction_key") == key), None)
    stats_list = ds.get_jurisdiction_stats()
    stat = next((s for s in stats_list if s["regulation"] == key),
                {"ropa": 0, "breaches": 0, "dsrs": 0, "dpias": 0})
    return JSONResponse({"key": key, "rules": rules, "config": config, "stats": stat})


@router.put("/api/jurisdictions/{key}/config")
@require_capability("module.sentinel.access")
async def api_update_jurisdiction_config(request: Request, key: str):
    """Update org-specific config for a jurisdiction (DPO, reg number, notes)."""
    from modules.sentinel.jurisdictions import get_jurisdiction
    if not get_jurisdiction(key):
        raise HTTPException(status_code=404, detail="Unknown jurisdiction")
    body = await _json_body(request)
    ds.update_jurisdiction_config(key, body)
    return JSONResponse({"ok": True})


@router.delete("/api/jurisdictions/{key}")
@require_capability("module.sentinel.access")
async def api_deactivate_jurisdiction(request: Request, key: str):
    """Deactivate a jurisdiction (keeps config row)."""
    ds.deactivate_jurisdiction(key)
    return JSONResponse({"ok": True})


@router.get("/api/jurisdiction-stats")
@require_module("sentinel")
async def api_jurisdiction_stats(request: Request):
    """Per-jurisdiction record counts across all sentinel entity types."""
    return JSONResponse({"items": ds.get_jurisdiction_stats()})


@router.post("/api/ai/research")
@require_capability("sentinel.ai.assess")
async def api_ai_research(request: Request):
    body = await _json_body(request)
    activity = body.get("activity_type", "")
    regulation = body.get("regulation", "GDPR")
    context = body.get("context", "")
    if not activity:
        raise HTTPException(400, "activity_type is required")
    from modules.sentinel.ai_service import ai_research
    text, err = await ai_research(activity, regulation, context)
    if err:
        raise HTTPException(500, err)
    return JSONResponse({"research": text})


@router.post("/api/ai/generate/{dpia_id}")
@require_capability("sentinel.ai.assess")
async def api_ai_generate(request: Request, dpia_id: int):
    dpia = ds.get_dpia(dpia_id)
    if not dpia:
        raise HTTPException(404, "DPIA not found")
    from modules.sentinel.ai_service import ai_generate_full_dpia
    text, err = await ai_generate_full_dpia(dpia)
    if err:
        raise HTTPException(500, err)
    ds.update_dpia(dpia_id, {"ai_full_dpia": text})
    return JSONResponse({"content": text})


@router.post("/api/ai/risks")
@require_capability("sentinel.ai.assess")
async def api_ai_risks(request: Request):
    body = await _json_body(request)
    activity = body.get("activity_type", "")
    regulation = body.get("regulation", "GDPR")
    categories = body.get("data_categories", [])
    if not activity:
        raise HTTPException(400, "activity_type is required")
    from modules.sentinel.ai_service import ai_suggest_risks
    risks, err = await ai_suggest_risks(activity, regulation, categories)
    if err:
        raise HTTPException(500, err)
    return JSONResponse({"risks": risks})


@router.post("/api/ai/score-ropa/{ropa_id}")
@require_capability("sentinel.ai.assess")
async def api_ai_score_ropa(request: Request, ropa_id: int):
    entry = ds.get_ropa(ropa_id)
    if not entry:
        raise HTTPException(404, "RoPA entry not found")
    from modules.sentinel.ai_service import ai_score_ropa
    result, err = await ai_score_ropa(entry)
    if err:
        raise HTTPException(500, err)
    update_data = {
        "risk_score": result.get("risk_score", "medium"),
        "ai_risk_notes": result.get("rationale", ""),
    }
    if result.get("dpia_required"):
        update_data["dpia_required"] = 1
    ds.update_ropa(ropa_id, update_data)
    return JSONResponse(result)


@router.post("/api/ai/breach-impact/{breach_id}")
@require_capability("sentinel.ai.assess")
async def api_ai_breach_impact(request: Request, breach_id: int):
    breach = ds.get_breach(breach_id)
    if not breach:
        raise HTTPException(404, "Breach not found")
    from modules.sentinel.ai_service import ai_assess_breach
    text, err = await ai_assess_breach(breach)
    if err:
        raise HTTPException(500, err)
    ds.update_breach(breach_id, {"ai_assessment": text})
    return JSONResponse({"assessment": text})


@router.post("/api/ai/dsr-draft/{dsr_id}")
@require_capability("sentinel.ai.assess")
async def api_ai_dsr_draft(request: Request, dsr_id: int):
    dsr = ds.get_dsr(dsr_id)
    if not dsr:
        raise HTTPException(404, "DSR not found")
    from modules.sentinel.ai_service import ai_draft_dsr_response
    text, err = await ai_draft_dsr_response(dsr)
    if err:
        raise HTTPException(500, err)
    ds.update_dsr(dsr_id, {"ai_draft": text})
    return JSONResponse({"draft": text})


@router.post("/api/ai/privacy-notice")
@require_capability("sentinel.ai.assess")
async def api_ai_privacy_notice(request: Request):
    body = await _json_body(request)
    settings = ds.get_all_settings()
    body.setdefault("org_name", settings.get("org_name", ""))
    body.setdefault("dpo_name", settings.get("dpo_name", ""))
    body.setdefault("dpo_email", settings.get("dpo_email", ""))
    from modules.sentinel.ai_service import ai_generate_privacy_notice
    text, err = await ai_generate_privacy_notice(body)
    if err:
        raise HTTPException(500, err)
    return JSONResponse({"notice": text})


@router.post("/api/ai/vendor-check/{vendor_id}")
@require_capability("sentinel.ai.assess")
async def api_ai_vendor_check(request: Request, vendor_id: int):
    vendor = ds.get_vendor(vendor_id)
    if not vendor:
        raise HTTPException(404, "Vendor not found")
    from modules.sentinel.ai_service import ai_vendor_check
    text, err = await ai_vendor_check(vendor)
    if err:
        raise HTTPException(500, err)
    ds.update_vendor(vendor_id, {"ai_assessment": text})
    return JSONResponse({"assessment": text})


@router.post("/api/ai/chat")
@require_capability("sentinel.ai.assess")
async def api_ai_chat(request: Request):
    body = await _json_body(request)
    message = body.get("message", "")
    regulation = body.get("regulation")
    history = body.get("history", [])
    if not message:
        raise HTTPException(400, "message is required")
    from modules.sentinel.ai_service import ai_chat
    text, err = await ai_chat(message, regulation=regulation, history=history)
    if err:
        raise HTTPException(500, err)
    return JSONResponse({"response": text})


@router.post("/api/ai/gap-analysis")
@require_capability("sentinel.ai.assess")
async def api_ai_gap_analysis(request: Request):
    body = await _json_body(request)
    reg_from = body.get("regulation_from", "GDPR")
    reg_to = body.get("regulation_to", "South Africa POPIA")
    activities = body.get("activities", "")
    from modules.sentinel.ai_service import ai_gap_analysis
    text, err = await ai_gap_analysis(reg_from, reg_to, activities)
    if err:
        raise HTTPException(500, err)
    return JSONResponse({"analysis": text})


@router.post("/api/ai/generate-policy")
@require_capability("sentinel.ai.assess")
async def api_ai_generate_policy(request: Request):
    """Draft a data protection policy document using AI."""
    body = await _json_body(request)
    policy_type = body.get("policy_type", "Data Protection Policy")
    topic = body.get("topic", "")
    regulation = body.get("regulation", "GDPR")

    topic_clause = f" with a focus on {topic}" if topic else ""
    user_msg = (
        f"Draft a complete {policy_type}{topic_clause} for compliance with {regulation}. "
        f"Structure it with: Purpose, Scope, Policy Statement, Responsibilities, "
        f"Procedures, Review, and Definitions sections. "
        f"Return only the policy document text, ready to use."
    )
    from modules.sentinel.ai_service import ai_chat
    text, err = await ai_chat(user_msg, regulation=regulation)
    if err:
        raise HTTPException(500, err)
    return JSONResponse({"content": text})
