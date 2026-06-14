from fastapi import FastAPI, Request, Form, Depends, HTTPException, Response
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import sqlite3, hashlib, json, io, os, secrets
from passlib.hash import bcrypt as _bcrypt
from datetime import datetime, timedelta
from typing import Optional
from database import init_db, DB_PATH, FRAMEWORKS
from ai_generator import generate_policy, generate_gap_analysis, get_api_key
from ask_service import (
    ask as ask_policy,
    init_index as init_ask_index,
    rebuild_all as rebuild_ask_index,
    reindex_document, reindex_control, reindex_risk,
    remove_from_index,
)
import roles as role_lib
from roles import (
    has_capability, has_role, can_edit_control, can_edit_risk,
    can_approve_policy, ROLE_LABELS, ALL_ROLES,
)
from dotenv import load_dotenv
load_dotenv()

app = FastAPI(title="ARIA", version="1.0.0")
app.mount("/static", StaticFiles(directory="static"), name="static")

# Paths allowed while a user has must_change_password=1 set.
_PW_CHANGE_ALLOWLIST = {"/change-password", "/logout", "/login"}

@app.middleware("http")
async def enforce_password_change(request: Request, call_next):
    """If the logged-in user must change password, funnel every page and
    non-ask API call to /change-password until they do."""
    uid = request.cookies.get("user_id")
    path = request.url.path
    if uid and path not in _PW_CHANGE_ALLOWLIST \
            and not path.startswith("/static") \
            and not path.startswith("/api/change-password"):
        try:
            import sqlite3 as _s
            _c = _s.connect(DB_PATH)
            row = _c.execute("SELECT must_change_password FROM users WHERE id=?", (int(uid),)).fetchone()
            _c.close()
            if row and row[0]:
                # For JSON endpoints return 403; for pages redirect
                if path.startswith("/api/"):
                    return JSONResponse({"error": "Password change required before continuing."}, 403)
                return RedirectResponse("/change-password", status_code=302)
        except Exception:
            pass
    return await call_next(request)
templates = Jinja2Templates(directory="templates")
templates.env.filters["tojson"] = lambda v: json.dumps(dict(v) if hasattr(v, 'keys') else v, default=str)
# Expose role/capability helpers to all templates
templates.env.globals["has_capability"] = has_capability
templates.env.globals["has_role"] = has_role
templates.env.globals["ROLE_LABELS"] = ROLE_LABELS

# ── Auth helpers ─────────────────────────────────────────────────────────────
# Passwords are hashed with bcrypt (12 rounds). Legacy accounts that still
# have unsalted SHA-256 digests are accepted transparently for one final
# login, at which point the hash is upgraded to bcrypt in place.
_BCRYPT_ROUNDS = 12


def hash_password(p: str) -> str:
    """Hash a plaintext password with bcrypt."""
    return _bcrypt.using(rounds=_BCRYPT_ROUNDS).hash(p)


def _is_legacy_sha256(stored: str) -> bool:
    """A legacy hash is the 64-char hex digest of sha256(plain)."""
    if not stored or len(stored) != 64:
        return False
    try:
        int(stored, 16)
        return True
    except ValueError:
        return False


def verify_password(plain: str, stored: str) -> bool:
    """Verify a password against either a bcrypt hash or a legacy SHA-256 hex digest."""
    if not stored:
        return False
    if stored.startswith("$2"):                          # bcrypt
        try:
            return _bcrypt.verify(plain, stored)
        except Exception:
            return False
    if _is_legacy_sha256(stored):                        # legacy
        return hashlib.sha256(plain.encode("utf-8")).hexdigest() == stored
    return False


def _upgrade_password_hash_if_legacy(user_id: int, plain: str, stored: str) -> None:
    """If the stored hash is legacy SHA-256, rehash with bcrypt and persist."""
    if _is_legacy_sha256(stored):
        try:
            conn = get_db()
            conn.execute(
                "UPDATE users SET password_hash=? WHERE id=?",
                (hash_password(plain), int(user_id))
            )
            conn.commit()
            conn.close()
        except Exception:
            pass

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def _load_user_roles(uid: int | str) -> list[str]:
    """Fetch the additive role list for a user. Safe if user_roles is empty."""
    try:
        conn = get_db()
        rows = conn.execute(
            "SELECT role_key FROM user_roles WHERE user_id=? ORDER BY role_key",
            (int(uid),)
        ).fetchall()
        conn.close()
        return [r[0] for r in rows]
    except Exception:
        return []


def get_current_user(request: Request):
    uid = request.cookies.get("user_id")
    role = request.cookies.get("user_role")      # legacy single-role cookie
    name = request.cookies.get("user_name")
    uname = request.cookies.get("username")
    if not uid:
        return None
    roles_list = _load_user_roles(uid)
    # Fall back to legacy role if user_roles row is missing (fresh install quirk)
    if not roles_list and role:
        roles_list = role_lib.migrate_legacy_role(role)
    return {
        "id": uid,
        "role": role,              # kept for any legacy template reference
        "roles": roles_list,
        "full_name": name,
        "username": uname,
    }


def require_user(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=302, headers={"Location": "/login"})
    return user


def require_capability(request: Request, capability: str):
    """Guard helper for routes. Returns the user if the capability is held,
    otherwise raises a 302 back to the dashboard (or 403 for API calls)."""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=302, headers={"Location": "/login"})
    if not has_capability(user, capability):
        raise HTTPException(status_code=403, detail=f"Missing capability: {capability}")
    return user


def require_admin(request: Request):
    """Legacy shim — now backed by the manage_users capability."""
    user = get_current_user(request)
    if not user or not has_capability(user, "manage_users"):
        raise HTTPException(status_code=302, headers={"Location": "/dashboard"})
    return user

def log_action(user, action, entity_type="", entity_id="", old_val="", new_val=""):
    try:
        conn = get_db()
        conn.execute("""INSERT INTO audit_log (user_id, username, action, entity_type, entity_id, old_value, new_value)
                        VALUES (?, ?, ?, ?, ?, ?, ?)""",
                     (user["id"], user["username"], action, entity_type, entity_id, old_val, new_val))
        conn.commit()
        conn.close()
    except: pass

# ── Routes ────────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    user = get_current_user(request)
    if user:
        return RedirectResponse("/dashboard")
    return RedirectResponse("/login")

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    # Constant-ish-time path: always run verify_password so a missing username
    # doesn't return measurably faster than a wrong password.
    stored = user["password_hash"] if user else "$2b$12$" + "x" * 53
    ok = verify_password(password, stored)
    if user and ok and not user["active"]:
        conn.close()
        return templates.TemplateResponse("login.html",
            {"request": request, "error": "This account has been deactivated. Contact an administrator."})
    if user and ok:
        conn.execute("UPDATE users SET last_login=? WHERE id=?",
                     (datetime.now().isoformat(), user["id"]))
        conn.commit()
    conn.close()
    if not (user and ok):
        return templates.TemplateResponse("login.html", {"request": request, "error": "Invalid credentials"})
    # Transparently upgrade legacy SHA-256 hashes to bcrypt on first successful login.
    _upgrade_password_hash_if_legacy(user["id"], password, user["password_hash"])
    # Forced password change takes precedence over dashboard redirect
    next_path = "/change-password" if (user["must_change_password"] or 0) else "/dashboard"
    response = RedirectResponse(next_path, status_code=302)
    response.set_cookie("user_id", str(user["id"]), httponly=True, max_age=86400)
    response.set_cookie("user_role", user["role"], httponly=True, max_age=86400)
    response.set_cookie("user_name", user["full_name"], httponly=True, max_age=86400)
    response.set_cookie("username", user["username"], httponly=True, max_age=86400)
    return response

@app.get("/logout")
async def logout():
    response = RedirectResponse("/login")
    for c in ["user_id", "user_role", "user_name", "username"]:
        response.delete_cookie(c)
    return response

def _must_change_pw(uid) -> bool:
    try:
        conn = get_db()
        row = conn.execute("SELECT must_change_password FROM users WHERE id=?", (int(uid),)).fetchone()
        conn.close()
        return bool(row and row[0])
    except Exception:
        return False

@app.get("/change-password", response_class=HTMLResponse)
async def change_password_page(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    return templates.TemplateResponse("change_password.html",
        {"request": request, "user": user, "forced": _must_change_pw(user["id"])})

@app.post("/change-password")
async def change_password(request: Request,
                           current_password: str = Form(""),
                           new_password: str = Form(...),
                           confirm_password: str = Form(...)):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    forced = _must_change_pw(user["id"])
    ctx = {"request": request, "user": user, "forced": forced}

    if new_password != confirm_password:
        ctx["error"] = "The two new password fields don't match."
        return templates.TemplateResponse("change_password.html", ctx)
    if len(new_password) < 8:
        ctx["error"] = "New password must be at least 8 characters."
        return templates.TemplateResponse("change_password.html", ctx)

    conn = get_db()
    row = conn.execute("SELECT password_hash FROM users WHERE id=?", (int(user["id"]),)).fetchone()
    # Current password is required unless this is a forced change right after an admin reset.
    if not forced:
        if not current_password or not verify_password(current_password, row["password_hash"]):
            conn.close()
            ctx["error"] = "Current password is incorrect."
            return templates.TemplateResponse("change_password.html", ctx)

    conn.execute(
        "UPDATE users SET password_hash=?, must_change_password=0 WHERE id=?",
        (hash_password(new_password), int(user["id"]))
    )
    conn.commit()
    conn.close()
    log_action(user, "Changed own password", "user", str(user["id"]))
    return RedirectResponse("/dashboard", status_code=302)

@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    conn = get_db()
    frameworks = conn.execute("SELECT * FROM frameworks ORDER BY name").fetchall()
    fw_stats = []
    for fw in frameworks:
        stats = conn.execute("""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN status='Implemented' THEN 1 ELSE 0 END) as implemented,
                SUM(CASE WHEN status='In Progress' THEN 1 ELSE 0 END) as in_progress,
                SUM(CASE WHEN status='Not Started' THEN 1 ELSE 0 END) as not_started,
                SUM(CASE WHEN status='Approved' THEN 1 ELSE 0 END) as approved,
                SUM(CASE WHEN status='Draft' THEN 1 ELSE 0 END) as draft,
                SUM(CASE WHEN status='Under Review' THEN 1 ELSE 0 END) as under_review,
                SUM(CASE WHEN status='Needs Update' THEN 1 ELSE 0 END) as needs_update
            FROM controls WHERE framework_id=?""", (fw["id"],)).fetchone()
        pct = round((stats["implemented"] / stats["total"] * 100) if stats["total"] > 0 else 0, 1)
        fw_stats.append({
            "id": fw["id"], "name": fw["name"], "color": fw["color"],
            "description": fw["description"],
            "total": stats["total"], "implemented": stats["implemented"],
            "in_progress": stats["in_progress"], "not_started": stats["not_started"],
            "approved": stats["approved"], "draft": stats["draft"],
            "under_review": stats["under_review"], "needs_update": stats["needs_update"],
            "pct": pct
        })
    totals = {
        "total": sum(f["total"] for f in fw_stats),
        "implemented": sum(f["implemented"] for f in fw_stats),
        "in_progress": sum(f["in_progress"] for f in fw_stats),
        "not_started": sum(f["not_started"] for f in fw_stats),
        "overall_pct": round(sum(f["implemented"] for f in fw_stats) / max(sum(f["total"] for f in fw_stats), 1) * 100, 1)
    }
    recent = conn.execute("""SELECT al.*, u.full_name FROM audit_log al
                             LEFT JOIN users u ON al.user_id=u.id
                             ORDER BY al.timestamp DESC LIMIT 8""").fetchall()
    conn.close()
    return templates.TemplateResponse("dashboard.html", {
        "request": request, "user": user,
        "fw_stats": fw_stats,
        "fw_stats_json": json.dumps(fw_stats),
        "totals": totals,
        "totals_json": json.dumps(totals),
        "recent": recent
    })

@app.get("/framework/{fw_id}", response_class=HTMLResponse)
async def framework_detail(request: Request, fw_id: int,
                            status: str = "", priority: str = "",
                            category: str = "", search: str = ""):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    conn = get_db()
    fw = conn.execute("SELECT * FROM frameworks WHERE id=?", (fw_id,)).fetchone()
    if not fw: raise HTTPException(404)

    q = "SELECT * FROM controls WHERE framework_id=?"
    params = [fw_id]
    if status:
        q += " AND status=?"; params.append(status)
    if priority:
        q += " AND priority=?"; params.append(priority)
    if category:
        q += " AND category=?"; params.append(category)
    if search:
        q += " AND (name LIKE ? OR ref LIKE ? OR description LIKE ?)";
        params += [f"%{search}%", f"%{search}%", f"%{search}%"]
    q += " ORDER BY ref"
    controls = conn.execute(q, params).fetchall()

    cats = conn.execute("SELECT DISTINCT category FROM controls WHERE framework_id=? ORDER BY category", (fw_id,)).fetchall()
    stats = conn.execute("""SELECT
        COUNT(*) as total,
        SUM(CASE WHEN status='Implemented' THEN 1 ELSE 0 END) as implemented,
        SUM(CASE WHEN status='Not Started' THEN 1 ELSE 0 END) as not_started,
        SUM(CASE WHEN status='In Progress' THEN 1 ELSE 0 END) as in_progress,
        SUM(CASE WHEN status='Approved' THEN 1 ELSE 0 END) as approved
        FROM controls WHERE framework_id=?""", (fw_id,)).fetchone()
    conn.close()
    return templates.TemplateResponse("framework.html", {
        "request": request, "user": user, "fw": fw,
        "controls": controls, "stats": stats, "categories": cats,
        "filters": {"status": status, "priority": priority, "category": category, "search": search}
    })

@app.post("/control/{ctrl_id}/update")
async def update_control(request: Request, ctrl_id: int,
                          status: str = Form(None), priority: str = Form(None),
                          owner: str = Form(None), target_date: str = Form(None),
                          review_date: str = Form(None), notes: str = Form(None),
                          document_title: str = Form(None), version: str = Form(None),
                          evidence_ref: str = Form(None)):
    user = get_current_user(request)
    if not user: return JSONResponse({"error": "Unauthorized"}, 401)

    conn = get_db()
    old = conn.execute("SELECT * FROM controls WHERE id=?", (ctrl_id,)).fetchone()
    if not can_edit_control(user, dict(old) if old else None):
        conn.close()
        return JSONResponse({"error": "You do not have permission to update this control."}, 403)
    updates = []
    params = []
    for field, val in [("status", status), ("priority", priority), ("owner", owner),
                        ("target_date", target_date), ("review_date", review_date),
                        ("notes", notes), ("document_title", document_title),
                        ("version", version), ("evidence_ref", evidence_ref)]:
        if val is not None:
            updates.append(f"{field}=?")
            params.append(val)
    if updates:
        updates.append("last_updated=?")
        params.append(datetime.now().strftime("%Y-%m-%d"))
        params.append(ctrl_id)
        conn.execute(f"UPDATE controls SET {', '.join(updates)} WHERE id=?", params)
        conn.commit()
        log_action(user, f"Updated control status to '{status}'", "control", str(ctrl_id),
                   old["status"] if old else "", status or "")
    conn.close()
    try:
        reindex_control(ctrl_id)
    except Exception:
        pass
    return JSONResponse({"success": True})

@app.get("/api/stats")
async def api_stats(request: Request):
    user = get_current_user(request)
    if not user: return JSONResponse({"error": "Unauthorized"}, 401)
    conn = get_db()
    frameworks = conn.execute("SELECT * FROM frameworks").fetchall()
    data = []
    for fw in frameworks:
        stats = conn.execute("""SELECT
            SUM(CASE WHEN status='Implemented' THEN 1 ELSE 0 END) as implemented,
            SUM(CASE WHEN status='In Progress' THEN 1 ELSE 0 END) as in_progress,
            SUM(CASE WHEN status='Not Started' THEN 1 ELSE 0 END) as not_started,
            SUM(CASE WHEN status='Approved' THEN 1 ELSE 0 END) as approved,
            SUM(CASE WHEN status='Draft' THEN 1 ELSE 0 END) as draft,
            SUM(CASE WHEN status='Needs Update' THEN 1 ELSE 0 END) as needs_update,
            COUNT(*) as total
            FROM controls WHERE framework_id=?""", (fw["id"],)).fetchone()
        data.append({"name": fw["name"], "color": fw["color"],
                     "implemented": stats["implemented"], "in_progress": stats["in_progress"],
                     "not_started": stats["not_started"], "approved": stats["approved"],
                     "draft": stats["draft"], "needs_update": stats["needs_update"],
                     "total": stats["total"]})
    conn.close()
    return JSONResponse(data)

@app.get("/documents", response_class=HTMLResponse)
async def documents(request: Request,
                    framework: str = "", status: str = "",
                    doc_type: str = "", search: str = ""):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    conn = get_db()

    q = "SELECT * FROM documents WHERE 1=1"
    params = []
    if framework: q += " AND framework=?"; params.append(framework)
    if status:    q += " AND status=?";    params.append(status)
    if doc_type:  q += " AND doc_type=?";  params.append(doc_type)
    if search:    q += " AND (title LIKE ? OR control_ref LIKE ? OR owner LIKE ?)"; params += [f"%{search}%"]*3
    q += " ORDER BY framework, doc_id"

    docs = conn.execute(q, params).fetchall()
    frameworks = conn.execute("SELECT DISTINCT name FROM frameworks ORDER BY name").fetchall()

    # Stats
    total     = conn.execute("SELECT COUNT(*) FROM documents").fetchone()[0]
    approved  = conn.execute("SELECT COUNT(*) FROM documents WHERE status='Approved'").fetchone()[0]
    draft     = conn.execute("SELECT COUNT(*) FROM documents WHERE status='Draft'").fetchone()[0]
    ai_gen    = conn.execute("SELECT COUNT(*) FROM documents WHERE comments LIKE '%AI Generated%'").fetchone()[0]
    conn.close()

    return templates.TemplateResponse("documents.html", {
        "request": request, "user": user, "docs": docs,
        "frameworks": frameworks,
        "filters": {"framework": framework, "status": status, "doc_type": doc_type, "search": search},
        "stats": {"total": total, "approved": approved, "draft": draft, "ai_gen": ai_gen}
    })

@app.post("/documents/add")
async def add_document(request: Request,
                       framework: str = Form(...), control_ref: str = Form(""),
                       title: str = Form(...), doc_type: str = Form("Policy"),
                       version: str = Form("1.0"), status: str = Form("Draft"),
                       owner: str = Form(""), approver: str = Form(""),
                       effective_date: str = Form(""), review_date: str = Form(""),
                       location: str = Form(""), comments: str = Form("")):
    user = get_current_user(request)
    if not user: return JSONResponse({"error": "Unauthorized"}, 401)
    if not has_capability(user, "create_policy"):
        return JSONResponse({"error": "You need policy author or compliance manager role to add documents."}, 403)
    # Separation of duties: only admin/compliance_manager can create as 'Approved' directly.
    if status == "Approved" and not has_capability(user, "approve_policy"):
        status = "Draft"
    conn = get_db()
    count = conn.execute("SELECT COUNT(*) FROM documents").fetchone()[0]
    doc_id = f"DOC-{count+1:04d}"
    now = datetime.now().isoformat()
    conn.execute("""INSERT INTO documents
        (doc_id, framework, control_ref, title, doc_type, version, status,
         owner, approver, effective_date, review_date, location, comments, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (doc_id, framework, control_ref, title, doc_type, version, status,
         owner, approver, effective_date or None, review_date or None,
         location, comments, now, now))
    conn.commit()
    conn.close()
    log_action(user, f"Added document {doc_id}: {title}", "document", doc_id)
    try:
        reindex_document(doc_id)
    except Exception:
        pass
    return JSONResponse({"success": True, "doc_id": doc_id})

@app.post("/documents/update/{doc_id}")
async def update_document(request: Request, doc_id: str,
                          title: str = Form(None), status: str = Form(None),
                          version: str = Form(None), owner: str = Form(None),
                          approver: str = Form(None), effective_date: str = Form(None),
                          review_date: str = Form(None), location: str = Form(None),
                          comments: str = Form(None)):
    user = get_current_user(request)
    if not user: return JSONResponse({"error": "Unauthorized"}, 401)

    conn = get_db()
    existing = conn.execute("SELECT * FROM documents WHERE doc_id=?", (doc_id,)).fetchone()
    if not existing:
        conn.close()
        return JSONResponse({"error": "Document not found"}, 404)
    doc = dict(existing)

    # Who can edit this document's content?
    is_own = (doc.get("owner") or "").strip().lower() in (
        (user.get("full_name") or "").strip().lower(),
        (user.get("username") or "").strip().lower(),
    )
    can_edit = has_capability(user, "edit_any_policy") or (
        has_capability(user, "edit_own_policy") and is_own
    )
    if not can_edit:
        conn.close()
        return JSONResponse({"error": "You do not have permission to edit this document."}, 403)

    # If the caller is trying to flip status to Approved, they need approve_policy
    # capability AND cannot self-approve (unless admin/compliance_manager).
    if status == "Approved" and status != doc.get("status"):
        if not can_approve_policy(user, doc):
            conn.close()
            return JSONResponse({
                "error": "You cannot approve this document. "
                         "Policy authors cannot approve their own drafts."
            }, 403)

    updates, params = [], []
    for field, val in [("title", title), ("status", status), ("version", version),
                        ("owner", owner), ("approver", approver),
                        ("effective_date", effective_date), ("review_date", review_date),
                        ("location", location), ("comments", comments)]:
        if val is not None:
            updates.append(f"{field}=?"); params.append(val)
    if updates:
        updates.append("updated_at=?"); params.append(datetime.now().isoformat())
        params.append(doc_id)
        conn.execute(f"UPDATE documents SET {', '.join(updates)} WHERE doc_id=?", params)
        conn.commit()
        log_action(user, f"Updated document {doc_id}", "document", doc_id)
    conn.close()
    try:
        reindex_document(doc_id)
    except Exception:
        pass
    return JSONResponse({"success": True})

@app.post("/documents/delete/{doc_id}")
async def delete_document(request: Request, doc_id: str):
    user = get_current_user(request)
    if not user: return JSONResponse({"error": "Unauthorized"}, 401)
    if not has_capability(user, "delete_policy"):
        return JSONResponse({"error": "Only System Administrators can delete documents."}, 403)
    conn = get_db()
    conn.execute("DELETE FROM documents WHERE doc_id=?", (doc_id,))
    conn.commit()
    conn.close()
    log_action(user, f"Deleted document {doc_id}", "document", doc_id)
    try:
        remove_from_index("document", doc_id)
    except Exception:
        pass
    return JSONResponse({"success": True})

@app.get("/documents/{doc_id}", response_class=HTMLResponse)
async def document_detail(request: Request, doc_id: str):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    conn = get_db()
    doc = conn.execute("SELECT * FROM documents WHERE doc_id=?", (doc_id,)).fetchone()
    conn.close()
    if not doc: raise HTTPException(404)
    return templates.TemplateResponse("document_detail.html", {"request": request, "user": user, "doc": doc})

@app.get("/risks", response_class=HTMLResponse)
async def risks_page(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    conn = get_db()
    risks = conn.execute("SELECT * FROM risks ORDER BY (likelihood*impact) DESC").fetchall()
    conn.close()
    return templates.TemplateResponse("risks.html", {"request": request, "user": user, "risks": risks})

@app.post("/risks/add")
async def add_risk(request: Request,
                   framework: str = Form(...), control_ref: str = Form(""),
                   description: str = Form(...), category: str = Form(""),
                   likelihood: int = Form(3), impact: int = Form(3),
                   owner: str = Form(""), mitigation: str = Form("")):
    user = get_current_user(request)
    if not user: return JSONResponse({"error": "Unauthorized"}, 401)
    if not has_capability(user, "add_risk"):
        return JSONResponse({"error": "You do not have permission to add risks."}, 403)
    conn = get_db()
    count = conn.execute("SELECT COUNT(*) FROM risks").fetchone()[0]
    risk_id = f"RISK-{count+1:04d}"
    conn.execute("""INSERT INTO risks (risk_id, framework, control_ref, description, category,
                    likelihood, impact, owner, mitigation, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Open')""",
                 (risk_id, framework, control_ref, description, category,
                  likelihood, impact, owner, mitigation))
    conn.commit()
    conn.close()
    log_action(user, f"Added risk {risk_id}", "risk", risk_id)
    try:
        reindex_risk(risk_id)
    except Exception:
        pass
    return JSONResponse({"success": True, "risk_id": risk_id})

@app.get("/mapping", response_class=HTMLResponse)
async def mapping_page(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    mapping = {
        "Access Control": {
            "ISO 27001": ["A.5.15", "A.5.16", "A.5.17", "A.5.18", "A.8.2", "A.8.3", "A.8.5"],
            "SOC 2 Type II": ["CC6.1", "CC6.2", "CC6.3", "CC6.4", "CC6.5"],
            "PCI DSS": ["7.1", "7.2", "8.1", "8.2", "8.3", "8.4"],
            "GDPR": ["Art.25", "Art.32"],
            "Zimbabwe CDPA": ["S.11"],
            "HIPAA": ["164.308(a)(4)", "164.312(a)"],
            "ISO 42001": []
        },
        "Incident Response": {
            "ISO 27001": ["A.5.24", "A.5.25", "A.5.26", "A.5.27", "A.5.28"],
            "SOC 2 Type II": ["CC7.3", "CC7.4", "CC7.5"],
            "PCI DSS": ["12.10"],
            "GDPR": ["Art.33", "Art.34"],
            "Zimbabwe CDPA": ["S.12"],
            "HIPAA": ["164.308(a)(6)", "164.404", "164.406", "164.408"],
            "ISO 42001": ["A.9.2"]
        },
        "Risk Assessment": {
            "ISO 27001": ["6.1.2", "6.1.3"],
            "SOC 2 Type II": ["CC3.1", "CC3.2", "CC3.3", "CC3.4"],
            "PCI DSS": ["12.3"],
            "GDPR": ["Art.35"],
            "Zimbabwe CDPA": ["S.11"],
            "HIPAA": ["164.308(a)(1)"],
            "ISO 42001": ["6.1", "6.1.2", "A.4.3"]
        },
        "Data Classification": {
            "ISO 27001": ["A.5.12", "A.5.13"],
            "SOC 2 Type II": ["C1.1"],
            "PCI DSS": ["3.1", "3.2"],
            "GDPR": ["Art.5", "Art.9"],
            "Zimbabwe CDPA": ["S.6", "S.16"],
            "HIPAA": ["164.514"],
            "ISO 42001": ["A.6.1", "A.6.2"]
        },
        "Third-Party Management": {
            "ISO 27001": ["A.5.19", "A.5.20", "A.5.21", "A.5.22"],
            "SOC 2 Type II": ["CC9.2"],
            "PCI DSS": ["12.8", "12.9"],
            "GDPR": ["Art.28", "Art.29"],
            "Zimbabwe CDPA": ["S.18", "S.19"],
            "HIPAA": ["164.308(b)"],
            "ISO 42001": ["8.7"]
        },
        "Awareness & Training": {
            "ISO 27001": ["7.2", "7.3", "A.6.3"],
            "SOC 2 Type II": ["CC1.4"],
            "PCI DSS": ["12.6"],
            "GDPR": [],
            "Zimbabwe CDPA": [],
            "HIPAA": ["164.308(a)(5)"],
            "ISO 42001": ["7.2", "7.3", "A.2.4", "A.2.5"]
        },
        "Logging & Monitoring": {
            "ISO 27001": ["A.8.15", "A.8.16"],
            "SOC 2 Type II": ["CC4.1", "CC7.1", "CC7.2"],
            "PCI DSS": ["10.1", "10.2", "10.3", "10.4", "10.5"],
            "GDPR": ["Art.30"],
            "Zimbabwe CDPA": [],
            "HIPAA": ["164.312(b)"],
            "ISO 42001": ["A.8.2"]
        },
        "Encryption": {
            "ISO 27001": ["A.8.24"],
            "SOC 2 Type II": ["CC6.7"],
            "PCI DSS": ["3.4", "3.5", "4.1", "4.2"],
            "GDPR": ["Art.32"],
            "Zimbabwe CDPA": ["S.11"],
            "HIPAA": ["164.312(e)"],
            "ISO 42001": []
        },
    }
    return templates.TemplateResponse("mapping.html", {
        "request": request, "user": user,
        "mapping": mapping, "frameworks": list(FRAMEWORKS.keys())
    })

@app.get("/export/excel")
async def export_excel(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    conn = get_db()
    frameworks = conn.execute("SELECT * FROM frameworks ORDER BY name").fetchall()
    wb = Workbook()
    wb.remove(wb.active)
    hdr_fill = PatternFill("solid", start_color="1A2744", end_color="1A2744")
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for fw in frameworks:
        ws = wb.create_sheet(fw["name"][:31])
        ws.sheet_view.showGridLines = False
        headers = ["Ref","Control Name","Category","Doc Type","Document Title","Owner",
                   "Status","Priority","Target Date","Review Date","Version","Notes"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[1].height = 24
        controls = conn.execute("SELECT * FROM controls WHERE framework_id=? ORDER BY ref", (fw["id"],)).fetchall()
        for i, ctrl in enumerate(controls, 2):
            bg = "F8FAFC" if i % 2 == 0 else "FFFFFF"
            row_fill = PatternFill("solid", start_color=bg, end_color=bg)
            vals = [ctrl["ref"], ctrl["name"], ctrl["category"], ctrl["doc_type"],
                    ctrl["document_title"], ctrl["owner"], ctrl["status"], ctrl["priority"],
                    ctrl["target_date"], ctrl["review_date"], ctrl["version"], ctrl["notes"]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=c, value=v or "")
                cell.fill = row_fill; cell.border = border
                cell.font = Font(name="Calibri", size=9)
                cell.alignment = Alignment(vertical="center", wrap_text=c in (2,3,5,12))
            ws.row_dimensions[i].height = 18
        widths = [10,28,22,12,35,18,16,12,14,14,10,30]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w
    conn.close()
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"ARIA_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": f"attachment; filename={fname}"})

@app.get("/audit-log", response_class=HTMLResponse)
async def audit_log_page(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    conn = get_db()
    logs = conn.execute("""SELECT al.*, u.full_name FROM audit_log al
                           LEFT JOIN users u ON al.user_id=u.id
                           ORDER BY al.timestamp DESC LIMIT 200""").fetchall()
    conn.close()
    return templates.TemplateResponse("audit_log.html", {"request": request, "user": user, "logs": logs})

@app.get("/api/controls/{fw_id}")
async def api_controls(request: Request, fw_id: int):
    user = get_current_user(request)
    if not user: return JSONResponse({"error": "Unauthorized"}, 401)
    conn = get_db()
    controls = conn.execute(
        "SELECT id, ref, name, description, doc_type, status FROM controls WHERE framework_id=? ORDER BY ref",
        (fw_id,)
    ).fetchall()
    conn.close()
    return JSONResponse([dict(c) for c in controls])

@app.get("/ai-generator", response_class=HTMLResponse)
async def ai_generator_page(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    conn = get_db()
    frameworks = conn.execute("SELECT * FROM frameworks ORDER BY name").fetchall()
    conn.close()
    api_configured = bool(get_api_key())
    return templates.TemplateResponse("ai_generator.html", {
        "request": request, "user": user,
        "frameworks": frameworks, "api_configured": api_configured
    })

@app.post("/api/generate-policy")
async def api_generate_policy(request: Request,
                               control_id: int = Form(...),
                               org_name: str = Form("Your Organisation")):
    user = get_current_user(request)
    if not user: return JSONResponse({"error": "Unauthorized"}, 401)
    if not has_capability(user, "generate_policy_ai"):
        return JSONResponse({
            "error": "You need Policy Author or Compliance Manager role to generate policies with AI."
        }, 403)

    conn = get_db()
    ctrl = conn.execute("""SELECT c.*, f.name as fw_name
                           FROM controls c JOIN frameworks f ON c.framework_id=f.id
                           WHERE c.id=?""", (control_id,)).fetchone()
    conn.close()

    if not ctrl:
        return JSONResponse({"error": "Control not found"}, 404)

    result = await generate_policy(
        framework=ctrl["fw_name"],
        control_ref=ctrl["ref"],
        control_name=ctrl["name"],
        control_description=ctrl["description"],
        doc_type=ctrl["doc_type"],
        org_name=org_name
    )

    if result["success"]:
        log_action(user, f"Generated AI policy for {ctrl['fw_name']} {ctrl['ref']}", "control", str(control_id))
        # Auto-save to Document Register
        conn = get_db()
        conn.execute("UPDATE controls SET last_updated=? WHERE id=?",
                     (datetime.now().strftime("%Y-%m-%d"), control_id))
        # Check if a doc for this control already exists
        existing = conn.execute(
            "SELECT doc_id FROM documents WHERE control_ref=? AND framework=?",
            (ctrl["ref"], ctrl["fw_name"])
        ).fetchone()
        now = datetime.now().isoformat()
        doc_title = f"{ctrl['name']} — {ctrl['doc_type']}"
        policy_body = result.get("content", "")
        if existing:
            # Bump version, refresh body
            old_ver = conn.execute("SELECT version FROM documents WHERE doc_id=?",
                                   (existing["doc_id"],)).fetchone()["version"]
            try:
                parts = old_ver.split(".")
                new_ver = f"{parts[0]}.{int(parts[1])+1}"
            except:
                new_ver = "1.1"
            conn.execute("""UPDATE documents SET version=?, status='Draft', updated_at=?,
                            comments='AI Generated — updated ' || ?,
                            body=?
                            WHERE doc_id=?""",
                         (new_ver, now, datetime.now().strftime("%Y-%m-%d"),
                          policy_body, existing["doc_id"]))
            saved_doc_id = existing["doc_id"]
        else:
            count = conn.execute("SELECT COUNT(*) FROM documents").fetchone()[0]
            doc_id = f"DOC-{count+1:04d}"
            conn.execute("""INSERT INTO documents
                (doc_id, framework, control_ref, title, doc_type, version, status,
                 owner, created_at, updated_at, comments, body)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (doc_id, ctrl["fw_name"], ctrl["ref"], doc_title,
                 ctrl["doc_type"], "1.0", "Draft",
                 user["full_name"], now, now,
                 f"AI Generated on {datetime.now().strftime('%Y-%m-%d')}",
                 policy_body))
            saved_doc_id = doc_id
        conn.commit()
        conn.close()
        # Refresh search index for this doc
        try:
            reindex_document(saved_doc_id)
        except Exception:
            pass

    return JSONResponse(result)

@app.post("/api/generate-gap-analysis")
async def api_gap_analysis(request: Request, framework_id: int = Form(...)):
    user = get_current_user(request)
    if not user: return JSONResponse({"error": "Unauthorized"}, 401)
    if not has_capability(user, "generate_policy_ai"):
        return JSONResponse({
            "error": "You need Policy Author or Compliance Manager role to run gap analysis."
        }, 403)

    conn = get_db()
    fw = conn.execute("SELECT * FROM frameworks WHERE id=?", (framework_id,)).fetchone()
    controls = conn.execute(
        "SELECT ref, name, status, priority FROM controls WHERE framework_id=?",
        (framework_id,)
    ).fetchall()
    conn.close()

    if not fw:
        return JSONResponse({"error": "Framework not found"}, 404)

    controls_data = [dict(c) for c in controls]
    result = await generate_gap_analysis(fw["name"], controls_data)
    if result["success"]:
        log_action(user, f"Generated gap analysis for {fw['name']}", "framework", str(framework_id))
    return JSONResponse(result)

@app.post("/api/export-word")
async def export_word(request: Request,
                      control_id: int = Form(...),
                      content: str = Form(...),
                      org_name: str = Form("Your Organisation")):
    """Generate a properly formatted Word .docx from AI-generated policy markdown."""
    user = get_current_user(request)
    if not user: return JSONResponse({"error": "Unauthorized"}, 401)

    conn = get_db()
    ctrl = conn.execute("""SELECT c.*, f.name as fw_name FROM controls c
                           JOIN frameworks f ON c.framework_id=f.id
                           WHERE c.id=?""", (control_id,)).fetchone()
    conn.close()
    if not ctrl: raise HTTPException(404)

    import tempfile, subprocess, json as _json
    fw_name = ctrl["fw_name"]
    ctrl_ref = ctrl["ref"]

    # Write input JSON for Node script
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
        _json.dump({
            "title": ctrl["name"],
            "framework": fw_name,
            "control_ref": ctrl_ref,
            "org_name": org_name,
            "content": content,
            "generated_at": datetime.now().strftime("%Y-%m-%d")
        }, f)
        json_path = f.name

    docx_path = json_path.replace('.json', '.docx')

    try:
        # Find generate_docx.js relative to this file
        script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "generate_docx.js")
        result = subprocess.run(
            [r"C:\Program Files\nodejs\node.exe", script_path, json_path, docx_path],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode != 0 or not os.path.exists(docx_path):
            raise Exception(result.stderr or "Node script failed")

        with open(docx_path, 'rb') as f:
            docx_bytes = f.read()

        fname = f"{fw_name.replace(' ','_')}_{ctrl_ref.replace('.','_')}_Policy.docx"
        return StreamingResponse(
            io.BytesIO(docx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": f"attachment; filename={fname}"}
        )
    except Exception as e:
        return JSONResponse({"error": f"Word generation failed: {str(e)}"}, 500)
    finally:
        for p in [json_path, docx_path]:
            try: os.unlink(p)
            except: pass

# ── Ask ARIA (AI Q&A) ────────────────────────────────────────────────────────

@app.get("/ask", response_class=HTMLResponse)
async def ask_page(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    conn = get_db()
    recent = conn.execute(
        """SELECT question, answer, covered, citations, created_at
           FROM ask_log
           WHERE username=?
           ORDER BY id DESC LIMIT 12""",
        (user["username"],)).fetchall()
    # Suggested starter questions — dynamic: pull 5 random policy titles if available
    suggestions_rows = conn.execute(
        """SELECT title FROM documents
           WHERE body IS NOT NULL AND length(body) > 200
           ORDER BY RANDOM() LIMIT 5""").fetchall()
    conn.close()
    suggestions = [
        f"What does our {r['title']} say?" for r in suggestions_rows
    ] or [
        "What is our remote working policy?",
        "How do I report a security incident?",
        "Can I use ChatGPT for work?",
        "What counts as confidential information?",
        "How long do we keep customer data?",
    ]
    return templates.TemplateResponse("ask.html", {
        "request": request, "user": user,
        "recent": recent, "suggestions": suggestions,
    })


@app.post("/api/ask")
async def api_ask(request: Request, question: str = Form(...)):
    user = get_current_user(request)
    if not user: return JSONResponse({"error": "Unauthorized"}, 401)
    question = (question or "").strip()
    if not question:
        return JSONResponse({"error": "Empty question"}, 400)
    if len(question) > 2000:
        return JSONResponse({"error": "Question too long"}, 400)

    result = await ask_policy(question, user=user)
    log_action(user, f"Asked ARIA: {question[:80]}", "ask", "")
    return JSONResponse(result)


@app.post("/api/ask/rebuild")
async def api_ask_rebuild(request: Request):
    user = get_current_user(request)
    if not user or not has_capability(user, "rebuild_index"):
        return JSONResponse({"error": "Admin only"}, 403)
    try:
        n = rebuild_ask_index()
        log_action(user, f"Rebuilt Ask ARIA search index ({n} chunks)", "ask_index", "")
        return JSONResponse({"success": True, "chunks_indexed": n})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, 500)


# ── Admin: User Management ────────────────────────────────────────────────────
# Alphabet for generated temp passwords — excludes ambiguous glyphs (0/O, 1/I/l).
_TEMP_PW_ALPHA = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789"


def _gen_temp_password(length: int = 12) -> str:
    """Generate a human-friendly one-time password for admin handoff."""
    return "".join(secrets.choice(_TEMP_PW_ALPHA) for _ in range(length))


def _render_admin_users(request: Request, user: dict, flash: dict | None = None):
    conn = get_db()
    users = conn.execute(
        """SELECT id, username, email, full_name, role, active,
                  must_change_password, created_at, last_login
             FROM users
             ORDER BY active DESC, username"""
    ).fetchall()
    rows = []
    for u in users:
        role_rows = conn.execute(
            "SELECT role_key FROM user_roles WHERE user_id=? ORDER BY role_key",
            (u["id"],)
        ).fetchall()
        rows.append({**dict(u), "role_keys": [r[0] for r in role_rows]})
    conn.close()
    return templates.TemplateResponse("admin_users.html", {
        "request": request,
        "user": user,
        "users": rows,
        "all_roles": ALL_ROLES,
        "role_labels": ROLE_LABELS,
        "role_descriptions": role_lib.ROLE_DESCRIPTIONS,
        "role_chip_tone": role_lib.ROLE_CHIP_TONE,
        "flash": flash or {},
    })


@app.get("/admin/users", response_class=HTMLResponse)
async def admin_users_page(request: Request):
    user = require_capability(request, "manage_users")
    return _render_admin_users(request, user)


@app.post("/admin/users/create")
async def admin_create_user(request: Request,
                             username: str = Form(...),
                             email: str = Form(...),
                             full_name: str = Form(...)):
    admin = require_capability(request, "manage_users")
    username = (username or "").strip()
    email = (email or "").strip().lower()
    full_name = (full_name or "").strip()
    if not (username and email and full_name):
        return _render_admin_users(request, admin,
            {"type": "error", "message": "Username, email, and full name are all required."})

    conn = get_db()
    dup = conn.execute("SELECT id FROM users WHERE username=? OR email=?",
                        (username, email)).fetchone()
    if dup:
        conn.close()
        return _render_admin_users(request, admin,
            {"type": "error", "message": "A user with that username or email already exists."})

    temp_pw = _gen_temp_password()
    cur = conn.execute(
        """INSERT INTO users
           (username, email, full_name, password_hash, role, active, must_change_password)
           VALUES (?, ?, ?, ?, ?, 1, 1)""",
        (username, email, full_name, hash_password(temp_pw), role_lib.EMPLOYEE)
    )
    new_id = cur.lastrowid
    conn.execute(
        "INSERT OR IGNORE INTO user_roles (user_id, role_key, granted_by) VALUES (?, ?, ?)",
        (new_id, role_lib.EMPLOYEE, int(admin["id"]))
    )
    conn.commit()
    conn.close()
    log_action(admin, f"Created user {username}", "user", str(new_id), "", username)
    return _render_admin_users(request, admin, {
        "type": "new_user",
        "username": username,
        "full_name": full_name,
        "temp_password": temp_pw,
    })


@app.post("/admin/users/{uid}/roles/grant")
async def admin_grant_role(request: Request, uid: int, role_key: str = Form(...)):
    admin = require_capability(request, "manage_users")
    role_key = (role_key or "").strip()
    if role_key not in ALL_ROLES:
        return _render_admin_users(request, admin,
            {"type": "error", "message": f"Unknown role: {role_key}"})
    conn = get_db()
    target = conn.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
    if not target:
        conn.close()
        return _render_admin_users(request, admin,
            {"type": "error", "message": "User not found."})
    conn.execute(
        "INSERT OR IGNORE INTO user_roles (user_id, role_key, granted_by) VALUES (?, ?, ?)",
        (uid, role_key, int(admin["id"]))
    )
    conn.commit()
    conn.close()
    log_action(admin, f"Granted role '{role_key}' to {target['username']}",
               "user", str(uid), "", role_key)
    return RedirectResponse("/admin/users", status_code=302)


@app.post("/admin/users/{uid}/roles/revoke")
async def admin_revoke_role(request: Request, uid: int, role_key: str = Form(...)):
    admin = require_capability(request, "manage_users")
    conn = get_db()
    target = conn.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
    if not target:
        conn.close()
        return _render_admin_users(request, admin,
            {"type": "error", "message": "User not found."})

    # Safety: never let the last admin account lose its admin role.
    if role_key == role_lib.ADMIN:
        others = conn.execute(
            "SELECT COUNT(*) FROM user_roles WHERE role_key=? AND user_id!=?",
            (role_lib.ADMIN, uid)
        ).fetchone()[0]
        if others == 0:
            conn.close()
            return _render_admin_users(request, admin, {
                "type": "error",
                "message": "Cannot revoke the last admin role — at least one active admin must remain."
            })

    conn.execute("DELETE FROM user_roles WHERE user_id=? AND role_key=?", (uid, role_key))
    # Ensure every user keeps at least one role; fall back to employee.
    remaining = conn.execute(
        "SELECT 1 FROM user_roles WHERE user_id=? LIMIT 1", (uid,)
    ).fetchone()
    if not remaining:
        conn.execute(
            "INSERT OR IGNORE INTO user_roles (user_id, role_key, granted_by) VALUES (?, ?, ?)",
            (uid, role_lib.EMPLOYEE, int(admin["id"]))
        )
    conn.commit()
    conn.close()
    log_action(admin, f"Revoked role '{role_key}' from {target['username']}",
               "user", str(uid), role_key, "")
    return RedirectResponse("/admin/users", status_code=302)


@app.post("/admin/users/{uid}/deactivate")
async def admin_deactivate_user(request: Request, uid: int):
    admin = require_capability(request, "manage_users")
    if int(admin["id"]) == uid:
        return _render_admin_users(request, admin,
            {"type": "error", "message": "You cannot deactivate your own account."})
    conn = get_db()
    target = conn.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
    if not target:
        conn.close()
        return _render_admin_users(request, admin,
            {"type": "error", "message": "User not found."})
    # Block deactivating the last active admin.
    is_target_admin = conn.execute(
        "SELECT 1 FROM user_roles WHERE user_id=? AND role_key=?",
        (uid, role_lib.ADMIN)
    ).fetchone()
    if is_target_admin:
        other_active_admins = conn.execute("""
            SELECT COUNT(*) FROM users u
            JOIN user_roles ur ON ur.user_id=u.id
            WHERE ur.role_key=? AND u.active=1 AND u.id!=?
        """, (role_lib.ADMIN, uid)).fetchone()[0]
        if other_active_admins == 0:
            conn.close()
            return _render_admin_users(request, admin, {
                "type": "error",
                "message": "Cannot deactivate the last active admin."
            })
    conn.execute("UPDATE users SET active=0 WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    log_action(admin, f"Deactivated user {target['username']}",
               "user", str(uid), "active", "inactive")
    return RedirectResponse("/admin/users", status_code=302)


@app.post("/admin/users/{uid}/activate")
async def admin_activate_user(request: Request, uid: int):
    admin = require_capability(request, "manage_users")
    conn = get_db()
    target = conn.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
    if not target:
        conn.close()
        return _render_admin_users(request, admin,
            {"type": "error", "message": "User not found."})
    conn.execute("UPDATE users SET active=1 WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    log_action(admin, f"Activated user {target['username']}",
               "user", str(uid), "inactive", "active")
    return RedirectResponse("/admin/users", status_code=302)


@app.post("/admin/users/{uid}/reset-password")
async def admin_reset_password(request: Request, uid: int):
    admin = require_capability(request, "manage_users")
    conn = get_db()
    target = conn.execute(
        "SELECT username, full_name FROM users WHERE id=?", (uid,)
    ).fetchone()
    if not target:
        conn.close()
        return _render_admin_users(request, admin,
            {"type": "error", "message": "User not found."})
    temp_pw = _gen_temp_password()
    conn.execute(
        "UPDATE users SET password_hash=?, must_change_password=1 WHERE id=?",
        (hash_password(temp_pw), uid)
    )
    conn.commit()
    conn.close()
    log_action(admin, f"Reset password for {target['username']}", "user", str(uid))
    return _render_admin_users(request, admin, {
        "type": "reset_pw",
        "username": target["username"],
        "full_name": target["full_name"],
        "temp_password": temp_pw,
    })


@app.on_event("startup")
async def startup():
    init_db()
    try:
        init_ask_index()
        # Warm the index on first boot if it's empty
        conn = get_db()
        empty = conn.execute("SELECT COUNT(*) FROM ask_index").fetchone()[0] == 0
        conn.close()
        if empty:
            rebuild_ask_index()
    except Exception as e:
        print(f"⚠️  Ask index init failed: {e}")
    print("🚀 ARIA started at http://localhost:8000")
