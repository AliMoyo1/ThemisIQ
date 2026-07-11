"""
ERM module — Data access layer.
Covers erm_enterprise_risks, erm_risk_appetite, erm_risk_library,
erm_regulatory_obligations, erm_assessments, and the shared risk_register view.
"""
import io
import logging
import re
from datetime import datetime
from difflib import get_close_matches
from core.timeutils import utcnow
from database import get_db, insert_returning_id, sql_now_offset, sql_now_ts, sql_days_between, sql_date_offset, sql_date_ts, sql_current_date


def _dict(row):
    return dict(row) if row else None


def _dicts(rows):
    return [dict(r) for r in rows]


def _now():
    return utcnow().strftime("%Y-%m-%d %H:%M:%S")


# ═════════════════════════════════════════════════════════════════════════════
# RISK RATING FRAMEWORK — active matrix lookup
# ═════════════════════════════════════════════════════════════════════════════

def get_active_framework_matrix(db):
    """Fetch the active framework's band metadata + likelihood×impact lookup.

    Call once per request/function and reuse across a loop — never re-query
    per risk row (25 matrix rows + up to a handful of bands, trivial cost).
    """
    bands = {r["band_key"]: dict(r) for r in db.execute(
        "SELECT b.* FROM erm_framework_bands b "
        "JOIN erm_risk_frameworks f ON f.id=b.framework_id WHERE f.is_active=1"
    ).fetchall()}
    matrix = {(r["likelihood"], r["impact"]): r["band_key"] for r in db.execute(
        "SELECT mb.* FROM erm_framework_matrix_bands mb "
        "JOIN erm_risk_frameworks f ON f.id=mb.framework_id WHERE f.is_active=1"
    ).fetchall()}
    return {"bands": bands, "matrix": matrix}


def resolve_band(fw_matrix, likelihood, impact):
    """Resolve a (likelihood, impact) pair to a band_key via the active framework."""
    key = (int(likelihood or 3), int(impact or 3))
    return fw_matrix["matrix"].get(key, "moderate")


# Reusable JOIN fragment for SQL-side aggregate counts against the active
# framework's matrix — alias the target table as "e". Cheaper than pulling
# every row into Python just to call resolve_band() when only a COUNT is needed.
_FW_BAND_JOIN = (
    "JOIN erm_risk_frameworks f ON f.is_active=1 "
    "JOIN erm_framework_matrix_bands mb ON mb.framework_id=f.id "
    "AND mb.likelihood=e.likelihood AND mb.impact=e.impact"
)


def get_framework_detail(framework_id):
    """Full framework detail by id: dimensions+levels, scales, bands, matrix, taxonomy.

    Flat shape — name/description/id/is_active/is_default/source/updated_at
    all at the top level — used identically as the GET detail response, the
    PUT/import request body, and the export response body (plus a
    "schema_version" key added only for export). This is generalized from
    get_active_framework()'s original shape (see below, kept nested for
    backward compatibility with the already-shipped Rating Guide).
    """
    db = get_db()
    try:
        fw = _dict(db.execute("SELECT * FROM erm_risk_frameworks WHERE id=%s", (framework_id,)).fetchone())
        if not fw:
            return None

        dims = []
        for d in db.execute(
            "SELECT * FROM erm_framework_impact_dimensions WHERE framework_id=%s ORDER BY order_idx",
            (fw["id"],),
        ).fetchall():
            levels = _dicts(db.execute(
                "SELECT level, description, threshold_label, threshold_min, threshold_max "
                "FROM erm_framework_impact_levels WHERE dimension_id=%s ORDER BY level",
                (d["id"],),
            ).fetchall())
            dims.append({"id": d["id"], "name": d["name"], "levels": levels})

        scales = {}
        for row in db.execute(
            "SELECT scale_type, level, label, description FROM erm_framework_scales "
            "WHERE framework_id=%s ORDER BY scale_type, level",
            (fw["id"],),
        ).fetchall():
            scales.setdefault(row["scale_type"], []).append(dict(row))

        bands = _dicts(db.execute(
            "SELECT band_key, label, color, sort_order FROM erm_framework_bands "
            "WHERE framework_id=%s ORDER BY sort_order",
            (fw["id"],),
        ).fetchall())

        matrix = _dicts(db.execute(
            "SELECT likelihood, impact, band_key FROM erm_framework_matrix_bands WHERE framework_id=%s",
            (fw["id"],),
        ).fetchall())

        tax_rows = _dicts(db.execute(
            "SELECT id, parent_id, name, order_idx FROM erm_framework_taxonomy "
            "WHERE framework_id=%s ORDER BY order_idx",
            (fw["id"],),
        ).fetchall())
        by_parent = {}
        for t in tax_rows:
            by_parent.setdefault(t["parent_id"], []).append(t)

        def build_tree(parent_id):
            return [{"id": n["id"], "name": n["name"], "children": build_tree(n["id"])}
                    for n in by_parent.get(parent_id, [])]

        return {
            "id": fw["id"], "name": fw["name"], "description": fw["description"],
            "is_active": bool(fw["is_active"]), "is_default": bool(fw["is_default"]),
            "source": fw["source"], "updated_at": fw.get("updated_at"),
            "dimensions": dims,
            "likelihood": scales.get("likelihood", []),
            "impact_scale": scales.get("impact", []),
            "control_effectiveness": scales.get("control_effectiveness", []),
            "bands": bands,
            "matrix": matrix,
            "taxonomy": build_tree(None),
        }
    finally:
        db.close()


def get_active_framework():
    """Active framework in the nested shape the Rating Guide already consumes.

    Kept byte-for-byte compatible with the shape shipped in slice 1
    (`{"framework": {id, name, description}, "dimensions": [...], ...}`) so
    the existing Risk Rating Guide page needs no changes. New code (the
    framework admin list/editor/import/export) uses get_framework_detail()'s
    flatter shape instead.
    """
    db = get_db()
    try:
        row = db.execute("SELECT id FROM erm_risk_frameworks WHERE is_active=1 LIMIT 1").fetchone()
    finally:
        db.close()
    if not row:
        return None
    detail = get_framework_detail(row["id"])
    if not detail:
        return None
    return {
        "framework": {"id": detail["id"], "name": detail["name"], "description": detail["description"]},
        "dimensions": detail["dimensions"],
        "likelihood": detail["likelihood"],
        "impact_scale": detail["impact_scale"],
        "control_effectiveness": detail["control_effectiveness"],
        "bands": detail["bands"],
        "matrix": detail["matrix"],
        "taxonomy": detail["taxonomy"],
    }


def list_frameworks():
    """Summary list of every framework for this tenant, for the framework admin list view."""
    db = get_db()
    try:
        rows = db.execute(
            "SELECT f.id, f.name, f.description, f.is_active, f.is_default, f.source, f.updated_at, "
            "(SELECT COUNT(*) FROM erm_framework_impact_dimensions d WHERE d.framework_id=f.id) AS dimension_count "
            "FROM erm_risk_frameworks f ORDER BY f.is_active DESC, f.name"
        ).fetchall()
        return [{**dict(r), "is_active": bool(r["is_active"]), "is_default": bool(r["is_default"])} for r in rows]
    finally:
        db.close()


def recompute_risk_bands(db):
    """Re-derive qualitative_score for every risk from the currently active framework.

    Operates on an already-open connection and does not commit — callers
    (activate_framework, and update_framework when editing the framework
    that's currently active) own the transaction and commit once alongside
    their own writes.
    """
    band_map = {(r["likelihood"], r["impact"]): r["band_key"] for r in db.execute(
        "SELECT mb.likelihood, mb.impact, mb.band_key FROM erm_framework_matrix_bands mb "
        "JOIN erm_risk_frameworks f ON f.id=mb.framework_id WHERE f.is_active=1"
    ).fetchall()}
    if not band_map:
        return
    for r in db.execute("SELECT id, likelihood, impact FROM erm_enterprise_risks").fetchall():
        band = band_map.get((r["likelihood"] or 3, r["impact"] or 3), "moderate")
        db.execute("UPDATE erm_enterprise_risks SET qualitative_score=%s WHERE id=%s", (band, r["id"]))


_BAND_KEY_RE = re.compile(r"^[a-z0-9_]+$")
_HEX_COLOR_RE = re.compile(r"^#[0-9A-Fa-f]{6}$")


def validate_framework_payload(payload):
    """Validate a full framework write payload (PUT body / import body).

    Returns a list of human-readable error strings; empty means valid.
    Collects every problem found rather than stopping at the first, so a
    hand-edited import file gets full feedback in one round-trip.
    """
    errors = []
    if not isinstance(payload, dict):
        return ["Payload must be a JSON object"]

    if not str(payload.get("name") or "").strip():
        errors.append("name is required")

    dims = payload.get("dimensions")
    if not isinstance(dims, list) or not dims:
        errors.append("dimensions must be a non-empty list")
        dims = []
    for i, d in enumerate(dims):
        if not isinstance(d, dict) or not str(d.get("name") or "").strip():
            errors.append(f"dimensions[{i}].name is required")
        levels = d.get("levels") if isinstance(d, dict) else None
        if not isinstance(levels, list) or len(levels) != 5:
            errors.append(f"dimensions[{i}].levels must have exactly 5 entries")
            continue
        seen = set()
        for lvl in levels:
            if not isinstance(lvl, dict):
                errors.append(f"dimensions[{i}] has a malformed level entry")
                continue
            n = lvl.get("level")
            if n not in (1, 2, 3, 4, 5) or n in seen:
                errors.append(f"dimensions[{i}] has a missing/duplicate level number")
            seen.add(n)
            if not str(lvl.get("description") or "").strip():
                errors.append(f"dimensions[{i}] level {n} description is required")

    for scale_name in ("likelihood", "impact_scale", "control_effectiveness"):
        rows = payload.get(scale_name)
        if not isinstance(rows, list) or len(rows) != 5:
            errors.append(f"{scale_name} must have exactly 5 entries")
            continue
        seen = set()
        for row in rows:
            if not isinstance(row, dict):
                errors.append(f"{scale_name} has a malformed entry")
                continue
            n = row.get("level")
            if n not in (1, 2, 3, 4, 5) or n in seen:
                errors.append(f"{scale_name} has a missing/duplicate level number")
            seen.add(n)
            if not str(row.get("label") or "").strip():
                errors.append(f"{scale_name} level {n} label is required")

    bands = payload.get("bands")
    band_keys = set()
    if not isinstance(bands, list) or not (2 <= len(bands) <= 10):
        errors.append("bands must be a list of 2-10 entries")
        bands = []
    for i, b in enumerate(bands):
        if not isinstance(b, dict):
            errors.append(f"bands[{i}] is malformed")
            continue
        key = b.get("band_key")
        if not isinstance(key, str) or not _BAND_KEY_RE.match(key):
            errors.append(f"bands[{i}].band_key must be lowercase letters/digits/underscore")
        elif key in band_keys:
            errors.append(f"bands[{i}].band_key '{key}' is duplicated")
        else:
            band_keys.add(key)
        if not str(b.get("label") or "").strip():
            errors.append(f"bands[{i}].label is required")
        if not _HEX_COLOR_RE.match(str(b.get("color") or "")):
            errors.append(f"bands[{i}].color must be a hex color like #RRGGBB")

    matrix = payload.get("matrix")
    if not isinstance(matrix, list) or len(matrix) != 25:
        errors.append("matrix must have exactly 25 entries (5x5)")
        matrix = []
    seen_cells = set()
    for i, m in enumerate(matrix):
        if not isinstance(m, dict):
            errors.append(f"matrix[{i}] is malformed")
            continue
        l, imp, key = m.get("likelihood"), m.get("impact"), m.get("band_key")
        if l not in (1, 2, 3, 4, 5) or imp not in (1, 2, 3, 4, 5):
            errors.append(f"matrix[{i}] likelihood/impact must be 1-5")
            continue
        if (l, imp) in seen_cells:
            errors.append(f"matrix has a duplicate cell for likelihood={l}, impact={imp}")
        seen_cells.add((l, imp))
        if key not in band_keys:
            errors.append(f"matrix cell ({l},{imp}) references unknown band_key '{key}'")
    missing_cells = {(l, i) for l in range(1, 6) for i in range(1, 6)} - seen_cells
    if missing_cells:
        errors.append(f"matrix is missing {len(missing_cells)} cell(s)")

    taxonomy = payload.get("taxonomy")
    if not isinstance(taxonomy, list):
        errors.append("taxonomy must be a list")
        taxonomy = []
    node_count = [0]

    def walk_taxonomy(nodes, depth):
        if depth > 20:
            errors.append("taxonomy nesting exceeds 20 levels")
            return
        for n in nodes:
            if not isinstance(n, dict) or not str(n.get("name") or "").strip():
                errors.append("a taxonomy node is missing a name")
                continue
            node_count[0] += 1
            if node_count[0] > 500:
                return
            children = n.get("children") or []
            if isinstance(children, list):
                walk_taxonomy(children, depth + 1)
            else:
                errors.append("a taxonomy node's children must be a list")

    walk_taxonomy(taxonomy, 1)
    if node_count[0] > 500:
        errors.append("taxonomy exceeds 500 total nodes")

    return errors


def _apply_framework_payload(db, framework_id, payload):
    """Replace a framework's dimensions/levels/scales/bands/matrix/taxonomy from payload.

    Deletes all existing child rows for framework_id then re-inserts from
    payload, mirroring the seed function's insert-loop idiom. Also updates
    name/description/updated_at on the framework row itself. Does not touch
    is_active/is_default/source — callers own those. Does not commit —
    caller commits once as part of its own unit of work. Safe as
    delete-then-reinsert-in-one-txn: both connection wrappers discard
    partial work on an unswallowed exception as long as db.close() runs in
    the caller's finally.
    """
    db.execute(
        "UPDATE erm_risk_frameworks SET name=%s, description=%s, updated_at=%s WHERE id=%s",
        (payload.get("name"), payload.get("description"), _now(), framework_id),
    )

    db.execute("DELETE FROM erm_framework_impact_dimensions WHERE framework_id=%s", (framework_id,))
    for dim_idx, d in enumerate(payload.get("dimensions") or []):
        dim_id = insert_returning_id(
            db,
            "INSERT INTO erm_framework_impact_dimensions (framework_id, name, order_idx) VALUES (%s,%s,%s)",
            (framework_id, d["name"], dim_idx),
        )
        for lvl in d.get("levels") or []:
            db.execute(
                "INSERT INTO erm_framework_impact_levels "
                "(dimension_id, level, description, threshold_label, threshold_min, threshold_max) "
                "VALUES (%s,%s,%s,%s,%s,%s)",
                (dim_id, lvl["level"], lvl.get("description"), lvl.get("threshold_label"),
                 lvl.get("threshold_min"), lvl.get("threshold_max")),
            )

    db.execute("DELETE FROM erm_framework_scales WHERE framework_id=%s", (framework_id,))
    for scale_type, key in (("likelihood", "likelihood"), ("impact", "impact_scale"),
                            ("control_effectiveness", "control_effectiveness")):
        for row in payload.get(key) or []:
            db.execute(
                "INSERT INTO erm_framework_scales (framework_id, scale_type, level, label, description) "
                "VALUES (%s,%s,%s,%s,%s)",
                (framework_id, scale_type, row["level"], row.get("label"), row.get("description")),
            )

    db.execute("DELETE FROM erm_framework_bands WHERE framework_id=%s", (framework_id,))
    for b in payload.get("bands") or []:
        db.execute(
            "INSERT INTO erm_framework_bands (framework_id, band_key, label, color, sort_order) "
            "VALUES (%s,%s,%s,%s,%s)",
            (framework_id, b["band_key"], b.get("label"), b.get("color"), b.get("sort_order", 0)),
        )

    db.execute("DELETE FROM erm_framework_matrix_bands WHERE framework_id=%s", (framework_id,))
    for m in payload.get("matrix") or []:
        db.execute(
            "INSERT INTO erm_framework_matrix_bands (framework_id, likelihood, impact, band_key) "
            "VALUES (%s,%s,%s,%s)",
            (framework_id, m["likelihood"], m["impact"], m["band_key"]),
        )

    db.execute("DELETE FROM erm_framework_taxonomy WHERE framework_id=%s", (framework_id,))

    def insert_taxonomy(nodes, parent_id):
        for idx, n in enumerate(nodes):
            node_id = insert_returning_id(
                db,
                "INSERT INTO erm_framework_taxonomy (framework_id, parent_id, name, order_idx) "
                "VALUES (%s,%s,%s,%s)",
                (framework_id, parent_id, n["name"], idx),
            )
            insert_taxonomy(n.get("children") or [], node_id)

    insert_taxonomy(payload.get("taxonomy") or [], None)


def create_framework_from_clone(name, description, clone_from_id):
    """Create a new framework by deep-copying an existing one's contents.

    There is no from-scratch/blank creation path — every tenant always has
    at least the immutable built-in framework to clone from, so a new
    framework is never left in an incomplete state that would fail
    validate_framework_payload(). Raises LookupError if clone_from_id
    doesn't exist.
    """
    source = get_framework_detail(clone_from_id)
    if not source:
        raise LookupError(f"Framework {clone_from_id} not found")
    db = get_db()
    try:
        new_id = insert_returning_id(
            db,
            "INSERT INTO erm_risk_frameworks (name, description, is_active, is_default, source) "
            "VALUES (%s,%s,0,0,'manual')",
            (name, description),
        )
        payload = {**source, "name": name, "description": description}
        _apply_framework_payload(db, new_id, payload)
        db.commit()
        return new_id
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def update_framework(framework_id, payload):
    """Replace an existing framework's contents. Raises LookupError if missing,
    PermissionError if the framework is the immutable built-in one."""
    db = get_db()
    try:
        fw = _dict(db.execute(
            "SELECT source, is_active FROM erm_risk_frameworks WHERE id=%s", (framework_id,)
        ).fetchone())
        if not fw:
            raise LookupError(f"Framework {framework_id} not found")
        if fw["source"] == "built_in":
            raise PermissionError("The built-in framework cannot be edited — clone it instead")
        _apply_framework_payload(db, framework_id, payload)
        if fw["is_active"]:
            recompute_risk_bands(db)
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def delete_framework(framework_id):
    """Delete a framework. Raises LookupError if missing, PermissionError if
    it's currently active or the immutable built-in one."""
    db = get_db()
    try:
        fw = _dict(db.execute(
            "SELECT source, is_active FROM erm_risk_frameworks WHERE id=%s", (framework_id,)
        ).fetchone())
        if not fw:
            raise LookupError(f"Framework {framework_id} not found")
        if fw["source"] == "built_in":
            raise PermissionError("The built-in framework cannot be deleted")
        if fw["is_active"]:
            raise PermissionError("Cannot delete the currently active framework — activate another one first")
        db.execute("DELETE FROM erm_risk_frameworks WHERE id=%s", (framework_id,))
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def activate_framework(framework_id):
    """Make framework_id the sole active framework and recompute all risk bands.

    Uses a single atomic UPDATE (CASE WHEN) rather than two separate
    statements ("clear others" then "set this one") so concurrent activation
    requests can never leave two frameworks marked active at once — a state
    get_active_framework_matrix() would silently and incorrectly merge
    bands/matrix from by dict key.
    """
    db = get_db()
    try:
        fw = db.execute("SELECT id FROM erm_risk_frameworks WHERE id=%s", (framework_id,)).fetchone()
        if not fw:
            raise LookupError(f"Framework {framework_id} not found")
        db.execute(
            "UPDATE erm_risk_frameworks SET is_active = CASE WHEN id=%s THEN 1 ELSE 0 END",
            (framework_id,),
        )
        recompute_risk_bands(db)
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def import_framework(payload, name_override=None):
    """Create a new framework from an imported payload (already validated by
    the caller via validate_framework_payload). Returns the new framework id."""
    db = get_db()
    try:
        name = name_override or payload.get("name") or "Imported Framework"
        new_id = insert_returning_id(
            db,
            "INSERT INTO erm_risk_frameworks (name, description, is_active, is_default, source) "
            "VALUES (%s,%s,0,0,'imported')",
            (name, payload.get("description")),
        )
        _apply_framework_payload(db, new_id, {**payload, "name": name})
        db.commit()
        return new_id
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# ENTERPRISE RISKS
# ═════════════════════════════════════════════════════════════════════════════

def list_enterprise_risks(category=None, status=None, source_module=None, board_only=False, limit=500, bu_id=None):
    db = get_db()
    try:
        where, params = [], []
        if category:
            where.append("category=%s"); params.append(category)
        if status:
            where.append("status=%s"); params.append(status)
        if source_module:
            where.append("source_module=%s"); params.append(source_module)
        if board_only:
            where.append("board_visibility=1")
        if bu_id is not None:
            where.append("business_unit_id=%s"); params.append(bu_id)
        clause = ("WHERE " + " AND ".join(where)) if where else ""
        rows = db.execute(
            f"SELECT e.*, u.full_name AS owner_name "
            f"FROM erm_enterprise_risks e "
            f"LEFT JOIN users u ON u.id = e.owner_id "
            f"{clause} ORDER BY "
            f"CASE status WHEN 'open' THEN 0 WHEN 'under_review' THEN 1 ELSE 2 END, "
            f"(likelihood*impact) DESC LIMIT %s",
            params + [limit],
        ).fetchall()
        return _dicts(rows)
    finally:
        db.close()


def get_enterprise_risk(risk_id):
    db = get_db()
    try:
        risk = _dict(db.execute(
            "SELECT e.*, u.full_name AS owner_name, 'erm' AS register_source "
            "FROM erm_enterprise_risks e LEFT JOIN users u ON u.id=e.owner_id "
            "WHERE e.id=%s", (risk_id,)
        ).fetchone())
        if risk:
            risk["dimension_scores"] = _get_dimension_scores(db, risk_id)
        return risk
    finally:
        db.close()


def create_enterprise_risk(data):
    db = get_db()
    try:
        dim_scores = data.pop("dimension_scores", None)
        if dim_scores:
            derived = _impact_from_dimensions(dim_scores)
            if derived is not None:
                data["impact"] = derived
        inh, res, qual = _compute_scores(db, data)
        new_id = insert_returning_id(db,
            """INSERT INTO erm_enterprise_risks
               (title, description, category, sub_category, likelihood, impact, velocity,
                strategic_objective, owner_id, reviewer_id, treatment, treatment_plan,
                residual_likelihood, residual_impact, status, board_visibility,
                regulation_links, review_date, source_module, source_risk_id, created_by,
                inherent_score, residual_score, qualitative_score,
                risk_statement, workflow_step, response_deadline)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (data.get("title"), data.get("description"), data.get("category", "Strategic Risk"),
             data.get("sub_category"), data.get("likelihood", 3), data.get("impact", 3),
             data.get("velocity", 3), data.get("strategic_objective"),
             data.get("owner_id"), data.get("reviewer_id"),
             data.get("treatment", "mitigate"), data.get("treatment_plan"),
             data.get("residual_likelihood"), data.get("residual_impact"),
             data.get("status", "open"), data.get("board_visibility", 0),
             data.get("regulation_links"), data.get("review_date"),
             data.get("source_module", "erm"), data.get("source_risk_id"),
             data.get("created_by"),
             inh, res, qual,
             data.get("risk_statement"), data.get("workflow_step", "draft"),
             data.get("response_deadline")),
        )
        if dim_scores:
            _save_dimension_scores(db, new_id, dim_scores)
        db.commit()
        return new_id
    finally:
        db.close()


def _compute_scores(db, data, existing=None):
    """Auto-derive inherent_score, residual_score, qualitative_score from L/I values."""
    L = data.get("likelihood") or (existing.get("likelihood") if existing else 3) or 3
    I = data.get("impact")     or (existing.get("impact")     if existing else 3) or 3
    RL = data.get("residual_likelihood") or (existing.get("residual_likelihood") if existing else None)
    RI = data.get("residual_impact")     or (existing.get("residual_impact")     if existing else None)
    inherent  = int(L) * int(I)
    residual  = int(RL) * int(RI) if RL and RI else None
    qual = resolve_band(get_active_framework_matrix(db), L, I)
    return inherent, residual, qual


def update_enterprise_risk(risk_id, data):
    db = get_db()
    try:
        existing = _dict(db.execute(
            "SELECT likelihood, impact, residual_likelihood, residual_impact "
            "FROM erm_enterprise_risks WHERE id=%s", (risk_id,)
        ).fetchone())
        dim_scores = data.pop("dimension_scores", None)
        if dim_scores:
            derived = _impact_from_dimensions(dim_scores)
            if derived is not None:
                data["impact"] = derived
        fields, vals = [], []
        for k in ("title", "description", "category", "sub_category", "likelihood", "impact",
                  "velocity", "strategic_objective", "owner_id", "reviewer_id", "treatment",
                  "treatment_plan", "residual_likelihood", "residual_impact", "status",
                  "board_visibility", "regulation_links", "review_date",
                  "risk_statement", "workflow_step", "response_deadline", "effectiveness_rating",
                  "last_reviewed"):
            if k in data:
                fields.append(f"{k}=%s"); vals.append(data[k])
        if any(k in data for k in ("likelihood", "impact", "residual_likelihood", "residual_impact")):
            inh, res, qual = _compute_scores(db, data, existing)
            fields += ["inherent_score=%s", "qualitative_score=%s"]
            vals   += [inh, qual]
            if res is not None:
                fields.append("residual_score=%s"); vals.append(res)
        if fields:
            fields.append("updated_at=%s"); vals.append(_now()); vals.append(risk_id)
            db.execute(f"UPDATE erm_enterprise_risks SET {','.join(fields)} WHERE id=%s", vals)
        if dim_scores:
            _save_dimension_scores(db, risk_id, dim_scores)
        db.commit()
    finally:
        db.close()


def delete_enterprise_risk(risk_id):
    db = get_db()
    try:
        db.execute("UPDATE erm_kris SET linked_risk_id=NULL WHERE linked_risk_id=%s", (risk_id,))
        db.execute("DELETE FROM erm_risk_dimension_scores WHERE risk_id=%s", (risk_id,))
        db.execute("DELETE FROM erm_risk_workflow_history WHERE risk_id=%s", (risk_id,))
        db.execute("UPDATE erm_regulatory_obligations SET linked_erm_risk_id=NULL WHERE linked_erm_risk_id=%s", (risk_id,))
        db.execute("UPDATE orm_events SET erm_risk_id=NULL WHERE erm_risk_id=%s", (risk_id,))
        db.execute("UPDATE ai_risk_predictions SET erm_risk_id=NULL WHERE erm_risk_id=%s", (risk_id,))
        db.execute("DELETE FROM cross_module_links WHERE target_module='erm' AND target_type='enterprise_risk' AND target_id=%s", (risk_id,))
        db.execute("DELETE FROM cross_module_links WHERE source_module='erm' AND source_type='enterprise_risk' AND source_id=%s", (risk_id,))
        db.execute("DELETE FROM erm_enterprise_risks WHERE id=%s", (risk_id,))
        db.commit()
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# RISK ↔ CONTROL LINKING
# ═════════════════════════════════════════════════════════════════════════════

def list_risk_controls(risk_id):
    """Return all controls linked to a risk, joined with canonical_controls for title/ref."""
    db = get_db()
    try:
        rows = db.execute(
            "SELECT rc.*, cc.title AS control_title, cc.ref AS control_ref "
            "FROM risk_controls rc JOIN canonical_controls cc ON rc.control_id = cc.id "
            "WHERE rc.risk_id=%s ORDER BY cc.title",
            (risk_id,),
        ).fetchall()
        return _dicts(rows)
    finally:
        db.close()


def link_risk_control(risk_id, control_id, user_id, weight=1.0):
    """Link a control to a risk. Weight is clamped to [0.1, 5.0]. Returns True."""
    weight = max(0.1, min(5.0, float(weight)))
    db = get_db()
    try:
        db.execute(
            "INSERT INTO risk_controls (risk_id, control_id, weight, direction, created_by) "
            "VALUES (%s, %s, %s, 'mitigates', %s) "
            "ON CONFLICT (risk_id, control_id) DO NOTHING",
            (risk_id, control_id, weight, user_id),
        )
        db.commit()
        return True
    finally:
        db.close()


def unlink_risk_control(risk_id, control_id):
    """Remove a control link from a risk. Returns True."""
    db = get_db()
    try:
        db.execute(
            "DELETE FROM risk_controls WHERE risk_id=%s AND control_id=%s",
            (risk_id, control_id),
        )
        db.commit()
        return True
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# DIMENSION SCORES — per-risk impact breakdown
# ═════════════════════════════════════════════════════════════════════════════

def _save_dimension_scores(db, risk_id, dimension_scores):
    """Replace all dimension scores for a risk. Each entry: {dimension_name, score}."""
    db.execute("DELETE FROM erm_risk_dimension_scores WHERE risk_id=%s", (risk_id,))
    for ds in (dimension_scores or []):
        name = str(ds.get("dimension_name") or "").strip()
        score = ds.get("score")
        if not name or score is None:
            continue
        score = max(1, min(5, int(score)))
        db.execute(
            "INSERT INTO erm_risk_dimension_scores (risk_id, dimension_name, score) "
            "VALUES (%s,%s,%s)",
            (risk_id, name, score),
        )


def _get_dimension_scores(db, risk_id):
    """Return [{dimension_name, score}] for a risk."""
    return [{"dimension_name": r["dimension_name"], "score": r["score"]}
            for r in db.execute(
                "SELECT dimension_name, score FROM erm_risk_dimension_scores "
                "WHERE risk_id=%s ORDER BY dimension_name",
                (risk_id,),
            ).fetchall()]


def _impact_from_dimensions(dimension_scores):
    """Derive overall impact as MAX of individual dimension scores."""
    scores = [int(ds.get("score") or 0) for ds in (dimension_scores or []) if ds.get("score")]
    return max(scores) if scores else None


# ═════════════════════════════════════════════════════════════════════════════
# RISK APPETITE
# ═════════════════════════════════════════════════════════════════════════════

def list_appetite():
    db = get_db()
    try:
        return _dicts(db.execute(
            "SELECT a.*, u.full_name AS approver_name "
            "FROM erm_risk_appetite a LEFT JOIN users u ON u.id=a.approved_by "
            "ORDER BY category"
        ).fetchall())
    finally:
        db.close()


def get_appetite(appetite_id):
    db = get_db()
    try:
        return _dict(db.execute("SELECT * FROM erm_risk_appetite WHERE id=%s", (appetite_id,)).fetchone())
    finally:
        db.close()


def upsert_appetite(data):
    """Create or update appetite for a category."""
    db = get_db()
    try:
        existing = db.execute(
            "SELECT id FROM erm_risk_appetite WHERE category=%s", (data.get("category"),)
        ).fetchone()
        if existing:
            fields, vals = [], []
            for k in ("appetite_level", "max_score", "description", "tolerance_notes",
                      "approved_by", "effective_date", "review_date"):
                if k in data:
                    fields.append(f"{k}=%s"); vals.append(data[k])
            if fields:
                fields.append("updated_at=%s"); vals.append(_now())
                vals.append(existing[0])
                db.execute(f"UPDATE erm_risk_appetite SET {','.join(fields)} WHERE id=%s", vals)
            db.commit()
            return existing[0]
        else:
            cur = insert_returning_id(db,
                "INSERT INTO erm_risk_appetite (category, appetite_level, max_score, description, "
                "tolerance_notes, approved_by, effective_date, review_date) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                (data.get("category"), data.get("appetite_level", "medium"),
                 data.get("max_score", 12), data.get("description"),
                 data.get("tolerance_notes"), data.get("approved_by"),
                 data.get("effective_date"), data.get("review_date")),
            )
            db.commit()
            return cur
    finally:
        db.close()


def update_appetite(appetite_id, data):
    """Update an existing appetite record by ID."""
    db = get_db()
    try:
        fields, vals = [], []
        for k in ("category", "appetite_level", "max_score", "description",
                  "tolerance_notes", "approved_by", "effective_date", "review_date"):
            if k in data:
                fields.append(f"{k}=%s"); vals.append(data[k])
        if fields:
            fields.append("updated_at=%s"); vals.append(_now())
            vals.append(appetite_id)
            db.execute(f"UPDATE erm_risk_appetite SET {','.join(fields)} WHERE id=%s", vals)
            db.commit()
    finally:
        db.close()


def delete_appetite(appetite_id):
    db = get_db()
    try:
        db.execute("DELETE FROM erm_risk_appetite WHERE id=%s", (appetite_id,))
        db.commit()
    finally:
        db.close()


def get_appetite_status():
    """For each category: current max score in erm_enterprise_risks vs appetite threshold."""
    db = get_db()
    try:
        appetites = _dicts(db.execute("SELECT * FROM erm_risk_appetite").fetchall())
        result = []
        for a in appetites:
            cat = a["category"]
            max_risk = db.execute(
                "SELECT MAX(likelihood*impact) FROM erm_enterprise_risks "
                "WHERE category=%s AND status NOT IN ('closed','accepted')", (cat,)
            ).fetchone()[0] or 0
            count_open = db.execute(
                "SELECT COUNT(*) FROM erm_enterprise_risks "
                "WHERE category=%s AND status NOT IN ('closed','accepted')", (cat,)
            ).fetchone()[0]
            a["current_max_score"] = max_risk
            a["open_count"] = count_open
            a["breached"] = max_risk > a["max_score"]
            # Top risk in this category
            top = db.execute(
                "SELECT id, title, (likelihood*impact) AS score FROM erm_enterprise_risks "
                "WHERE category=%s AND status NOT IN ('closed','accepted') "
                "ORDER BY (likelihood*impact) DESC LIMIT 1", (cat,)
            ).fetchone()
            a["top_risk"] = _dict(top)
            result.append(a)
        return result
    finally:
        db.close()


def mark_appetite_notified(appetite_id, is_breached: bool):
    """Set or clear last_breach_notified_at for a risk appetite row.

    Call with is_breached=True after emitting a breach event.
    Call with is_breached=False when the breach clears, so the next breach fires again.
    """
    db = get_db()
    try:
        value = _now() if is_breached else None
        db.execute(
            "UPDATE erm_risk_appetite SET last_breach_notified_at=%s WHERE id=%s",
            (value, appetite_id),
        )
        db.commit()
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# RISK LIBRARY
# ═════════════════════════════════════════════════════════════════════════════

def list_library(category=None, industry=None, limit=200):
    db = get_db()
    try:
        where, params = ["is_active=1"], []
        if category:
            where.append("category=%s"); params.append(category)
        if industry:
            where.append("(applicable_industries LIKE %s OR applicable_industries='all')")
            params.append(f"%{industry}%")
        clause = "WHERE " + " AND ".join(where)
        return _dicts(db.execute(
            f"SELECT * FROM erm_risk_library {clause} ORDER BY category, title LIMIT %s",
            params + [limit],
        ).fetchall())
    finally:
        db.close()


def get_library_item(item_id):
    db = get_db()
    try:
        return _dict(db.execute("SELECT * FROM erm_risk_library WHERE id=%s", (item_id,)).fetchone())
    finally:
        db.close()


def update_library_item(item_id, data):
    db = get_db()
    try:
        fields, vals = [], []
        for k in ("title", "description", "category", "default_likelihood", "default_impact",
                  "typical_treatment", "suggested_controls", "applicable_industries",
                  "regulatory_references", "tags", "is_active"):
            if k in data:
                fields.append(f"{k}=%s"); vals.append(data[k])
        if fields:
            vals.append(item_id)
            db.execute(f"UPDATE erm_risk_library SET {','.join(fields)} WHERE id=%s", vals)
            db.commit()
    finally:
        db.close()


def delete_library_item(item_id):
    db = get_db()
    try:
        db.execute("UPDATE erm_risk_library SET is_active=0 WHERE id=%s", (item_id,))
        db.commit()
    finally:
        db.close()


def create_library_item(data):
    db = get_db()
    try:
        cur = insert_returning_id(db,
            "INSERT INTO erm_risk_library (title, description, category, default_likelihood, "
            "default_impact, typical_treatment, suggested_controls, applicable_industries, "
            "regulatory_references, tags) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (data.get("title"), data.get("description"), data.get("category"),
             data.get("default_likelihood", 3), data.get("default_impact", 3),
             data.get("typical_treatment", "mitigate"), data.get("suggested_controls"),
             data.get("applicable_industries"), data.get("regulatory_references"),
             data.get("tags")),
        )
        db.commit()
        return cur
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# REGULATORY OBLIGATIONS
# ═════════════════════════════════════════════════════════════════════════════

def list_obligations(status=None, regulator=None, limit=500):
    db = get_db()
    try:
        where, params = [], []
        if status:
            where.append("o.status=%s"); params.append(status)
        if regulator:
            where.append("o.regulator=%s"); params.append(regulator)
        clause = ("WHERE " + " AND ".join(where)) if where else ""
        return _dicts(db.execute(
            f"SELECT o.*, u.full_name AS owner_name "
            f"FROM erm_regulatory_obligations o LEFT JOIN users u ON u.id=o.owner_id "
            f"{clause} ORDER BY o.due_date ASC NULLS LAST, o.regulator LIMIT %s",
            params + [limit],
        ).fetchall())
    finally:
        db.close()


def get_obligation(obl_id):
    db = get_db()
    try:
        return _dict(db.execute("SELECT * FROM erm_regulatory_obligations WHERE id=%s", (obl_id,)).fetchone())
    finally:
        db.close()


def create_obligation(data):
    db = get_db()
    try:
        cur = insert_returning_id(db,
            "INSERT INTO erm_regulatory_obligations "
            "(regulator, regulation_name, obligation, applicable_departments, evidence_required, "
            "owner_id, due_date, status, linked_controls, linked_erm_risk_id, notes, created_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (data.get("regulator"), data.get("regulation_name"), data.get("obligation"),
             data.get("applicable_departments"), data.get("evidence_required"),
             data.get("owner_id"), data.get("due_date"), data.get("status", "pending"),
             data.get("linked_controls"), data.get("linked_erm_risk_id"),
             data.get("notes"), data.get("created_by")),
        )
        db.commit()
        return cur
    finally:
        db.close()


def update_obligation(obl_id, data):
    db = get_db()
    try:
        fields, vals = [], []
        for k in ("regulator", "regulation_name", "obligation", "applicable_departments",
                  "evidence_required", "owner_id", "due_date", "status",
                  "linked_controls", "linked_erm_risk_id", "notes"):
            if k in data:
                fields.append(f"{k}=%s"); vals.append(data[k])
        if fields:
            fields.append("updated_at=%s"); vals.append(_now()); vals.append(obl_id)
            db.execute(f"UPDATE erm_regulatory_obligations SET {','.join(fields)} WHERE id=%s", vals)
            db.commit()
    finally:
        db.close()


def delete_obligation(obl_id):
    db = get_db()
    try:
        db.execute("DELETE FROM erm_regulatory_obligations WHERE id=%s", (obl_id,))
        db.commit()
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# SELF-ASSESSMENTS
# ═════════════════════════════════════════════════════════════════════════════

def list_assessments(limit=200):
    db = get_db()
    try:
        return _dicts(db.execute(
            "SELECT a.*, u.full_name AS creator_name, "
            "(SELECT COUNT(*) FROM erm_assessment_questions WHERE assessment_id=a.id) AS question_count "
            "FROM erm_assessments a LEFT JOIN users u ON u.id=a.created_by "
            "ORDER BY a.created_at DESC LIMIT %s", (limit,)
        ).fetchall())
    finally:
        db.close()


def get_assessment(assessment_id):
    db = get_db()
    try:
        a = _dict(db.execute("SELECT * FROM erm_assessments WHERE id=%s", (assessment_id,)).fetchone())
        if a:
            a["questions"] = _dicts(db.execute(
                "SELECT * FROM erm_assessment_questions WHERE assessment_id=%s ORDER BY order_idx",
                (assessment_id,)
            ).fetchall())
        return a
    finally:
        db.close()


def create_assessment(data):
    db = get_db()
    try:
        cur = insert_returning_id(db,
            "INSERT INTO erm_assessments (title, type, description, target_audience, status, due_date, created_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (data.get("title"), data.get("type", "risk"), data.get("description"),
             data.get("target_audience"), data.get("status", "draft"),
             data.get("due_date"), data.get("created_by")),
        )
        db.commit()
        return cur
    finally:
        db.close()


def update_assessment(assessment_id, data):
    db = get_db()
    try:
        fields, vals = [], []
        for k in ("title", "type", "description", "target_audience", "status", "due_date"):
            if k in data:
                fields.append(f"{k}=%s"); vals.append(data[k])
        if fields:
            fields.append("updated_at=%s"); vals.append(_now()); vals.append(assessment_id)
            db.execute(f"UPDATE erm_assessments SET {','.join(fields)} WHERE id=%s", vals)
            db.commit()
    finally:
        db.close()


def delete_assessment(assessment_id):
    db = get_db()
    try:
        db.execute("DELETE FROM erm_assessment_responses WHERE assessment_id=%s", (assessment_id,))
        db.execute("DELETE FROM erm_assessment_questions WHERE assessment_id=%s", (assessment_id,))
        db.execute("DELETE FROM erm_assessments WHERE id=%s", (assessment_id,))
        db.commit()
    finally:
        db.close()


def add_question(assessment_id, data):
    db = get_db()
    try:
        cur = insert_returning_id(db,
            "INSERT INTO erm_assessment_questions "
            "(assessment_id, question, question_type, options, weight, order_idx, required) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (assessment_id, data.get("question"), data.get("question_type", "scale"),
             data.get("options"), data.get("weight", 1.0),
             data.get("order_idx", 0), data.get("required", 1)),
        )
        db.commit()
        return cur
    finally:
        db.close()


def delete_question(question_id):
    db = get_db()
    try:
        db.execute("DELETE FROM erm_assessment_questions WHERE id=%s", (question_id,))
        db.commit()
    finally:
        db.close()


def save_response(data):
    db = get_db()
    try:
        # Upsert: delete existing response for this question+respondent then insert
        db.execute(
            "DELETE FROM erm_assessment_responses WHERE assessment_id=%s AND question_id=%s AND respondent_id=%s",
            (data.get("assessment_id"), data.get("question_id"), data.get("respondent_id")),
        )
        cur = insert_returning_id(db,
            "INSERT INTO erm_assessment_responses "
            "(assessment_id, question_id, respondent_id, response, score, notes) "
            "VALUES (%s,%s,%s,%s,%s,%s)",
            (data.get("assessment_id"), data.get("question_id"), data.get("respondent_id"),
             data.get("response"), data.get("score"), data.get("notes")),
        )
        db.commit()
        return cur
    finally:
        db.close()


def list_responses(assessment_id):
    db = get_db()
    try:
        return _dicts(db.execute(
            "SELECT r.*, q.question, q.weight, u.full_name AS respondent_name "
            "FROM erm_assessment_responses r "
            "JOIN erm_assessment_questions q ON q.id=r.question_id "
            "LEFT JOIN users u ON u.id=r.respondent_id "
            "WHERE r.assessment_id=%s ORDER BY r.question_id, r.submitted_at",
            (assessment_id,)
        ).fetchall())
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# UNIFIED RISK REGISTER VIEW (erm_enterprise_risks UNION risk_register)
# ═════════════════════════════════════════════════════════════════════════════

def get_unified_register(filters=None, limit=1000):
    """Return all ERM enterprise risks + risk_register entries as a unified list."""
    filters = filters or {}
    cat    = filters.get("category")
    status = filters.get("status")
    db = get_db()
    try:
        # ── ERM enterprise risks ──────────────────────────────────────────
        erm_where, erm_params = [], []
        if cat:    erm_where.append("category=%s");  erm_params.append(cat)
        if status: erm_where.append("status=%s");    erm_params.append(status)
        else:      erm_where.append("status != 'closed'")
        ew = ("WHERE " + " AND ".join(erm_where)) if erm_where else ""
        erm_rows = _dicts(db.execute(
            f"SELECT e.id, e.title, e.description, e.category, e.sub_category, e.likelihood, e.impact, "
            f"(e.likelihood*e.impact) AS risk_score, e.status, e.treatment, e.owner_id, "
            f"e.board_visibility, "
            f"'erm' AS register_source, e.source_module, e.created_at, "
            f"u.full_name AS owner_name "
            f"FROM erm_enterprise_risks e "
            f"LEFT JOIN users u ON u.id=e.owner_id "
            f"{ew} ORDER BY (e.likelihood*e.impact) DESC LIMIT %s",
            erm_params + [limit]
        ).fetchall())

        # ── Platform risk_register ────────────────────────────────────────
        rr_where, rr_params = [], []
        if cat:    rr_where.append("r.category=%s");  rr_params.append(cat)
        if status: rr_where.append("r.status=%s");    rr_params.append(status)
        else:      rr_where.append("r.status != 'closed'")
        rw = ("WHERE " + " AND ".join(rr_where)) if rr_where else ""
        rr_rows = _dicts(db.execute(
            f"SELECT r.id, r.title, r.description, r.category, r.likelihood, r.impact, "
            f"r.risk_score, r.status, r.treatment, r.owner_id, "
            f"0 AS board_visibility, "
            f"'risk_register' AS register_source, r.source_module, r.created_at, "
            f"u.full_name AS owner_name "
            f"FROM risk_register r "
            f"LEFT JOIN users u ON u.id=r.owner_id "
            f"{rw} ORDER BY r.risk_score DESC LIMIT %s",
            rr_params + [limit]
        ).fetchall())

        all_risks = erm_rows + rr_rows
        all_risks.sort(key=lambda x: x.get("risk_score") or 0, reverse=True)
        return all_risks[:limit]
    finally:
        db.close()


def get_register_stats():
    """Stats for ERM dashboard: by_level, by_module, by_category, heat_map."""
    db = get_db()
    try:
        fw_matrix = get_active_framework_matrix(db)

        # All open risks from both tables
        erm = _dicts(db.execute(
            "SELECT likelihood, impact, (likelihood*impact) AS score, category, source_module "
            "FROM erm_enterprise_risks WHERE status NOT IN ('closed','accepted')"
        ).fetchall())
        rr = _dicts(db.execute(
            "SELECT likelihood, impact, risk_score AS score, category, source_module "
            "FROM risk_register WHERE status NOT IN ('closed')"
        ).fetchall())
        all_risks = erm + rr

        by_level = {b: 0 for b in fw_matrix["bands"]}
        by_module = {}
        by_category = {}
        heat_map = [[0]*5 for _ in range(5)]

        for r in all_risks:
            lvl = resolve_band(fw_matrix, r.get("likelihood"), r.get("impact"))
            by_level[lvl] = by_level.get(lvl, 0) + 1
            mod = r.get("source_module") or "erm"
            by_module[mod] = by_module.get(mod, 0) + 1
            cat = r.get("category") or "operational"
            by_category[cat] = by_category.get(cat, 0) + 1
            l = min(max((r.get("likelihood") or 1) - 1, 0), 4)
            i = min(max((r.get("impact") or 1) - 1, 0), 4)
            heat_map[l][i] += 1

        return {
            "total": len(all_risks),
            "by_level": by_level,
            "by_module": by_module,
            "by_category": by_category,
            "heat_map": heat_map,
        }
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# RISK FEED (recent events from all modules)
# ═════════════════════════════════════════════════════════════════════════════

def get_risk_feed(limit=20):
    """Recent risk-related events from the platform events table."""
    db = get_db()
    try:
        rows = _dicts(db.execute(
            "SELECT event_type, source_module, source_entity_type, source_entity_id, "
            "payload, created_at FROM events "
            "WHERE event_type LIKE '%%.risk%%' OR event_type LIKE '%%breach%%' "
            "OR event_type LIKE '%%incident%%' OR event_type LIKE '%%escalat%%' "
            "OR event_type LIKE 'erm.%%' OR event_type LIKE 'orm.%%' "
            "ORDER BY created_at DESC LIMIT %s",
            (limit,)
        ).fetchall())
        return rows
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# DASHBOARD STATS
# ═════════════════════════════════════════════════════════════════════════════

def get_dashboard_stats():
    db = get_db()
    try:
        total_erm = db.execute("SELECT COUNT(*) FROM erm_enterprise_risks WHERE status!='closed'").fetchone()[0]
        critical = db.execute(
            f"SELECT COUNT(*) FROM erm_enterprise_risks e {_FW_BAND_JOIN} "
            "WHERE mb.band_key='critical' AND e.status!='closed'"
        ).fetchone()[0]
        high = db.execute(
            f"SELECT COUNT(*) FROM erm_enterprise_risks e {_FW_BAND_JOIN} "
            "WHERE mb.band_key IN ('high','critical') AND e.status!='closed'"
        ).fetchone()[0]
        total_rr = db.execute("SELECT COUNT(*) FROM risk_register WHERE status!='closed'").fetchone()[0]
        appetite_breaches = db.execute(
            "SELECT COUNT(*) FROM erm_enterprise_risks e "
            "JOIN erm_risk_appetite a ON a.category=e.category "
            "WHERE (e.likelihood*e.impact) > a.max_score AND e.status!='closed'"
        ).fetchone()[0]
        overdue_obligations = db.execute(
            "SELECT COUNT(*) FROM erm_regulatory_obligations "
            f"WHERE status NOT IN ('compliant') AND due_date < {sql_current_date()}"
        ).fetchone()[0]
        open_assessments = db.execute(
            "SELECT COUNT(*) FROM erm_assessments WHERE status='active'"
        ).fetchone()[0]
        board_visible = db.execute(
            "SELECT COUNT(*) FROM erm_enterprise_risks WHERE board_visibility=1 AND status!='closed'"
        ).fetchone()[0]

        # Trend: compare current open count to 30 days ago
        erm_30d = db.execute(
            "SELECT COUNT(*) FROM erm_enterprise_risks WHERE status!='closed' "
            f"AND created_at < {sql_now_ts('-30 days')}"
        ).fetchone()[0]
        rr_30d = db.execute(
            "SELECT COUNT(*) FROM risk_register WHERE status!='closed' "
            f"AND created_at < {sql_now_ts('-30 days')}"
        ).fetchone()[0]
        total_30d = erm_30d + rr_30d
        trend_total = (total_erm + total_rr) - total_30d

        crit_30d = db.execute(
            f"SELECT COUNT(*) FROM erm_enterprise_risks e {_FW_BAND_JOIN} "
            "WHERE mb.band_key='critical' AND e.status!='closed' "
            f"AND e.created_at < {sql_now_ts('-30 days')}"
        ).fetchone()[0]
        trend_critical = critical - crit_30d

        # Top 5 critical risks
        top_critical = _dicts(db.execute(
            "SELECT e.id, e.title, e.category, e.likelihood, e.impact, "
            "(e.likelihood*e.impact) AS score, e.source_module, e.status, "
            "'erm' AS register_source "
            "FROM erm_enterprise_risks e "
            "WHERE e.status IN ('open','under_review') "
            "ORDER BY (e.likelihood*e.impact) DESC LIMIT 5"
        ).fetchall())

        # Actions required: unowned + past-review-date + overdue obligations
        actions = []
        # Unowned ERM risks (no owner_id)
        unowned = _dicts(db.execute(
            "SELECT id, title, 'unowned_risk' AS action_type FROM erm_enterprise_risks "
            "WHERE owner_id IS NULL AND status IN ('open','under_review') LIMIT 5"
        ).fetchall())
        for r in unowned:
            actions.append({"type": "unowned_risk", "id": r["id"],
                            "text": f"Unowned risk: {r['title']}", "link": "/erm/register"})

        # Past review date
        overdue_review = _dicts(db.execute(
            "SELECT id, title FROM erm_enterprise_risks "
            f"WHERE review_date < {sql_current_date()} AND status IN ('open','under_review') "
            "AND review_date IS NOT NULL LIMIT 5"
        ).fetchall())
        for r in overdue_review:
            actions.append({"type": "overdue_review", "id": r["id"],
                            "text": f"Review overdue: {r['title']}", "link": "/erm/register"})

        # Overdue obligations
        overdue_obls = _dicts(db.execute(
            "SELECT id, regulation_name FROM erm_regulatory_obligations "
            f"WHERE status NOT IN ('compliant') AND due_date < {sql_current_date()} LIMIT 5"
        ).fetchall())
        for o in overdue_obls:
            actions.append({"type": "overdue_obligation", "id": o["id"],
                            "text": f"Overdue obligation: {o['regulation_name']}",
                            "link": "/erm/obligations"})

        return {
            "total_enterprise_risks": total_erm,
            "total_register_risks": total_rr,
            "total_risks": total_erm + total_rr,
            "critical": critical,
            "high": high,
            "appetite_breaches": appetite_breaches,
            "overdue_obligations": overdue_obligations,
            "open_assessments": open_assessments,
            "board_visible": board_visible,
            "trend_total": trend_total,
            "trend_critical": trend_critical,
            "top_critical_risks": top_critical,
            "actions_required": actions[:8],
        }
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# CHAT
# ═════════════════════════════════════════════════════════════════════════════

def list_chat(user_id, limit=50):
    db = get_db()
    try:
        rows = _dicts(db.execute(
            "SELECT * FROM erm_chat_messages WHERE user_id=%s ORDER BY created_at DESC LIMIT %s",
            (user_id, limit)
        ).fetchall())
        rows.reverse()
        return rows
    finally:
        db.close()


def save_chat(user_id, role, content, provider=None):
    db = get_db()
    try:
        db.execute(
            "INSERT INTO erm_chat_messages (user_id, role, content, provider) VALUES (%s,%s,%s,%s)",
            (user_id, role, content, provider)
        )
        db.commit()
    finally:
        db.close()


def clear_chat(user_id):
    db = get_db()
    try:
        db.execute("DELETE FROM erm_chat_messages WHERE user_id=%s", (user_id,))
        db.commit()
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# KEY RISK INDICATORS (KRIs)
# ═════════════════════════════════════════════════════════════════════════════

def list_kris(linked_risk_id=None):
    db = get_db()
    try:
        where, params = [], []
        if linked_risk_id:
            where.append("k.linked_risk_id=%s"); params.append(linked_risk_id)
        clause = ("WHERE " + " AND ".join(where)) if where else ""
        rows = db.execute(
            f"SELECT k.*, u.full_name AS owner_name, r.title AS risk_title "
            f"FROM erm_kris k "
            f"LEFT JOIN users u ON u.id=k.owner_id "
            f"LEFT JOIN erm_enterprise_risks r ON r.id=k.linked_risk_id "
            f"{clause} ORDER BY k.name",
            params
        ).fetchall()
        return _dicts(rows)
    finally:
        db.close()


def get_kri(kri_id):
    db = get_db()
    try:
        return _dict(db.execute(
            "SELECT k.*, u.full_name AS owner_name, r.title AS risk_title "
            "FROM erm_kris k "
            "LEFT JOIN users u ON u.id=k.owner_id "
            "LEFT JOIN erm_enterprise_risks r ON r.id=k.linked_risk_id "
            "WHERE k.id=%s", (kri_id,)
        ).fetchone())
    finally:
        db.close()


def create_kri(data):
    db = get_db()
    try:
        cur = insert_returning_id(db,
            "INSERT INTO erm_kris (name, description, linked_risk_id, metric_type, "
            "threshold_warn, threshold_crit, current_value, unit, frequency, owner_id) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (data.get("name"), data.get("description"), data.get("linked_risk_id"),
             data.get("metric_type", "count"), data.get("threshold_warn"),
             data.get("threshold_crit"), data.get("current_value", 0),
             data.get("unit"), data.get("frequency", "monthly"), data.get("owner_id"))
        )
        db.commit()
        return cur
    finally:
        db.close()


def update_kri(kri_id, data):
    db = get_db()
    try:
        fields, vals = [], []
        for k in ("name", "description", "linked_risk_id", "metric_type",
                  "threshold_warn", "threshold_crit", "unit", "frequency",
                  "owner_id", "status", "trend"):
            if k in data:
                fields.append(f"{k}=%s"); vals.append(data[k])
        # Handle value update: also record history
        if "current_value" in data:
            fields.append("current_value=%s"); vals.append(data["current_value"])
            fields.append("last_updated=%s"); vals.append(_now())
            db.execute(
                "INSERT INTO erm_kri_history (kri_id, value) VALUES (%s,%s)",
                (kri_id, data["current_value"])
            )
        if fields:
            fields.append("updated_at=%s"); vals.append(_now()); vals.append(kri_id)
            db.execute(f"UPDATE erm_kris SET {','.join(fields)} WHERE id=%s", vals)
            db.commit()
    finally:
        db.close()


def delete_kri(kri_id):
    db = get_db()
    try:
        db.execute("DELETE FROM erm_kris WHERE id=%s", (kri_id,))
        db.commit()
    finally:
        db.close()


def get_kri_history(kri_id, limit=12):
    db = get_db()
    try:
        return _dicts(db.execute(
            "SELECT * FROM erm_kri_history WHERE kri_id=%s ORDER BY recorded_at DESC LIMIT %s",
            (kri_id, limit)
        ).fetchall())
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# WORKFLOW HISTORY
# ═════════════════════════════════════════════════════════════════════════════

_WORKFLOW_STEPS = ["draft", "identified", "assessed", "treated", "monitored", "closed"]


def transition_workflow(risk_id, to_step, user_id, notes=None):
    """Advance a risk to the next workflow step. Returns the new step or raises ValueError."""
    db = get_db()
    try:
        risk = db.execute(
            "SELECT workflow_step FROM erm_enterprise_risks WHERE id=%s", (risk_id,)
        ).fetchone()
        if not risk:
            raise ValueError("Risk not found")
        from_step = risk["workflow_step"] or "draft"
        if to_step not in _WORKFLOW_STEPS:
            raise ValueError(f"Invalid step: {to_step}")
        from_idx = _WORKFLOW_STEPS.index(from_step) if from_step in _WORKFLOW_STEPS else 0
        to_idx   = _WORKFLOW_STEPS.index(to_step)
        if to_idx > from_idx + 1:
            next_step = _WORKFLOW_STEPS[from_idx + 1]
            raise ValueError(
                f"Cannot skip steps: complete '{next_step}' before moving to '{to_step}'"
            )
        # Step-status map (unchanged)
        step_status_map = {
            "closed":   "closed",
            "treated":  "mitigated",
            "assessed": "under_review",
        }
        new_status = step_status_map.get(to_step)
        _now_str = _now()
        if new_status:
            cur = db.execute(
                "UPDATE erm_enterprise_risks SET workflow_step=%s, status=%s, updated_at=%s "
                "WHERE id=%s AND COALESCE(workflow_step,'draft')=%s",
                (to_step, new_status, _now_str, risk_id, from_step)
            )
        else:
            cur = db.execute(
                "UPDATE erm_enterprise_risks SET workflow_step=%s, updated_at=%s "
                "WHERE id=%s AND COALESCE(workflow_step,'draft')=%s",
                (to_step, _now_str, risk_id, from_step)
            )
        if cur.rowcount == 0:
            db.rollback()
            raise ValueError("Risk was updated by someone else — reload and try again")
        # History INSERT now AFTER the successful conditional update
        db.execute(
            "INSERT INTO erm_risk_workflow_history (risk_id, from_step, to_step, changed_by, notes) "
            "VALUES (%s,%s,%s,%s,%s)", (risk_id, from_step, to_step, user_id, notes)
        )
        db.commit()
        return to_step
    finally:
        db.close()


def get_workflow_history(risk_id):
    db = get_db()
    try:
        return _dicts(db.execute(
            "SELECT h.*, u.full_name AS changed_by_name "
            "FROM erm_risk_workflow_history h "
            "LEFT JOIN users u ON u.id=h.changed_by "
            "WHERE h.risk_id=%s ORDER BY h.changed_at ASC",
            (risk_id,)
        ).fetchall())
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# RISK STATEMENT LIBRARY
# ═════════════════════════════════════════════════════════════════════════════

def list_statements(category=None, tags=None, limit=200):
    db = get_db()
    try:
        where, params = [], []
        if category:
            where.append("category=%s"); params.append(category)
        if tags:
            where.append("tags LIKE %s"); params.append(f"%{tags}%")
        clause = ("WHERE " + " AND ".join(where)) if where else ""
        return _dicts(db.execute(
            f"SELECT * FROM erm_risk_statements {clause} "
            f"ORDER BY usage_count DESC, created_at DESC LIMIT %s",
            params + [limit]
        ).fetchall())
    finally:
        db.close()


def create_statement(data):
    db = get_db()
    try:
        full = data.get("full_statement") or (
            f"Due to {data.get('cause', '')}, there is a risk that "
            f"{data.get('event', '')}, resulting in {data.get('consequence', '')}."
        )
        cur = insert_returning_id(db,
            "INSERT INTO erm_risk_statements (category, cause, event, consequence, "
            "full_statement, tags, industry) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (data.get("category"), data.get("cause"), data.get("event"),
             data.get("consequence"), full, data.get("tags"), data.get("industry"))
        )
        db.commit()
        return cur
    finally:
        db.close()


def update_statement(stmt_id, data):
    db = get_db()
    try:
        fields, vals = [], []
        for k in ("category", "cause", "event", "consequence", "full_statement", "tags", "industry"):
            if k in data:
                fields.append(f"{k}=%s"); vals.append(data[k])
        if fields:
            vals.append(stmt_id)
            db.execute(f"UPDATE erm_risk_statements SET {','.join(fields)} WHERE id=%s", vals)
            db.commit()
    finally:
        db.close()


def delete_statement(stmt_id):
    db = get_db()
    try:
        db.execute("DELETE FROM erm_risk_statements WHERE id=%s", (stmt_id,))
        db.commit()
    finally:
        db.close()


def use_statement(stmt_id):
    """Increment usage count and return the statement."""
    db = get_db()
    try:
        db.execute(
            "UPDATE erm_risk_statements SET usage_count=usage_count+1 WHERE id=%s", (stmt_id,)
        )
        db.commit()
        return _dict(db.execute("SELECT * FROM erm_risk_statements WHERE id=%s", (stmt_id,)).fetchone())
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# REPORTING DATA
# ═════════════════════════════════════════════════════════════════════════════

def get_trend_data(period_days=30):
    """Return weekly risk count snapshots over the given period."""
    db = get_db()
    try:
        rows = db.execute(
            f"SELECT date(created_at) AS day, COUNT(*) AS new_risks "
            f"FROM erm_enterprise_risks "
            f"WHERE created_at >= {sql_date_ts(f'-{int(period_days)} days')} "
            f"GROUP BY day ORDER BY day ASC"
        ).fetchall()
        # Also get total open per day using current status snapshot
        total_open = db.execute(
            "SELECT COUNT(*) FROM erm_enterprise_risks WHERE status NOT IN ('closed','accepted')"
        ).fetchone()[0]
        critical_open = db.execute(
            f"SELECT COUNT(*) FROM erm_enterprise_risks e {_FW_BAND_JOIN} "
            "WHERE e.status NOT IN ('closed','accepted') AND mb.band_key='critical'"
        ).fetchone()[0]
        return {
            "daily": _dicts(rows),
            "total_open": total_open,
            "critical_open": critical_open,
            "period_days": period_days
        }
    finally:
        db.close()


def get_risk_aging():
    """Bucket open risks by how long they have been open."""
    db = get_db()
    try:
        rows = db.execute(
            "SELECT id, title, category, qualitative_score, "
            f"""CAST({sql_days_between("'now'", "created_at")} AS INTEGER) AS days_open """
            "FROM erm_enterprise_risks "
            "WHERE status NOT IN ('closed','accepted') ORDER BY days_open DESC"
        ).fetchall()
        buckets = {"lt30": [], "d30_90": [], "d90_180": [], "gt180": []}
        for r in rows:
            d = r["days_open"] or 0
            key = "lt30" if d < 30 else "d30_90" if d < 90 else "d90_180" if d < 180 else "gt180"
            buckets[key].append(dict(r))
        return {k: {"count": len(v), "risks": v[:5]} for k, v in buckets.items()}
    finally:
        db.close()


def get_executive_dashboard():
    """Board-level view: board_visibility=1 risks + appetite status + top 3 actions."""
    db = get_db()
    try:
        board_risks = _dicts(db.execute(
            "SELECT e.*, u.full_name AS owner_name "
            "FROM erm_enterprise_risks e "
            "LEFT JOIN users u ON u.id=e.owner_id "
            "WHERE e.board_visibility=1 AND e.status NOT IN ('closed','accepted') "
            "ORDER BY (e.likelihood*e.impact) DESC LIMIT 10"
        ).fetchall())
        appetite = _dicts(db.execute(
            "SELECT a.*, "
            "(SELECT MAX(likelihood*impact) FROM erm_enterprise_risks "
            " WHERE category=a.category AND status NOT IN ('closed','accepted')) AS current_score "
            "FROM erm_risk_appetite a ORDER BY category"
        ).fetchall())
        for a in appetite:
            a["breached"] = bool(a.get("current_score") and a["current_score"] > a["max_score"])
        actions = _dicts(db.execute(
            "SELECT id, title, category, qualitative_score, owner_id "
            "FROM erm_enterprise_risks "
            f"WHERE (owner_id IS NULL OR review_date < {sql_current_date()}) "
            "AND status NOT IN ('closed','accepted') LIMIT 3"
        ).fetchall())
        total = db.execute(
            "SELECT COUNT(*) FROM erm_enterprise_risks WHERE status NOT IN ('closed','accepted')"
        ).fetchone()[0]
        critical = db.execute(
            f"SELECT COUNT(*) FROM erm_enterprise_risks e {_FW_BAND_JOIN} "
            "WHERE e.status NOT IN ('closed','accepted') AND mb.band_key='critical'"
        ).fetchone()[0]
        return {
            "board_risks": board_risks, "appetite": appetite,
            "actions": actions, "total_open": total, "critical": critical
        }
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL RISK REGISTER IMPORT
# ═════════════════════════════════════════════════════════════════════════════

_log = logging.getLogger(__name__)


def _norm_header(s):
    return re.sub(r'[^a-z0-9]', '', str(s).lower().strip())


_HEADER_MAP = {
    "risktitle": "title", "title": "title", "riskname": "title", "name": "title",
    "risk": "title",
    "riskdescription": "description", "description": "description", "details": "description",
    "riskimpact": "impact_description", "impactdescription": "impact_description",
    "mitigation": "treatment_plan", "mitigationplan": "treatment_plan",
    "controls": "treatment_plan", "existingcontrols": "treatment_plan",
    "controlmeasures": "treatment_plan", "mitigatingcontrols": "treatment_plan",
    "section": "sub_category", "department": "sub_category", "businessunit": "sub_category",
    "unit": "sub_category", "division": "sub_category", "area": "sub_category",
    "priority": "priority", "urgency": "priority",
    "likelihood": "likelihood", "probability": "likelihood", "frequency": "likelihood",
    "impact": "impact", "consequence": "impact", "severity": "impact",
    "aggregateinherentrisk": "_skip", "inherentrisk": "_skip", "riskscore": "_skip",
    "inherentscore": "_skip", "grossrisk": "_skip",
    "controleffectiveness": "effectiveness_rating", "effectiveness": "effectiveness_rating",
    "controlstrength": "effectiveness_rating",
    "aggregateresidual": "_skip", "residualrisk": "_skip", "residualscore": "_skip",
    "netrisk": "_skip", "residual": "_skip",
    "riskmitigation": "treatment", "riskmitigationplan": "treatment",
    "risktreatment": "treatment", "treatment": "treatment", "response": "treatment",
    "riskresponse": "treatment", "treatmentstrategy": "treatment",
    "reviewdate": "review_frequency", "reviewfrequency": "review_frequency",
    "followupdate": "review_date", "nextreview": "review_date", "duedate": "review_date",
    "nextreviewdate": "review_date", "followup": "review_date", "targetdate": "review_date",
    "statusupdate": "status", "status": "status", "riskstatus": "status",
    "responsibility": "owner_name", "owner": "owner_name", "riskowner": "owner_name",
    "responsible": "owner_name", "assignedto": "owner_name", "accountable": "owner_name",
    "category": "category", "riskcategory": "category", "risktype": "category",
    "subcategory": "sub_category",
    "boardvisibility": "board_visibility", "boardvisible": "board_visibility",
    "strategicobjective": "strategic_objective",
    "regulationlinks": "regulation_links", "regulation": "regulation_links",
    "residuallikelihood": "residual_likelihood", "residualimpact": "residual_impact",
    "riskstatement": "risk_statement",
}

_PRIORITY_MAP = {"p1": 5, "p2": 4, "p3": 3, "high": 5, "critical": 5, "medium": 3, "low": 1}

_TREATMENT_TEXT = {
    "risk mitigation": "mitigate", "mitigate": "mitigate", "mitigation": "mitigate",
    "risk mitigatio": "mitigate", "reduce": "mitigate", "control": "mitigate",
    "risk avoidance": "avoid", "avoid": "avoid", "avoidance": "avoid", "terminate": "avoid",
    "risk transfer": "transfer", "transfer": "transfer", "share": "transfer", "insure": "transfer",
    "risk acceptance": "accept", "accept": "accept", "acceptance": "accept", "tolerate": "accept",
    "risk treatment": "mitigate",
}

_STATUS_TEXT = {
    "wip": "open", "work in progress": "open", "in progress": "open", "open": "open",
    "new": "open", "draft": "open", "identified": "open", "active": "open",
    "under review": "under_review", "underreview": "under_review", "review": "under_review",
    "mitigated": "mitigated", "controlled": "mitigated", "treated": "mitigated",
    "accepted": "accepted", "tolerated": "accepted",
    "closed": "closed", "complete": "closed", "done": "closed", "resolved": "closed",
}

_CATEGORY_KEYWORDS = {
    "revenue": "Financial Risk", "financial": "Financial Risk", "finance": "Financial Risk",
    "credit": "Credit Risk", "market": "Market Risk",
    "infrastructure": "Technology Risk", "technology": "Technology Risk",
    "it": "Technology Risk", "cyber": "Technology Risk", "information security": "Technology Risk",
    "people": "Operational Risk", "hr": "Operational Risk", "process": "Operational Risk",
    "operational": "Operational Risk",
    "supply chain": "Third Party Risk", "vendor": "Third Party Risk",
    "third party": "Third Party Risk", "supplier": "Third Party Risk",
    "compliance": "Compliance & Legal Risk", "legal": "Compliance & Legal Risk",
    "regulatory": "Compliance & Legal Risk",
    "strategic": "Strategic Risk", "strategy": "Strategic Risk",
    "reputational": "Reputational Risk", "reputation": "Reputational Risk", "brand": "Reputational Risk",
    "environmental": "Environmental Risk", "climate": "Environmental Risk",
    "suply": "Third Party Risk", "supply": "Third Party Risk",
    "revenue": "Financial Risk", "income": "Financial Risk", "profit": "Financial Risk",
}


def _get_taxonomy_categories():
    db = get_db()
    try:
        row = db.execute("SELECT id FROM erm_risk_frameworks WHERE is_active=1 LIMIT 1").fetchone()
        if not row:
            return []
        fw_id = row[0] if isinstance(row, (tuple, list)) else row["id"]
        rows = db.execute(
            "SELECT name FROM erm_framework_taxonomy WHERE framework_id=%s AND parent_id IS NULL ORDER BY order_idx",
            (fw_id,),
        ).fetchall()
        return [r[0] if isinstance(r, (tuple, list)) else r["name"] for r in rows]
    finally:
        db.close()


def _fuzzy_match_category(raw, taxonomy_cats):
    if not raw or not raw.strip():
        return ("Operational Risk", "default")
    clean = raw.strip()
    lower = clean.lower()
    for tc in taxonomy_cats:
        if tc.lower() == lower:
            return (tc, "exact")
    matches = get_close_matches(clean, taxonomy_cats, n=1, cutoff=0.45)
    if matches:
        return (matches[0], "fuzzy")
    for kw, cat in _CATEGORY_KEYWORDS.items():
        if kw in lower:
            if cat in taxonomy_cats:
                return (cat, "keyword")
    return (clean, "unmapped")


def _fuzzy_match_owner(raw_name, users):
    if not raw_name or not raw_name.strip():
        return (None, "", "empty")
    clean = raw_name.strip()
    lower = clean.lower()
    for u in users:
        if (u.get("full_name") or "").lower() == lower:
            return (u["id"], u["full_name"], "exact")
    for u in users:
        fn = (u.get("full_name") or "").lower()
        if lower in fn or fn in lower:
            return (u["id"], u["full_name"], "partial")
    names = [u.get("full_name", "") for u in users]
    matches = get_close_matches(clean, names, n=1, cutoff=0.5)
    if matches:
        uid = next((u["id"] for u in users if u.get("full_name") == matches[0]), None)
        return (uid, matches[0], "fuzzy")
    return (None, clean, "unmatched")


def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _clamp(val, lo, hi, default):
    try:
        v = int(val)
        return max(lo, min(hi, v))
    except (TypeError, ValueError):
        return default


def parse_risk_register_excel(file_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    taxonomy_cats = _get_taxonomy_categories()
    db = get_db()
    try:
        users = _dicts(db.execute("SELECT id, full_name, email FROM users").fetchall())
    finally:
        db.close()

    header_row = None
    header_map = {}
    raw_headers = []

    for row_idx in range(1, min(ws.max_row + 1, 20)):
        cells = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        normed = [_norm_header(c) for c in cells if c is not None]
        matched = sum(1 for n in normed if n in _HEADER_MAP)
        if matched >= 3:
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row_idx, col_idx).value
                if val is None:
                    continue
                n = _norm_header(val)
                field = _HEADER_MAP.get(n)
                if field and field != "_skip":
                    header_map[col_idx] = field
            raw_headers = [str(ws.cell(row_idx, c).value or "").strip() for c in range(1, ws.max_column + 1)]
            header_row = row_idx
            break

    if header_row is None:
        return {"error": "Could not detect a header row. Need at least 3 recognized column names."}

    has_cat_col = "category" in header_map.values()
    current_category = ""
    rows = []
    warnings = []
    category_map = {}
    owner_map = {}
    title_col = None
    desc_col = None
    for ci, field in header_map.items():
        if field == "title" and title_col is None:
            title_col = ci
        if field == "description" and desc_col is None:
            desc_col = ci

    if not has_cat_col and header_row:
        first_cell = str(ws.cell(header_row, 1).value or "").strip()
        first_norm = _norm_header(first_cell)
        if first_cell and first_norm not in _HEADER_MAP:
            current_category = first_cell
            matched, conf = _fuzzy_match_category(current_category, taxonomy_cats)
            category_map[current_category] = {"mapped": matched, "confidence": conf}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell_a = ws.cell(row_idx, 1).value
        cell_a_str = str(cell_a or "").strip()

        if not has_cat_col and desc_col:
            desc_val = ws.cell(row_idx, desc_col).value
            desc_norm = _norm_header(desc_val) if desc_val else ""
            if desc_norm in ("riskdescription", "description", "risktitle"):
                current_category = cell_a_str
                if current_category and current_category not in category_map:
                    matched, conf = _fuzzy_match_category(current_category, taxonomy_cats)
                    category_map[current_category] = {"mapped": matched, "confidence": conf}
                continue

        title_val = ws.cell(row_idx, title_col).value if title_col else cell_a
        title_str = str(title_val or "").strip()
        if not title_str:
            continue

        row_data = {}
        extras = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row_idx, col_idx).value
            if val is None:
                continue
            field = header_map.get(col_idx)
            if field is None:
                hdr = raw_headers[col_idx - 1] if col_idx - 1 < len(raw_headers) else f"Column {col_idx}"
                if hdr and str(val).strip():
                    extras.append(f"{hdr}: {str(val).strip()[:200]}")
                continue
            if field == "_skip":
                continue
            row_data[field] = val

        risk = {"title": title_str, "source_module": "erm", "board_visibility": 0}

        desc_parts = []
        if row_data.get("description"):
            desc_parts.append(str(row_data["description"]).strip())
        if row_data.get("impact_description"):
            desc_parts.append("Impact: " + str(row_data["impact_description"]).strip())
        if row_data.get("review_frequency"):
            freq = str(row_data["review_frequency"]).strip()
            if freq.lower() not in ("review date", ""):
                desc_parts.append("Review frequency: " + freq)
        if extras:
            desc_parts.append("--\n" + "\n".join(extras))
        risk["description"] = "\n\n".join(desc_parts) if desc_parts else ""

        if has_cat_col and row_data.get("category"):
            cat_raw = str(row_data["category"]).strip()
            if cat_raw not in category_map:
                matched, conf = _fuzzy_match_category(cat_raw, taxonomy_cats)
                category_map[cat_raw] = {"mapped": matched, "confidence": conf}
            risk["category"] = cat_raw
        elif current_category:
            risk["category"] = current_category
        else:
            risk["category"] = "Operational Risk"

        risk["sub_category"] = str(row_data.get("sub_category", "")).strip() or ""
        risk["likelihood"] = _clamp(row_data.get("likelihood"), 1, 5, 3)
        risk["impact"] = _clamp(row_data.get("impact"), 1, 5, 3)

        if risk["likelihood"] == 3 and row_data.get("likelihood") is None:
            warnings.append(f"Row {row_idx}: no likelihood value, defaulting to 3")
        if risk["impact"] == 3 and row_data.get("impact") is None:
            warnings.append(f"Row {row_idx}: no impact value, defaulting to 3")

        pri = str(row_data.get("priority", "")).strip().lower()
        risk["velocity"] = _PRIORITY_MAP.get(pri, 3)

        treat_raw = str(row_data.get("treatment", "")).strip().lower()
        risk["treatment"] = _TREATMENT_TEXT.get(treat_raw, "mitigate")

        risk["treatment_plan"] = str(row_data.get("treatment_plan", "")).strip() or ""

        status_raw = str(row_data.get("status", "")).strip().lower()
        risk["status"] = _STATUS_TEXT.get(status_raw, "open")

        risk["review_date"] = _parse_date(row_data.get("review_date"))
        risk["strategic_objective"] = str(row_data.get("strategic_objective", "")).strip() or None
        risk["regulation_links"] = str(row_data.get("regulation_links", "")).strip() or None
        risk["risk_statement"] = str(row_data.get("risk_statement", "")).strip() or None

        rl = row_data.get("residual_likelihood")
        ri = row_data.get("residual_impact")
        if rl is not None:
            risk["residual_likelihood"] = _clamp(rl, 1, 5, None)
        if ri is not None:
            risk["residual_impact"] = _clamp(ri, 1, 5, None)

        eff = row_data.get("effectiveness_rating")
        if eff is not None:
            risk["effectiveness_rating"] = _clamp(eff, 1, 5, None)

        owner_raw = str(row_data.get("owner_name", "")).strip()
        if owner_raw:
            if owner_raw not in owner_map:
                uid, display, conf = _fuzzy_match_owner(owner_raw, users)
                owner_map[owner_raw] = {"user_id": uid, "display_name": display, "confidence": conf}
            risk["owner_name"] = owner_raw
            risk["owner_id"] = owner_map[owner_raw]["user_id"]
        else:
            risk["owner_name"] = ""
            risk["owner_id"] = None

        if row_data.get("board_visibility"):
            bv = str(row_data["board_visibility"]).strip().lower()
            risk["board_visibility"] = 1 if bv in ("1", "yes", "true", "y") else 0

        rows.append(risk)

    by_category = {}
    by_status = {}
    with_scores = 0
    for r in rows:
        cat = r.get("category", "Unknown")
        by_category[cat] = by_category.get(cat, 0) + 1
        st = r.get("status", "open")
        by_status[st] = by_status.get(st, 0) + 1
        if r.get("likelihood") and r.get("impact"):
            with_scores += 1

    col_display = {}
    for ci, field in header_map.items():
        hdr = raw_headers[ci - 1] if ci - 1 < len(raw_headers) else f"Col {ci}"
        col_display[hdr] = field

    return {
        "rows": rows,
        "summary": {
            "total": len(rows),
            "by_category": by_category,
            "by_status": by_status,
            "with_scores": with_scores,
            "without_scores": len(rows) - with_scores,
        },
        "column_map": col_display,
        "category_map": category_map,
        "owner_map": owner_map,
        "warnings": warnings[:50],
        "taxonomy_categories": taxonomy_cats or [],
        "sheet_name": ws.title,
    }


def bulk_import_risks(rows, created_by, category_overrides=None, owner_overrides=None):
    cat_over = category_overrides or {}
    own_over = owner_overrides or {}
    imported = 0
    skipped = 0
    errors = []
    for i, row in enumerate(rows):
        try:
            if not row.get("title", "").strip():
                skipped += 1
                continue
            cat_raw = row.get("category", "")
            if cat_raw in cat_over and cat_over[cat_raw]:
                row["category"] = cat_over[cat_raw]
            else:
                cm = _fuzzy_match_category(cat_raw, _get_taxonomy_categories())
                row["category"] = cm[0]
            owner_raw = row.get("owner_name", "")
            if owner_raw in own_over and own_over[owner_raw]:
                row["owner_id"] = own_over[owner_raw]
            row["created_by"] = created_by
            row.pop("owner_name", None)
            row.pop("impact_description", None)
            row.pop("review_frequency", None)
            row.pop("priority", None)
            row.pop("effectiveness_rating", None)
            create_enterprise_risk(row)
            imported += 1
        except Exception as exc:
            _log.warning("Import row %d failed: %s", i + 1, exc)
            errors.append({"row": i + 1, "title": row.get("title", ""), "error": str(exc)[:200]})
    return {"imported": imported, "skipped": skipped, "errors": errors}
