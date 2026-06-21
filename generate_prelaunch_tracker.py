"""
Generate ThemisIQ Pre-Launch Tracker as an Excel workbook.
Run: python generate_prelaunch_tracker.py
Output: ThemisIQ_PreLaunch_Tracker.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colours
NAVY = "1E3A8A"
GREEN_BG = "D1FAE5"
GREEN_FG = "059669"
RED_BG = "FEE2E2"
RED_FG = "DC2626"
ORANGE_BG = "FFF7ED"
ORANGE_FG = "EA580C"
YELLOW_BG = "FEF9C3"
YELLOW_FG = "CA8A04"
BLUE_BG = "DBEAFE"
GREY_BG = "F1F5F9"
WHITE = "FFFFFF"
LIGHT_BORDER = "D1D5DB"

header_font = Font(bold=True, color=WHITE, size=11, name="Calibri")
header_fill = PatternFill("solid", fgColor=NAVY)
section_font = Font(bold=True, color=NAVY, size=11, name="Calibri")
section_fill = PatternFill("solid", fgColor=BLUE_BG)
done_fill = PatternFill("solid", fgColor=GREEN_BG)
pending_fill = PatternFill("solid", fgColor=ORANGE_BG)
normal_font = Font(size=10, name="Calibri")
done_font = Font(size=10, name="Calibri", color=GREEN_FG, bold=True)
pending_font = Font(size=10, name="Calibri", color=ORANGE_FG, bold=True)
border = Border(
    left=Side(style="thin", color=LIGHT_BORDER),
    right=Side(style="thin", color=LIGHT_BORDER),
    top=Side(style="thin", color=LIGHT_BORDER),
    bottom=Side(style="thin", color=LIGHT_BORDER),
)
wrap = Alignment(wrap_text=True, vertical="top")


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def style_section(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    cell.fill = section_fill
    cell.border = border
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).border = border


def add_row(ws, row, data, status=None):
    for c, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = normal_font
        cell.alignment = wrap
        cell.border = border
    if status:
        status_col = len(data)
        cell = ws.cell(row=row, column=status_col)
        if status == "Done":
            cell.font = done_font
            cell.fill = done_fill
        else:
            cell.font = pending_font
            cell.fill = pending_fill


def build_workbook():
    wb = Workbook()

    # ── Sheet 1: Security Findings ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Security Findings"
    ws1.sheet_properties.tabColor = "DC2626"

    headers = ["ID", "Severity", "Category", "Title", "Description", "Phase", "Status"]
    ws1.append(headers)
    style_header(ws1, 1, len(headers))

    findings = [
        ["F-01", "CRITICAL", "Secret Exposure", "API Key Exposed in Git History",
         "Third-party API key committed to git. Key revoked, .env removed from tracking.", "Phase 2", "Done"],
        ["F-02", "HIGH", "Secret Exposure", "Email Credentials in .env.example",
         "Real Gmail address and app password in example env file. Credentials revoked.", "Phase 2", "Done"],
        ["F-03", "HIGH", "Cryptography", "API Key in URL Query Parameter",
         "Gemini API key passed as ?key= URL param. Moved to x-goog-api-key header.", "Phase 2", "Done"],
        ["F-04", "HIGH", "SSRF", "Webhook URL Validation Missing",
         "Webhook endpoint accepted internal network addresses. Added SSRF protection.", "Phase 2", "Done"],
        ["F-05", "HIGH", "XSS", "Stored XSS via Email Templates",
         "User variables not escaped in HTML email bodies. Applied html.escape().", "Phase 2", "Done"],
        ["F-06", "HIGH", "Authentication", "Missing Secure Flag on Cookies",
         "Session/CSRF cookies lacked Secure flag. Added Secure=True in production.", "Phase 2", "Done"],
        ["F-07", "MEDIUM", "Authentication", "No Rate Limiting on MFA Endpoint",
         "MFA verify had no rate limiting allowing brute-force. Extended rate limiter.", "Phase 3", "Done"],
        ["F-08", "MEDIUM", "Credentials", "Hardcoded Seed Passwords",
         "Seed script set identical passwords. Replaced with random + must_change_password.", "Phase 3", "Done"],
        ["F-09", "MEDIUM", "Cryptography", "PostgreSQL Without TLS",
         "No TLS on DB connection. Added PGSSLMODE=require via systemd override.", "Phase 3", "Done"],
        ["F-10", "MEDIUM", "Audit Trail", "Failed Logins Not Logged",
         "Auth failures not in audit log. Added log_audit() with IP and username.", "Phase 4", "Done"],
        ["F-11", "MEDIUM", "Audit Trail", "Deletion Events Missing Identifiers",
         "Deletion audit logs had only numeric IDs. Added key_prefix and URL lookup.", "Phase 4", "Done"],
        ["F-12", "MEDIUM", "Logic", "Email Password Sentinel Collision",
         "Bullet char sentinel could match real password. Replaced with '__unchanged__'.", "Phase 4", "Done"],
        ["F-13", "MEDIUM", "Data Integrity", "Bulk Import Partial Commit",
         "Per-row inserts left partial data on failure. Rewritten with atomic transaction.", "Phase 4", "Done"],
        ["F-14", "LOW", "Security Headers", "Missing COOP and CORP Headers",
         "COOP and CORP headers absent. Both added to security headers middleware.", "Phase 5", "Done"],
        ["F-15", "LOW", "Security Headers", "CSP Allows External Font CDN",
         "CSP permitted external font sources. Fonts already self-hosted. Restricted.", "Phase 5", "Done"],
        ["F-16", "LOW", "Dependencies", "Three CVEs in Dependencies",
         "jinja2, python-multipart, python-dotenv had CVEs. All upgraded.", "Phase 6", "Done"],
    ]
    for r, f in enumerate(findings, 2):
        add_row(ws1, r, f, f[-1])
        sev = f[1]
        sev_cell = ws1.cell(row=r, column=2)
        if sev == "CRITICAL":
            sev_cell.font = Font(bold=True, color=RED_FG, size=10, name="Calibri")
            sev_cell.fill = PatternFill("solid", fgColor=RED_BG)
        elif sev == "HIGH":
            sev_cell.font = Font(bold=True, color=ORANGE_FG, size=10, name="Calibri")
            sev_cell.fill = PatternFill("solid", fgColor=ORANGE_BG)
        elif sev == "MEDIUM":
            sev_cell.font = Font(bold=True, color=YELLOW_FG, size=10, name="Calibri")
            sev_cell.fill = PatternFill("solid", fgColor=YELLOW_BG)

    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 35
    ws1.column_dimensions["E"].width = 60
    ws1.column_dimensions["F"].width = 10
    ws1.column_dimensions["G"].width = 10

    # ── Sheet 2: ZAP Scan Results ───────────────────────────────────────────
    ws2 = wb.create_sheet("ZAP Scan Results")
    ws2.sheet_properties.tabColor = "059669"

    headers = ["Warning", "Count", "Risk", "Detail", "Disposition"]
    ws2.append(headers)
    style_header(ws2, 1, len(headers))

    zap_warnings = [
        ["Cache-control on robots.txt [10015]", "x1", "LOW",
         "robots.txt lacks cache-control. Nginx serves this directly.", "Accepted"],
        ["X-Content-Type-Options on robots.txt [10021]", "x1", "LOW",
         "Header set by FastAPI middleware, not propagated to Nginx static.", "Accepted"],
        ["HSTS Not Set [10035]", "x3 -> x1", "MEDIUM",
         "Main domain and sitemap fixed via Cloudflare. robots.txt remains (Nginx).", "Remediated"],
        ["Non-Storable Content [10049]", "x3", "LOW",
         "403 responses lack Cache-Control. 403s should not be cached.", "Accepted"],
        ["CSP No Fallback Directive [10055]", "x8", "LOW",
         "worker-src/manifest-src not explicit. default-src 'self' covers per spec.", "Accepted"],
        ["Timestamp Disclosure [10096]", "x2", "LOW",
         "Unix timestamps in 403 error responses. Not exploitable.", "Accepted"],
        ["Modern Web Application [10109]", "x3", "INFO",
         "Informational: site uses modern JavaScript frameworks.", "Informational"],
        ["CORP Missing on robots.txt [90004]", "x1", "LOW",
         "CORP header not propagated to Nginx-served static files.", "Accepted"],
    ]
    for r, w in enumerate(zap_warnings, 2):
        add_row(ws2, r, w)

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 55
    ws2.column_dimensions["E"].width = 14

    # ── Sheet 3: Pre-Launch Checklist ───────────────────────────────────────
    ws3 = wb.create_sheet("Pre-Launch Checklist")
    ws3.sheet_properties.tabColor = "1E3A8A"

    headers = ["Category", "Item", "Details", "Owner", "Status"]
    ws3.append(headers)
    style_header(ws3, 1, len(headers))

    row = 2

    # Section: Session and Cookies
    style_section(ws3, row, len(headers), "Session and Cookie Security")
    row += 1
    items = [
        ["Cookies", "Session cookie flags", "HttpOnly, SameSite=Strict, Secure, max_age", "", "Done"],
        ["Cookies", "CSRF cookie flags", "HttpOnly, SameSite, Secure, max_age=3600", "", "Done"],
        ["Cookies", "CSRF via HMAC", "Token derived from session via HMAC, no separate cookie needed", "", "Done"],
        ["Cookies", "Logout cookie cleanup", "delete_cookie with path, samesite, secure flags", "", "Done"],
        ["Cookies", "POST-only logout", "GET /logout redirects without destroying session", "", "Done"],
        ["Cookies", "CSRF origin-check middleware", "Blocks cross-origin POST/PUT/DELETE/PATCH", "", "Done"],
        ["Cookies", "HSTS header", "max-age=63072000, includeSubDomains, preload (prod only)", "", "Done"],
    ]
    for item in items:
        add_row(ws3, row, item, item[-1])
        row += 1

    # Section: Authentication
    style_section(ws3, row, len(headers), "Authentication and Access Control")
    row += 1
    items = [
        ["Auth", "bcrypt password hashing", "Cost factor 12", "", "Done"],
        ["Auth", "Password complexity", "8+ chars, upper, lower, digit, special", "", "Done"],
        ["Auth", "TOTP MFA", "RFC 6238, backup codes, pyotp", "", "Done"],
        ["Auth", "Session token hashed in DB", "SHA-256, prevents token DB dump attacks", "", "Done"],
        ["Auth", "Rate limiting", "5 attempts / 5 min for login, MFA, password change", "", "Done"],
        ["Auth", "DB-backed rate limiting", "PostgreSQL rate_limit_attempts table", "", "Done"],
        ["Auth", "Org isolation", "Enforced across all admin user management routes", "", "Done"],
        ["Auth", "RBAC", "Capability-based access control", "", "Done"],
        ["Auth", "Licence enforcement", "Per-tenant module licensing", "", "Done"],
    ]
    for item in items:
        add_row(ws3, row, item, item[-1])
        row += 1

    # Section: Infrastructure
    style_section(ws3, row, len(headers), "Infrastructure and Operations")
    row += 1
    items = [
        ["Infra", "PostgreSQL migration", "Migrated from SQLite to PostgreSQL", "", "Done"],
        ["Infra", "PostgreSQL TLS", "PGSSLMODE=require via systemd override", "", "Done"],
        ["Infra", "Systemd service", "Environment isolation via drop-in configuration", "", "Done"],
        ["Infra", "Nginx reverse proxy", "Port 8080, HTTPS termination via Cloudflare", "", "Done"],
        ["Infra", "Database backups", "pg_dump at 2 AM daily, 7-day retention, /project/backup_db.sh", "", "Done"],
        ["Infra", "Static asset caching", "1 year max-age, immutable for /static/", "", "Done"],
        ["Infra", "Log rotation", "Configure logrotate for app and backup logs", "", "Pending"],
        ["Infra", "Uptime monitoring", "External ping service (UptimeRobot or similar)", "", "Pending"],
        ["Infra", "Email delivery", "Verify SMTP credentials and test full flow", "", "Pending"],
        ["Infra", "Nginx rate limiting", "Add rate limiting at reverse proxy level", "", "Pending"],
        ["Infra", "Firewall rules", "Verify only ports 80, 443 exposed", "", "Pending"],
    ]
    for item in items:
        add_row(ws3, row, item, item[-1])
        row += 1

    # Section: Cloudflare
    style_section(ws3, row, len(headers), "Cloudflare Coverage")
    row += 1
    items = [
        ["Cloudflare", "SSL/TLS termination", "Full Strict mode", "", "Done"],
        ["Cloudflare", "DDoS mitigation", "Automatic, always-on", "", "Done"],
        ["Cloudflare", "HSTS enforcement", "6 months, includeSubDomains", "", "Done"],
        ["Cloudflare", "CDN caching", "Static asset caching at edge", "", "Done"],
        ["Cloudflare", "Bot management", "Basic bot detection enabled", "", "Done"],
        ["Cloudflare", "WAF rules", "Review and enable OWASP Core Rule Set", "", "Pending"],
        ["Cloudflare", "Page rules", "Configure caching for API vs static routes", "", "Pending"],
        ["Cloudflare", "Edge rate limiting", "Rate limiting rules for login/API endpoints", "", "Pending"],
    ]
    for item in items:
        add_row(ws3, row, item, item[-1])
        row += 1

    # Section: Monitoring
    style_section(ws3, row, len(headers), "Monitoring and Analytics")
    row += 1
    items = [
        ["PostHog", "JS snippet", "Integrated via meta tags in base_shell.html", "", "Done"],
        ["PostHog", "User identification", "posthog.identify() with ID, email, name, role", "", "Done"],
        ["PostHog", "Autocapture", "Pageviews, page leaves, clicks", "", "Done"],
        ["PostHog", "CSP allowlist", "PostHog domains in script-src and connect-src", "", "Done"],
        ["Sentry", "DSN configured", "SENTRY_DSN in /project/.env on VPS", "", "Done"],
        ["Sentry", "SDK integrated", "Python Sentry SDK in application", "", "Done"],
        ["Sentry", "CSP allowlist", "sentry.io domains in connect-src", "", "Done"],
        ["Sentry", "Test event verified", "Confirmed in Sentry dashboard 19 June 2026", "", "Done"],
        ["Sentry", "Alert rules", "Configure notifications for new errors", "", "Pending"],
        ["Sentry", "Release tracking", "Tag deploys with git commit hash", "", "Pending"],
    ]
    for item in items:
        add_row(ws3, row, item, item[-1])
        row += 1

    # Section: Landing Page
    style_section(ws3, row, len(headers), "Landing Page")
    row += 1
    items = [
        ["Landing", "Tailwind CSS compiled", "CDN replaced with purged 25KB CSS", "", "Done"],
        ["Landing", "Three.js deduplication", "Removed eager-loaded duplicate script", "", "Done"],
        ["Landing", "WebGL zero-dimension guard", "Guard on Spline 3D canvas init", "", "Done"],
        ["Landing", "Console errors eliminated", "260+ errors reduced to 0 (6 Spline warnings remain)", "", "Done"],
        ["Landing", "Image optimization", "Compress hero images, add WebP fallbacks", "", "Pending"],
        ["Landing", "SEO meta tags", "Verify Open Graph, Twitter Card, description", "", "Pending"],
    ]
    for item in items:
        add_row(ws3, row, item, item[-1])
        row += 1

    # Section: Deploy
    style_section(ws3, row, len(headers), "Final Deploy Steps")
    row += 1
    items = [
        ["Deploy", "Push latest code", "Cookie hardening and landing page commits", "", "Pending"],
        ["Deploy", "git pull on VPS", "Pull latest from master on /project", "", "Pending"],
        ["Deploy", "Restart service", "systemctl restart themisiq-app", "", "Pending"],
        ["Deploy", "Run test suite", "pytest on production after deploy", "", "Pending"],
        ["Deploy", "Smoke test all modules", "Login, navigate each module, verify no 500s", "", "Pending"],
    ]
    for item in items:
        add_row(ws3, row, item, item[-1])
        row += 1

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 55
    ws3.column_dimensions["D"].width = 10
    ws3.column_dimensions["E"].width = 10

    # ── Sheet 4: Post-Launch ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Post-Launch")
    ws4.sheet_properties.tabColor = "7C3AED"

    headers = ["Priority", "Item", "Details", "Status"]
    ws4.append(headers)
    style_header(ws4, 1, len(headers))

    post_launch = [
        ["P1", "Authenticated ZAP scan", "Run ZAP with valid session credentials for deeper coverage", "Pending"],
        ["P1", "Sentry alert rules", "Configure notifications for new production errors", "Pending"],
        ["P1", "Uptime monitoring", "External ping service with alerting", "Pending"],
        ["P2", "API rate limiting", "Application-level rate limiting beyond login endpoints", "Pending"],
        ["P2", "CSP report-uri", "Collect and monitor CSP violation reports", "Pending"],
        ["P2", "Cloudflare WAF rules", "Enable OWASP Core Rule Set", "Pending"],
        ["P2", "Log rotation", "Configure logrotate for app and backup logs", "Pending"],
        ["P3", "Dependency audit automation", "Scheduled pip-audit in CI pipeline", "Pending"],
        ["P3", "CI/CD pipeline", "Automated test runs on push", "Pending"],
        ["P3", "Log aggregation", "Centralized logging with rotation and search", "Pending"],
        ["P3", "DB connection pooling", "Review and optimize connection pool settings", "Pending"],
        ["P3", "Backup restoration test", "Verify dump can be restored cleanly", "Pending"],
        ["P3", "HSTS preload submission", "Submit to preload list after stable HSTS period", "Pending"],
        ["P4", "Quarterly ZAP scans", "Scheduled security re-testing", "Pending"],
        ["P4", "SOC 2 / ISO 27001 evidence", "Collect audit logs, access control evidence", "Pending"],
        ["P4", "Image optimization", "Compress landing page images, add WebP fallbacks", "Pending"],
    ]
    for r, item in enumerate(post_launch, 2):
        add_row(ws4, r, item, item[-1])

    ws4.column_dimensions["A"].width = 10
    ws4.column_dimensions["B"].width = 30
    ws4.column_dimensions["C"].width = 55
    ws4.column_dimensions["D"].width = 10

    # ── Sheet 5: Bug Fixes ──────────────────────────────────────────────────
    ws5 = wb.create_sheet("Bug Fixes")
    ws5.sheet_properties.tabColor = "059669"

    headers = ["Bug", "Module", "Root Cause", "Fix", "Status"]
    ws5.append(headers)
    style_header(ws5, 1, len(headers))

    bugs = [
        ["GRID dashboard 500", "GRID", "PostgreSQL strict GROUP BY",
         "Added MAX() wrappers and full GROUP BY column list", "Done"],
        ["PostHog JS syntax error", "Platform", "Jinja2 auto-escaping in script blocks",
         "Moved PostHog config to meta tags, read from DOM", "Done"],
        ["MFA disabling on setup visit", "Auth", "start_enrollment called on every GET /mfa/setup",
         "Conditional: only call if not already enrolled", "Done"],
        ["Forced password change CSRF", "Auth", "CSRF token derived from cookie, not session",
         "HMAC-derived CSRF token bound to session", "Done"],
        ["Org deletion FK violation", "Admin", "audit_log.org_id and api_keys blocked DROP",
         "Null audit_log.org_id and delete api_keys before DROP", "Done"],
        ["User deletion FK violation", "Admin", "Multiple FK references to user not cleaned up",
         "Clean all FK-referenced rows before DELETE", "Done"],
        ["New org user CSRF failure", "Admin", "CSRF cookie not set for forced password change",
         "Session-based CSRF derivation", "Done"],
        ["Mobile sidebar positioning", "UI", "Module sidebar not beside icon nav on mobile",
         "Fixed CSS positioning for mobile drawer", "Done"],
        ["Starlette TemplateResponse API", "Admin", "New Starlette requires request as first arg",
         "Updated all TemplateResponse calls", "Done"],
        ["New User modal not opening", "Admin", "Wrong CSS class name in template",
         "Fixed class name typo", "Done"],
    ]
    for r, b in enumerate(bugs, 2):
        add_row(ws5, r, b, b[-1])

    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 12
    ws5.column_dimensions["C"].width = 40
    ws5.column_dimensions["D"].width = 50
    ws5.column_dimensions["E"].width = 10

    # ── Save ────────────────────────────────────────────────────────────────
    path = "ThemisIQ_PreLaunch_Tracker.xlsx"
    wb.save(path)
    print(f"Saved: {path}")


if __name__ == "__main__":
    build_workbook()
