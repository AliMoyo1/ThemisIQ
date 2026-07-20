"""
One-off extractor: reads the AI controls taxonomy workbook and writes
seeds/ai_controls_seed.json. Run manually on a dev machine that has the
source file; the app itself never reads the xlsx, only the committed
JSON output of this script.

Usage: python scripts/extract_ai_controls.py [path-to-Taxonomy.xlsx]
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

_ALLOWED_PILLARS = {"People", "Process", "Systems", "Technology", "Tools"}
_PILLAR_ALIASES = {"System": "Systems"}


def _normalize_pillar(raw: str) -> str:
    """The sheet mixes 'Process/Systems', 'Process/Tools', trailing
    whitespace, and 'System' vs 'Systems'. Take the first slash-segment,
    strip, then apply the singular->plural alias. Anything still outside
    the 5-value set is a data problem in the source, not a case to
    silently swallow -- raise so it gets fixed at extraction time."""
    first = (raw or "").split("/")[0].strip()
    first = _PILLAR_ALIASES.get(first, first)
    if first not in _ALLOWED_PILLARS:
        raise ValueError(f"Unrecognized pillar value: {raw!r} (normalized to {first!r})")
    return first


def _title_from_control(text: str) -> str:
    text = (text or "").strip()
    match = re.match(r"^(.*?[.!?])(\s|$)", text)
    if match:
        candidate = match.group(1).strip()
        if len(candidate) <= 120:
            return candidate
    return text[:120].strip()


def extract(xlsx_path: str) -> list:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ref, control, pillar_raw = row[1], row[2], row[3]
        if not ref or not control:
            continue
        rows.append({
            "ref": str(ref).strip(),
            "title": _title_from_control(control),
            "description": str(control).strip(),
            "pillar": _normalize_pillar(pillar_raw),
        })
    return rows


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else str(
        Path.home() / "Downloads" / "Taxonomy.xlsx"
    )
    rows = extract(xlsx_path)
    out_path = Path(__file__).parent.parent / "seeds" / "ai_controls_seed.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Extracted {len(rows)} controls -> {out_path}")
    pillars = {}
    for r in rows:
        pillars[r["pillar"]] = pillars.get(r["pillar"], 0) + 1
    for p, c in sorted(pillars.items()):
        print(f"  {p}: {c}")


if __name__ == "__main__":
    main()
