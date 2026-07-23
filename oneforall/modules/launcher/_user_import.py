"""
PLAN-33 Phase 3: bulk user creation from an Excel sheet, SBU/parent-company
aware. Mirrors the two-phase preview/commit shape already used by ERM's risk
register importer (modules/erm/data_service.py: parse_risk_register_excel /
bulk_import_risks) -- same fuzzy-matching-via-difflib approach, same
{rows, summary, warnings, ...} preview payload shape, same per-row
try/except commit so one bad row doesn't kill the batch.
"""
import io
import re
import secrets
import string
from difflib import get_close_matches

from core.auth import hash_password
from core.rbac import ALL_ROLES, EMPLOYEE, SUPER_ADMIN
from database import get_db, insert_returning_id

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s.]+\.[^@\s]+$")

_HEADER_MAP = {
    "fullname": "full_name", "name": "full_name", "employeename": "full_name",
    "email": "email", "emailaddress": "email",
    "username": "username", "userid": "username",
    "roles": "roles", "role": "roles",
    "businessunit": "business_unit", "businessunitsbu": "business_unit",
    "sbu": "business_unit", "unit": "business_unit", "department": "business_unit",
    "organization": "organization", "organisation": "organization",
    "company": "organization", "tenant": "organization",
}


def _norm_header(s):
    # Strip parenthetical hints ("Username (optional)" -> "username") before
    # normalizing -- these hints appear in the template this module itself
    # generates, so without this the template's own headers would fail to
    # match _HEADER_MAP.
    s = re.sub(r"\(.*?\)", "", str(s))
    return re.sub(r"[^a-z0-9]", "", s.lower().strip())


def _slugify_username(full_name: str) -> str:
    base = re.sub(r"[^a-z0-9.]", "", full_name.lower().replace(" ", "."))
    return base[:28] or "user"


def _fuzzy_match_bu(raw, bu_names):
    if not raw or not str(raw).strip():
        return (None, "empty")
    clean = str(raw).strip()
    lower = clean.lower()
    for name in bu_names:
        if name.lower() == lower:
            return (name, "exact")
    matches = get_close_matches(clean, bu_names, n=1, cutoff=0.5)
    if matches:
        return (matches[0], "fuzzy")
    return (clean, "unmapped")


def parse_users_excel(file_bytes, is_super: bool, caller_org_id):
    """Parse an uploaded Excel sheet of users into a preview payload.

    Columns (fuzzy-matched by header name): Full Name, Email,
    Username (optional), Roles (comma-separated), Business Unit (SBU),
    Organization (super-admin only -- ignored for org-scoped admins, whose
    import is always forced into their own org).
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    db = get_db()
    try:
        bu_rows = db.execute(
            "SELECT id, name FROM business_units WHERE is_active=1 ORDER BY name"
        ).fetchall()
        business_units = [{"id": b["id"], "name": b["name"]} for b in bu_rows]
        existing = db.execute("SELECT username, email FROM users").fetchall()
        existing_usernames = {r["username"].lower() for r in existing}
        existing_emails = {r["email"].lower() for r in existing}
        org_rows = []
        if is_super:
            org_rows = db.execute(
                "SELECT id, name FROM organizations WHERE status='active' ORDER BY name"
            ).fetchall()
    finally:
        db.close()

    bu_names = [b["name"] for b in business_units]
    bu_by_name = {b["name"].lower(): b["id"] for b in business_units}
    org_by_name = {o["name"].lower(): o["id"] for o in org_rows}

    header_row = None
    header_map = {}
    raw_headers = []
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        cells = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        normed = [_norm_header(c) for c in cells if c is not None]
        matched = sum(1 for n in normed if n in _HEADER_MAP)
        if matched >= 2:
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row_idx, col_idx).value
                if val is None:
                    continue
                field = _HEADER_MAP.get(_norm_header(val))
                if field:
                    header_map[col_idx] = field
            raw_headers = [str(ws.cell(row_idx, c).value or "").strip() for c in range(1, ws.max_column + 1)]
            header_row = row_idx
            break

    if header_row is None:
        return {"error": "Could not detect a header row. Need at least Full Name and Email columns."}

    name_col = next((ci for ci, f in header_map.items() if f == "full_name"), None)
    if name_col is None:
        return {"error": "Could not find a 'Full Name' column."}

    rows = []
    warnings = []
    bu_map = {}
    org_map = {}
    seen_emails_in_sheet = set()
    seen_usernames_in_sheet = set()

    for row_idx in range(header_row + 1, ws.max_row + 1):
        name_val = ws.cell(row_idx, name_col).value
        full_name = str(name_val or "").strip()
        if not full_name:
            continue

        row_data = {}
        for col_idx, field in header_map.items():
            val = ws.cell(row_idx, col_idx).value
            if val is not None:
                row_data[field] = val

        user = {"row": row_idx, "full_name": full_name}

        email = str(row_data.get("email", "")).strip().lower()
        if not email or not _EMAIL_RE.match(email):
            warnings.append(f"Row {row_idx}: missing or invalid email for '{full_name}' -- skipped")
            user["email"] = email
            user["error"] = "invalid_email"
            rows.append(user)
            continue
        user["email"] = email

        dup = email in existing_emails or email in seen_emails_in_sheet
        seen_emails_in_sheet.add(email)
        user["duplicate_email"] = dup
        if dup:
            warnings.append(f"Row {row_idx}: email '{email}' already exists -- will be skipped on import")

        username = str(row_data.get("username", "")).strip().lower()
        if not username:
            username = _slugify_username(full_name)
        dup_uname = username in existing_usernames or username in seen_usernames_in_sheet
        if dup_uname and not dup:
            # Auto-suffix rather than warn/skip -- username collisions are
            # cosmetic (email is the real identity key) unlike email dupes.
            suffix = 1
            candidate = f"{username}{suffix}"
            while candidate in existing_usernames or candidate in seen_usernames_in_sheet:
                suffix += 1
                candidate = f"{username}{suffix}"
            username = candidate
        seen_usernames_in_sheet.add(username)
        user["username"] = username

        roles_raw = str(row_data.get("roles", "")).strip()
        role_keys = [r.strip().lower().replace(" ", "_") for r in roles_raw.split(",") if r.strip()]
        valid_roles = [r for r in role_keys if r in ALL_ROLES]
        invalid_roles = [r for r in role_keys if r not in ALL_ROLES]
        if invalid_roles:
            warnings.append(f"Row {row_idx}: unrecognized role(s) {invalid_roles} ignored")
        if not valid_roles:
            valid_roles = [EMPLOYEE]
        if SUPER_ADMIN in valid_roles and not is_super:
            valid_roles.remove(SUPER_ADMIN)
            warnings.append(f"Row {row_idx}: Super Administrator role can only be granted by a super admin -- dropped")
        user["roles"] = valid_roles

        bu_raw = str(row_data.get("business_unit", "")).strip()
        if bu_raw:
            if bu_raw not in bu_map:
                matched, conf = _fuzzy_match_bu(bu_raw, bu_names)
                bu_map[bu_raw] = {
                    "matched": matched,
                    "bu_id": bu_by_name.get((matched or "").lower()),
                    "confidence": conf,
                }
            user["business_unit_raw"] = bu_raw
            user["business_unit_id"] = bu_map[bu_raw]["bu_id"]
        else:
            user["business_unit_raw"] = ""
            user["business_unit_id"] = None

        org_raw = str(row_data.get("organization", "")).strip()
        if is_super and org_raw:
            if org_raw not in org_map:
                m = get_close_matches(org_raw, [o["name"] for o in org_rows], n=1, cutoff=0.5)
                matched_org = m[0] if m else None
                org_map[org_raw] = {
                    "matched": matched_org,
                    "org_id": org_by_name.get((matched_org or "").lower()),
                }
                if not matched_org:
                    warnings.append(f"Row {row_idx}: organization '{org_raw}' not found -- will use your own org")
            user["organization_raw"] = org_raw
            user["organization_id"] = org_map[org_raw]["org_id"]
        else:
            user["organization_raw"] = ""
            user["organization_id"] = caller_org_id

        rows.append(user)

    valid_count = sum(1 for r in rows if not r.get("error") and not r.get("duplicate_email"))

    col_display = {}
    for ci, field in header_map.items():
        hdr = raw_headers[ci - 1] if ci - 1 < len(raw_headers) else f"Col {ci}"
        col_display[hdr] = field

    return {
        "rows": rows,
        "summary": {
            "total": len(rows),
            "valid": valid_count,
            "duplicates": sum(1 for r in rows if r.get("duplicate_email")),
            "invalid": sum(1 for r in rows if r.get("error")),
            "with_bu": sum(1 for r in rows if r.get("business_unit_id")),
        },
        "column_map": col_display,
        "bu_map": bu_map,
        "org_map": org_map,
        "warnings": warnings[:50],
        "business_units": business_units,
        "organizations": [{"id": o["id"], "name": o["name"]} for o in org_rows] if is_super else [],
        "sheet_name": ws.title,
    }


def _gen_temp_password(length: int = 12) -> str:
    alpha = string.ascii_letters + string.digits + "!@#$%"
    return "".join(secrets.choice(alpha) for _ in range(length))


def bulk_create_users(rows, admin: dict, bu_overrides=None, org_overrides=None):
    """Create each valid row as a real user. Org-scoped admins are always
    forced into their own org regardless of any Organization column/override
    (PLAN-30 org-scoping). One try/except per row so a single bad row doesn't
    abort the whole batch. Returns created count + a one-time credentials
    handoff list (temp passwords are never persisted anywhere else)."""
    bu_over = bu_overrides or {}
    org_over = org_overrides or {}
    is_super = bool(admin.get("is_super_admin"))
    caller_org_id = admin.get("org_id")

    created = 0
    skipped = 0
    errors = []
    credentials = []

    db = get_db()
    try:
        for row in rows:
            try:
                if row.get("error") or row.get("duplicate_email"):
                    skipped += 1
                    continue
                email = row.get("email", "")
                username = row.get("username", "")
                full_name = row.get("full_name", "")
                if not (email and username and full_name):
                    skipped += 1
                    continue

                dup = db.execute(
                    "SELECT id FROM users WHERE username=%s OR email=%s",
                    (username, email),
                ).fetchone()
                if dup:
                    skipped += 1
                    continue

                if is_super:
                    target_org_id = org_over.get(row.get("organization_raw", "")) or row.get("organization_id") or caller_org_id
                else:
                    target_org_id = caller_org_id

                bu_raw = row.get("business_unit_raw", "")
                business_unit_id = bu_over.get(bu_raw) if bu_raw in bu_over else row.get("business_unit_id")

                temp_pw = _gen_temp_password()
                initials = "".join(w[0].upper() for w in full_name.split()[:2]) or "?"
                new_id = insert_returning_id(db, """
                    INSERT INTO users
                    (username, email, full_name, password_hash, is_active,
                     must_change_password, avatar_initials, org_id, business_unit_id)
                    VALUES (%s, %s, %s, %s, 1, 1, %s, %s, %s)
                """, (username, email, full_name, hash_password(temp_pw), initials,
                      target_org_id, business_unit_id))

                role_keys = row.get("roles") or [EMPLOYEE]
                for rk in role_keys:
                    if rk == SUPER_ADMIN and not is_super:
                        continue
                    db.execute(
                        "INSERT INTO user_roles (user_id, role_key, granted_by) "
                        "VALUES (%s, %s, %s) ON CONFLICT DO NOTHING",
                        (new_id, rk, int(admin["id"])),
                    )
                db.commit()
                created += 1
                credentials.append({
                    "username": username, "full_name": full_name,
                    "email": email, "temp_password": temp_pw,
                })
            except Exception as exc:
                db.rollback()
                errors.append({"row": row.get("row"), "full_name": row.get("full_name", ""), "error": str(exc)[:200]})
    finally:
        db.close()

    return {"created": created, "skipped": skipped, "errors": errors, "credentials": credentials}
